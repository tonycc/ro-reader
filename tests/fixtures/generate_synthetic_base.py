"""合成 base 文件生成脚本。

用途：
- 人工 QA：用 Excel 打开生成的文件，肉眼检查布局合理
- 端到端 CLI 验证：不依赖单元测试 fixture，直接跑 ro-generate
- 真实 base 文件入库决策待与团队确认；在确认前用合成数据替代

覆盖场景（AGENTS.md "测试 fixture"）：
- combo / rod / reel 三类
- `SHIP QTY` 实际出货数量
- 多 INV#（触发 needs_input invoice_no，Phase 2 实现）
- 一行缺 SAP 的失败案例（按需取用）

用法：
    uv run python tests/fixtures/generate_synthetic_base.py
    # → 生成 tests/fixtures/synthetic_base.xlsx
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

DATA_BASE_HEADER = [
    "SAP",
    "Material Description",
    "Category",
    "GS MODEL",
    "MOQ",
    "FOB LT",
    "品牌",
    "RFID",
    "包装",
    "inner case value",
    "round value",
    "N/W",
    "G/W",
    "L",
    "W",
    "H",
    "CBM",
    "主件编号",
]

PO_RECORD_HEADER = [
    "SHIP TO",
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "BRAND",
    "FINALQTY",
    "SK/YM USD FOB",
    "GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "FACTORY DOC NO.",
    "ORDER DATE (EMAIL)",
    "PO DELIVERY DATE",
    "CTNS",
    "N/W",
    "G/W",
    "TOTAL CBM",
    "外箱",
    "SHIP QTY",
    "BALANCE QTY",
]

# ————————————————————————————————————————
# 产品主数据
# ————————————————————————————————————————

DATA_BASE_ROWS: list[dict[str, Any]] = [
    {
        "SAP": "21-44640",
        "Material Description": "CB2500.B2 Combo",
        "Category": 1,  # combo
        "GS MODEL": "Q1",
        "MOQ": 100,
        "FOB LT": 60,
        "品牌": "Quantum",
        "RFID": "Y",
        "包装": "carton",
        "round value": 24,
        "N/W": Decimal("12.5"),
        "G/W": Decimal("13.8"),
        "L": Decimal("60"),
        "W": Decimal("40"),
        "H": Decimal("30"),
        "CBM": Decimal("0.072"),
    },
    {
        "SAP": "21-44641",
        "Material Description": "CR3000.B2 Rod",
        "Category": 2,  # rod
        "GS MODEL": "Q2-ROD",
        "MOQ": 200,
        "FOB LT": 45,
        "品牌": "Quantum",
        "包装": "bulk",
        "round value": 12,
        "N/W": Decimal("8.0"),
        "G/W": Decimal("9.2"),
        "L": Decimal("180"),
        "W": Decimal("20"),
        "H": Decimal("20"),
        "CBM": Decimal("0.072"),
    },
    {
        "SAP": "21-44642",
        "Material Description": "RL4000.B2 Reel",
        "Category": 3,  # reel
        "GS MODEL": "Q3-REEL",
        "MOQ": 50,
        "FOB LT": 40,
        "品牌": "Lew's",
        "包装": "box",
        "round value": 36,
        "N/W": Decimal("4.5"),
        "G/W": Decimal("5.2"),
        "L": Decimal("40"),
        "W": Decimal("30"),
        "H": Decimal("20"),
        "CBM": Decimal("0.024"),
    },
]

# ————————————————————————————————————————
# 订单数据
# ————————————————————————————————————————

PO_RECORD_ROWS: list[dict[str, Any]] = [
    # PO 4500030844：3 行（combo + rod + reel），均有实际出货数量
    {
        "SHIP TO": "EMAX HQ",
        "PO NO.": "4500030844",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2 Combo",
        "BRAND": "Quantum",
        "FINALQTY": 240,
        "SK/YM USD FOB": Decimal("28.0"),
        "GS PTE FOB": Decimal("32.8"),
        "EMAX PTE": Decimal("38.0"),
        "INV#": "INV-2601-001",
        "FACTORY DOC NO.": "FDOC-2601-A",
        "外箱": 24,
        "CTNS": 10,
        "TOTAL CBM": Decimal("0.72"),
        "SHIP QTY": 100,
        "BALANCE QTY": 140,
    },
    {
        "SHIP TO": "EMAX HQ",
        "PO NO.": "4500030844",
        "ITEM LINE#": "20",
        "SAP Number": "21-44641",
        "DESCRIPTION": "CR3000.B2 Rod",
        "BRAND": "Quantum",
        "FINALQTY": 120,
        "SK/YM USD FOB": Decimal("18.5"),
        "GS PTE FOB": Decimal("22.0"),
        "EMAX PTE": Decimal("26.0"),
        "INV#": "INV-2601-001",
        "FACTORY DOC NO.": "FDOC-2601-A",
        "外箱": 12,
        "CTNS": 10,
        "TOTAL CBM": Decimal("0.72"),
        "SHIP QTY": 60,
        "BALANCE QTY": 60,
    },
    {
        "SHIP TO": "EMAX HQ",
        "PO NO.": "4500030844",
        "ITEM LINE#": "30",
        "SAP Number": "21-44642",
        "DESCRIPTION": "RL4000.B2 Reel",
        "BRAND": "Lew's",
        "FINALQTY": 360,
        "SK/YM USD FOB": Decimal("12.0"),
        "GS PTE FOB": Decimal("15.5"),
        "EMAX PTE": Decimal("18.5"),
        "INV#": "INV-2601-001",
        "FACTORY DOC NO.": "FDOC-2601-A",
        "外箱": 36,
        "CTNS": 10,
        "TOTAL CBM": Decimal("0.24"),
        "SHIP QTY": 360,
        "BALANCE QTY": 0,
    },
    # PO 4500099999：1 行，简单成功路径
    {
        "SHIP TO": "EMAX HQ",
        "PO NO.": "4500099999",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2 Combo",
        "BRAND": "Quantum",
        "FINALQTY": 100,
        "SK/YM USD FOB": Decimal("28.0"),
        "GS PTE FOB": Decimal("32.8"),
        "EMAX PTE": Decimal("38.0"),
        "INV#": "INV-2603-001",
        "FACTORY DOC NO.": "FDOC-2603",
        "外箱": 24,
        "CTNS": 5,  # 100/24=4.17，留作回退测试时用整 5
        "TOTAL CBM": Decimal("0.36"),
        "SHIP QTY": 100,
        "BALANCE QTY": 0,
    },
    # PO 4500088888：包含一行缺 SAP 的失败案例（按需取用）
    {
        "SHIP TO": "EMAX HQ",
        "PO NO.": "4500088888",
        "ITEM LINE#": "10",
        "SAP Number": None,  # ← 阻断点
        "DESCRIPTION": "未指定 SAP",
        "BRAND": "Quantum",
        "FINALQTY": 50,
        "GS PTE FOB": Decimal("32.8"),
        "INV#": "INV-2604-001",
        "FACTORY DOC NO.": "FDOC-2604",
        "外箱": 24,
        "SHIP QTY": 50,
        "BALANCE QTY": 0,
    },
]

# ————————————————————————————————————————
# 客户PO数据
# ————————————————————————————————————————

CUSTOMER_PO_HEADER = [
    "Purchasing Document",
    "Item",
    "Purchasing Doc. Type",
    "Item Category",
    "Purch. Doc. Category",
    "Purchasing Group",
    "Document Date",
    "Vendor/supplying plant",
    "Material",
    "Old Material #",
    "Short Text",
    "Plant",
    "Deletion Indicator",
    "Order Quantity",
    "Order Unit",
    "Net price",
    "Currency",
    "Price Unit",
    "Still to be delivered (qty)",
    "Confirm. Category",
    "Delivery Date",
    "Ship Date",
    "order date",
    "ship to",
    "manufacturer",
    "final destination",
]

CUSTOMER_PO_ROWS: list[dict[str, Any]] = [
    # 与 PO 4500030844 / 4500099999 关联的客户PO行
    {
        "Purchasing Document": "4500030844",
        "Item": "00010",
        "Material": "21-44640",
        "Short Text": "CB2500.B2 Combo",
        "Order Quantity": 240,
        "Order Unit": "PCS",
        "Net price": 38.0,
        "Currency": "USD",
        "Delivery Date": "2026-05-20",
        "ship to": "Rather Outdoors Corporation",
        "manufacturer": "E MAX SPORT PTE. LTD.",
        "final destination": "USA",
    },
    {
        "Purchasing Document": "4500030844",
        "Item": "00020",
        "Material": "21-44641",
        "Short Text": "CR3000.B2 Rod",
        "Order Quantity": 120,
        "Order Unit": "PCS",
        "Net price": 26.0,
        "Currency": "USD",
        "Delivery Date": "2026-05-20",
        "ship to": "Rather Outdoors Corporation",
        "manufacturer": "E MAX SPORT PTE. LTD.",
        "final destination": "USA",
    },
    {
        "Purchasing Document": "4500099999",
        "Item": "00010",
        "Material": "21-44640",
        "Short Text": "CB2500.B2 Combo",
        "Order Quantity": 100,
        "Order Unit": "PCS",
        "Net price": 38.0,
        "Currency": "USD",
        "Delivery Date": "2026-03-20",
        "ship to": "EMAX HQ",
        "manufacturer": "E MAX SPORT PTE. LTD.",
        "final destination": "USA",
    },
]


# ————————————————————————————————————————
# 写盘
# ————————————————————————————————————————


def build_workbook() -> Workbook:
    """构造一个 Workbook 对象（不写盘，便于测试调用）。"""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(ws_db, DATA_BASE_HEADER, DATA_BASE_ROWS)

    ws_po = wb.create_sheet("PO record")
    _write_sheet(ws_po, PO_RECORD_HEADER, PO_RECORD_ROWS)

    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(ws_cp, CUSTOMER_PO_HEADER, CUSTOMER_PO_ROWS, header_row=1, first_data_row=2)

    return wb


def write_synthetic_base(output_path: Path) -> Path:
    """把合成 base 写到指定路径，返回绝对路径。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(output_path)
    return output_path.resolve()


def _write_sheet(
    ws: Worksheet,
    headers: list[str],
    rows: list[dict[str, Any]],
    header_row: int = 4,
    first_data_row: int = 5,
) -> None:
    """写入 sheet 表头和数据。默认表头第 4 行 + 数据第 5 行（与产品方案 §9 一致）。"""
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            if value is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=value)


# ————————————————————————————————————————
# 入口
# ————————————————————————————————————————


DEFAULT_OUTPUT = Path(__file__).resolve().parent / "synthetic_base.xlsx"


def main() -> None:
    output = write_synthetic_base(DEFAULT_OUTPUT)
    print(f"已生成合成 base：{output}")
    print(
        "覆盖：\n"
        "  - 3 个产品（combo/rod/reel）\n"
        "  - PO 4500030844：3 行带 SHIP QTY（出货行筛选测试）\n"
        "  - PO 4500099999：1 行（简单成功路径）\n"
        "  - PO 4500088888：1 行缺 SAP（阻断错误测试）"
    )


if __name__ == "__main__":
    main()
