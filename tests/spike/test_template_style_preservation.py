"""Spike A: 模板样式保留验证。

按 implementation-guide §4.1 流程：
  1. 加载 GS Invoice 模板
  2. 写入表头字段（INVOICE#、DATE）
  3. 在 start_row 写 2 行数据
  4. 在 start_row + 2 处 insert_rows(1) + 复制 start_row + 1 的样式
  5. 写入第 3 条数据
  6. 写入合计单元格
  7. 保存为新文件
  8. 重新打开，断言所有样式属性

验收（implementation-guide §4.1）：
  - 文件能被 openpyxl 重新打开，无 corruption 警告
  - 合并单元格区域集合相等
  - 列宽 ws.column_dimensions[col].width 完全一致
  - 模板原有行高保持不变
  - 新插入行的行高来自样板行
  - 公式在插入行中正确写入

LibreOffice PDF 视觉对比环节因本机未安装 LibreOffice，跳过。
"""

from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_PATH = REPO_ROOT / "templates" / "gs" / "invoice.xlsx"

# Phase 0 Spike A 已完成（见 CLAUDE.md）。其样板模板 templates/gs/invoice.xlsx
# 在模板改名为 invoice&pl.xlsx 后已不存在，且断言行号（18/19/27）绑定旧版式。
# insert_rows / 行高平移 / 样式复制的回归覆盖现由 test_renderer.py 承担
# （test_totals_shifted_by_insertion_count / test_inserted_rows_have_styles 等）。
# 缺模板时整体跳过，避免 CI 收集报错；若日后恢复该模板则自动重新参与。
pytestmark = pytest.mark.skipif(
    not TEMPLATE_PATH.exists(),
    reason="Phase 0 spike 样板模板 templates/gs/invoice.xlsx 已随模板改名移除；行为已由 test_renderer.py 覆盖",
)

# 业务认知（来自 §6.7 题目检查）：
#   表头行 13-14
#   样板数据行 18-24
#   合计行 27
START_ROW = 18  # 第一条数据写入位置
TEMPLATE_STYLE_ROW = 19  # 复制样式的样板行
TOTAL_ROW = 27


def template_sheet(wb):
    """Return the invoice template sheet under current template naming."""
    return wb["INV"] if "INV" in wb.sheetnames else wb.active


def copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    """把 src_row 的每个单元格样式复制到 dst_row 的同列单元格。

    openpyxl 的 insert_rows 默认不复制样式，需要手动处理。
    """
    src_height = ws.row_dimensions[src_row].height
    if src_height is not None:
        ws.row_dimensions[dst_row].height = src_height
    for col_idx in range(1, max_col + 1):
        src_cell: Cell = ws.cell(row=src_row, column=col_idx)
        dst_cell: Cell = ws.cell(row=dst_row, column=col_idx)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)


def render_invoice(template: Path, output: Path) -> None:
    """用样板模板装配一份 Invoice 文件。

    样本输入：3 行数据，跨两个 PO，触发"插入行"路径。
    """
    wb = load_workbook(template)
    ws = template_sheet(wb)

    # —— 写表头 ——
    ws["H6"] = "INV-SPIKE-001"  # INVOICE #
    ws["H7"] = date(2026, 1, 31)  # DATE

    # —— 数据行 ——
    rows = [
        ("4500030844", "CB2500.B2", "21-44640", 32.8, 100),
        ("4500030844", "CB3000.B2", "21-44641", 33.1, 200),
        ("4500030844", "CB4000.B2", "21-44642", 34.4, 80),
    ]

    # 先把样板行 START_ROW、START_ROW + 1 用真实数据覆盖
    write_data_row(ws, START_ROW, rows[0])
    write_data_row(ws, START_ROW + 1, rows[1])

    # —— 插入第 3 行（验证 insert_rows + 样式复制）——
    insert_at = START_ROW + 2
    insert_styled_row(ws, insert_at, style_src=TEMPLATE_STYLE_ROW)
    write_data_row(ws, insert_at, rows[2])

    wb.save(output)


def insert_styled_row(ws: Worksheet, insert_at: int, style_src: int) -> None:
    """在 insert_at 处插入一行，并执行 openpyxl 不会自动做的事：

    - 平移 insert_at 之后所有 row_dimensions 一格（openpyxl 只移单元格内容，不移 row_dimensions）
    - 把 style_src 行的样式复制到新插入的行
    """
    # 1. 先手动把 row_dimensions 下移一格（必须在 insert_rows 之前做，否则会和被移位的内容错位）
    #    倒序处理，避免覆盖
    existing_rows = sorted(
        (r for r in ws.row_dimensions if r >= insert_at),
        reverse=True,
    )
    for orig_row in existing_rows:
        src_dim = ws.row_dimensions[orig_row]
        new_dim = ws.row_dimensions[orig_row + 1]
        new_dim.height = src_dim.height
        new_dim.hidden = src_dim.hidden
        new_dim.outlineLevel = src_dim.outlineLevel
    # 清掉原 insert_at 的行高，让样板复制接管
    if insert_at in ws.row_dimensions:
        ws.row_dimensions[insert_at].height = None

    # 2. 让 openpyxl 处理单元格内容、公式、合并区域的下移
    ws.insert_rows(insert_at)

    # 3. 把样板行样式复制到新行
    copy_row_style(ws, src_row=style_src, dst_row=insert_at, max_col=ws.max_column)


def write_data_row(ws: Worksheet, row: int, data: tuple[str, str, str, float, int]) -> None:
    po, model, sap, unit_price, qty = data
    ws.cell(row=row, column=2).value = po  # B = PO#
    ws.cell(row=row, column=3).value = model  # C = MODEL
    ws.cell(row=row, column=4).value = sap  # D = SAP
    ws.cell(row=row, column=5).value = unit_price  # E = Unit price
    ws.cell(row=row, column=6).value = qty  # F = Qty
    ws.cell(row=row, column=7).value = "PCS"  # G = Unit
    ws.cell(row=row, column=8).value = f"=E{row}*F{row}"  # H = formula amount


@pytest.fixture
def baseline() -> dict:
    """加载原模板，提取需要保留的不变量。"""
    wb = load_workbook(TEMPLATE_PATH)
    ws = template_sheet(wb)
    return {
        "merged": {str(r) for r in ws.merged_cells.ranges},
        "col_widths": {col: dim.width for col, dim in ws.column_dimensions.items()},
        "row_heights": {row: dim.height for row, dim in ws.row_dimensions.items()},
        "max_col": ws.max_column,
        "print_area": ws.print_area,
        "template_style_row_height": ws.row_dimensions[TEMPLATE_STYLE_ROW].height,
    }


@pytest.fixture
def rendered(tmp_path: Path) -> Path:
    output = tmp_path / "spike_invoice.xlsx"
    render_invoice(TEMPLATE_PATH, output)
    return output


def test_file_reopens_cleanly(rendered: Path) -> None:
    """断言：文件能被 openpyxl 重新打开。"""
    wb = load_workbook(rendered)
    assert template_sheet(wb).title in wb.sheetnames


def test_merged_cells_preserved(rendered: Path, baseline: dict) -> None:
    """断言：原合并单元格区域全部保留。

    注意：因为我们在 row 20 插入了一行，原来在 row 20 之后的合并区域应该被
    openpyxl 自动下移；row 20 之前的合并保持不变。
    模板原有合并都在 row 1-4，不受插入影响，期望集合相等。
    """
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    rendered_merged = {str(r) for r in ws.merged_cells.ranges}
    assert rendered_merged == baseline["merged"], (
        f"合并单元格区域变化\n  before: {baseline['merged']}\n  after:  {rendered_merged}"
    )


def test_column_widths_preserved(rendered: Path, baseline: dict) -> None:
    """断言：每列的列宽保持一致。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    for col, expected_width in baseline["col_widths"].items():
        actual_width = ws.column_dimensions[col].width
        assert actual_width == expected_width, (
            f"列 {col} 宽度变化：{expected_width} → {actual_width}"
        )


def test_pre_insert_row_heights_preserved(rendered: Path, baseline: dict) -> None:
    """断言：插入位置之前的行高保持不变。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    insert_at = START_ROW + 2
    for row, expected_height in baseline["row_heights"].items():
        if row >= insert_at:
            continue  # 插入行之后的行高会被 openpyxl 下移，分开检查
        actual_height = ws.row_dimensions[row].height
        assert actual_height == expected_height, (
            f"row {row} 行高变化：{expected_height} → {actual_height}"
        )


def test_inserted_row_has_template_style_height(rendered: Path, baseline: dict) -> None:
    """断言：新插入行的行高来自 TEMPLATE_STYLE_ROW。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    insert_at = START_ROW + 2
    actual_height = ws.row_dimensions[insert_at].height
    expected_height = baseline["template_style_row_height"]
    assert actual_height == expected_height, (
        f"新插入行行高 {actual_height} 不等于样板行行高 {expected_height}"
    )


def test_inserted_row_has_styled_cells(rendered: Path) -> None:
    """断言：新插入行的单元格有 font / border 等样式（不全是默认值）。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    insert_at = START_ROW + 2
    has_styled = False
    for col_idx in range(2, 9):
        cell = ws.cell(row=insert_at, column=col_idx)
        if cell.has_style:
            has_styled = True
            break
    assert has_styled, "新插入行所有单元格都没有样式，说明样式复制失败"


def test_data_written_correctly(rendered: Path) -> None:
    """断言：3 行数据正确写入。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    insert_at = START_ROW + 2
    expected = [
        (START_ROW, "4500030844", "CB2500.B2", "21-44640", 32.8, 100),
        (START_ROW + 1, "4500030844", "CB3000.B2", "21-44641", 33.1, 200),
        (insert_at, "4500030844", "CB4000.B2", "21-44642", 34.4, 80),
    ]
    for row, po, model, sap, price, qty in expected:
        assert ws.cell(row=row, column=2).value == po
        assert ws.cell(row=row, column=3).value == model
        assert ws.cell(row=row, column=4).value == sap
        assert ws.cell(row=row, column=5).value == price
        assert ws.cell(row=row, column=6).value == qty
        assert ws.cell(row=row, column=7).value == "PCS"
        # H 列保留为公式
        assert ws.cell(row=row, column=8).value == f"=E{row}*F{row}"


def test_total_formula_after_insertion(rendered: Path) -> None:
    """断言：合计行公式仍然存在且引用范围合理。

    原公式 =SUM(F16:F26)。插入一行后 openpyxl 应自动平移为 =SUM(F16:F27)
    （或类似的扩展），合计行本身位置也下移到 28。
    """
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    new_total_row = TOTAL_ROW + 1  # 插入一行后下移
    f_total = ws.cell(row=new_total_row, column=6).value
    h_total = ws.cell(row=new_total_row, column=8).value
    assert isinstance(f_total, str) and f_total.startswith("=SUM(F"), f"F 合计公式异常：{f_total}"
    assert isinstance(h_total, str) and h_total.startswith("=SUM(H"), f"H 合计公式异常：{h_total}"


def test_print_area_unchanged(rendered: Path, baseline: dict) -> None:
    """断言：打印区域与原模板一致（原模板未设置打印区域，应保持为空）。"""
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    assert ws.print_area == baseline["print_area"]


def test_post_insert_row_heights_shifted(rendered: Path, baseline: dict) -> None:
    """断言：插入位置之后的行高随行号下移一格仍保持。

    例如：原 row 27 的高度，应出现在新 row 28 上。
    """
    wb = load_workbook(rendered)
    ws = template_sheet(wb)
    insert_at = START_ROW + 2
    failures = []
    for orig_row, expected_height in baseline["row_heights"].items():
        if orig_row < insert_at:
            continue
        new_row = orig_row + 1
        actual_height = ws.row_dimensions[new_row].height
        if actual_height != expected_height:
            failures.append(f"row {orig_row}→{new_row}: {expected_height} → {actual_height}")
    if failures:
        # 把这条转为 xfail 信息：openpyxl 行高平移行为未知，需要观察
        pytest.fail("插入行后部分行高未正确下移：\n" + "\n".join(failures))
