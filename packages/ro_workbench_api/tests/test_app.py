"""工作台 API 边界测试。"""

from __future__ import annotations

import shutil
import time
from collections.abc import Iterator
from pathlib import Path
from typing import Any, cast

import pytest
from fastapi import HTTPException
from fastapi.responses import FileResponse
from fastapi.testclient import TestClient
from pytest import MonkeyPatch
from ro_generator.models import DocumentRequest, GenerationResult
from ro_workbench_api.app import (
    SessionInfo,
    _cleanup_expired_sessions,
    _lock,
    _sessions,
    app,
)

FIXTURE = Path(__file__).resolve().parents[3] / "tests" / "fixtures" / "synthetic_base.xlsx"

client = TestClient(app)


def _clear_sessions() -> None:
    with _lock:
        _sessions.clear()


@pytest.fixture(autouse=True)
def clean_sessions() -> Iterator[None]:
    _clear_sessions()
    yield
    _clear_sessions()


def _response_json(resp: Any) -> dict[str, Any]:
    return cast(dict[str, Any], resp.json())


# --- Session lifecycle ---


def test_open_session_returns_po_list() -> None:
    resp = client.post("/api/session/open", json={"base_file": str(FIXTURE)})
    assert resp.status_code == 200
    data = _response_json(resp)
    assert data["ok"] is True
    assert len(data["session_id"]) == 12
    assert len(data["po_list"]) == 3
    po_nos = {p["po_no"] for p in data["po_list"]}
    assert "4500030844" in po_nos
    assert "4500099999" in po_nos
    assert "4500088888" in po_nos


def test_open_session_returns_invoice_options_by_seller() -> None:
    resp = client.post("/api/session/open", json={"base_file": str(FIXTURE)})

    assert resp.status_code == 200
    data = _response_json(resp)
    po = next(p for p in data["po_list"] if p["po_no"] == "4500099999")
    assert po["invoice_options_by_seller"]["GS PTE"] == ["INV-2603-001"]
    assert po["invoice_options_by_seller"]["EMAX PTE"] == ["INV-2603-001-P"]
    assert po["exportable_documents_by_seller"]["GS PTE"] == ["PI", "PO", "INVOICE_PL"]


def test_open_session_reuses_existing_for_same_base() -> None:
    resp1 = client.post("/api/session/open", json={"base_file": str(FIXTURE)})
    resp2 = client.post("/api/session/open", json={"base_file": str(FIXTURE)})
    assert resp1.status_code == 200
    assert resp2.status_code == 200
    assert resp1.json()["session_id"] == resp2.json()["session_id"]


def test_open_session_creates_new_for_different_base(tmp_path: Path) -> None:
    other = tmp_path / "copy_base.xlsx"
    shutil.copy(FIXTURE, other)
    try:
        resp1 = client.post("/api/session/open", json={"base_file": str(FIXTURE)})
        resp2 = client.post("/api/session/open", json={"base_file": str(other)})
        assert resp1.json()["session_id"] != resp2.json()["session_id"]
    finally:
        other.unlink(missing_ok=True)


def test_close_session_removes_it() -> None:
    resp = client.post("/api/session/open", json={"base_file": str(FIXTURE)})
    sid = resp.json()["session_id"]
    resp2 = client.post("/api/session/close", json={"session_id": sid})
    assert resp2.status_code == 200
    assert resp2.json()["status"] == "closed"
    # Double close is safe
    resp3 = client.post("/api/session/close", json={"session_id": sid})
    assert resp3.json()["status"] == "not_found"


# --- Invalid session rejection ---


def test_dry_run_rejects_invalid_session() -> None:
    resp = client.post(
        "/api/po/4500099999/dry-run",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "GS PTE",
            "buyer": "EMAX PTE",
            "document": "PI",
        },
        headers={"X-Session-Id": "nonexistent"},
    )
    assert resp.status_code == 400
    assert "INVALID_SESSION" in str(resp.json()["detail"])


def test_preview_rejects_invalid_session() -> None:
    resp = client.post(
        "/api/po/4500099999/preview",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "GS PTE",
            "buyer": "EMAX PTE",
            "document": "PI",
        },
        headers={"X-Session-Id": "nonexistent"},
    )
    assert resp.status_code == 400


def test_export_rejects_invalid_session() -> None:
    resp = client.post(
        "/api/po/4500099999/export",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "GS PTE",
            "buyer": "EMAX PTE",
            "document": "PI",
        },
        headers={"X-Session-Id": "nonexistent"},
    )
    assert resp.status_code == 400


def test_get_po_issues_requires_valid_session() -> None:
    resp = client.get(
        f"/api/po/4500088888/issues?base_file={FIXTURE}",
        headers={"X-Session-Id": "nonexistent"},
    )

    assert resp.status_code == 400


def test_get_po_issues_returns_blocking_details() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.get(
        f"/api/po/4500088888/issues?base_file={FIXTURE}",
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    data = _response_json(resp)
    assert data["po_no"] == "4500088888"
    assert data["blocking_count"] >= 1
    assert data["blocking_errors"][0]["kind"] == "blocking_error"
    assert data["blocking_errors"][0]["message"]


def test_export_uses_xlsx_output_format(monkeypatch: MonkeyPatch, tmp_path: Path) -> None:
    captured: dict[str, Any] = {}

    def fake_generate(request: DocumentRequest) -> GenerationResult:
        captured["output_format"] = request.output_format
        output = tmp_path / "YM-GS-PI-4500099999.xlsx"
        output.write_bytes(b"fake xlsx")
        return GenerationResult(
            status="success",
            files=("YM-GS-PI-4500099999.xlsx",),
            output_file=str(output),
        )

    monkeypatch.setattr("ro_workbench_api.app.generate", fake_generate)
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.post(
        "/api/po/4500099999/export",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "YM",
            "document": "PI",
        },
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    assert captured["output_format"] == "xlsx"
    assert resp.json()["output_file"].endswith(".xlsx")


def test_export_invoice_pl_requests_combined_documents(
    monkeypatch: MonkeyPatch, tmp_path: Path
) -> None:
    captured: dict[str, Any] = {}

    def fake_generate(request: DocumentRequest) -> GenerationResult:
        captured["documents"] = request.documents
        captured["output_format"] = request.output_format
        output = tmp_path / "SK-GS-INVOICE&PL-4500099999.xlsx"
        output.write_bytes(b"fake xlsx")
        return GenerationResult(
            status="success",
            files=("SK-GS-INVOICE&PL-4500099999.xlsx",),
            output_file=str(output),
        )

    monkeypatch.setattr("ro_workbench_api.app.generate", fake_generate)
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.post(
        "/api/po/4500099999/export",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "SK",
            "document": "INVOICE_PL",
        },
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    assert captured["documents"] == ("INVOICE", "PL")
    assert captured["output_format"] == "xlsx"
    data = _response_json(resp)
    assert data["files"] == ["SK-GS-INVOICE&PL-4500099999.xlsx"]
    assert data["output_file"].endswith(".xlsx")


def test_get_invoice_groups_uses_session_snapshot() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.get("/api/invoices", headers={"X-Session-Id": sid})

    assert resp.status_code == 200
    groups = resp.json()["invoices"]
    assert groups
    assert all("invoice_group_key" in group for group in groups)
    assert all("invoice_month" not in group for group in groups)


def test_invoice_inspection_returns_rows_and_issue_counts() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]
    groups = client.get("/api/invoices", headers={"X-Session-Id": sid}).json()["invoices"]
    group = next(item for item in groups if item["display_invoice_no"] == "INV-2601-001")

    resp = client.get(
        f"/api/invoice/{group['invoice_group_key']}/inspection",
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    data = resp.json()
    assert data["invoice_group_key"] == group["invoice_group_key"]
    assert data["display_invoice_no"] == "INV-2601-001"
    assert data["line_count"] == len(data["rows"]) == 3
    assert data["blocking_count"] == len(data["blocking_errors"])
    assert data["warnings_count"] == len(data["warnings"])
    assert all(row["ship_qty"] > 0 for row in data["rows"])
    assert "base_file" not in data


def test_invoice_inspection_requires_valid_session() -> None:
    resp = client.get(
        "/api/invoice/invgrp::missing/inspection",
        headers={"X-Session-Id": "missing"},
    )

    assert resp.status_code == 400
    assert resp.json()["detail"]["code"] == "INVALID_SESSION"


def test_preview_invoice_group_uses_session_without_base_file() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]
    group = client.get("/api/invoices", headers={"X-Session-Id": sid}).json()["invoices"][0]

    resp = client.post(
        f"/api/invoice/{group['invoice_group_key']}/preview",
        json={"seller": "GS PTE", "document": "INVOICE"},
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    assert resp.json()["status"] == "success"
    assert resp.json()["invoice_group_key"] == group["invoice_group_key"]
    assert "base_file" not in resp.json()


def test_export_invoice_group_returns_zip() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]
    group = client.get("/api/invoices", headers={"X-Session-Id": sid}).json()["invoices"][0]

    resp = client.post(
        f"/api/invoice/{group['invoice_group_key']}/export",
        json={"seller": "GS PTE", "documents": ["INVOICE", "PL"]},
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    assert resp.json()["status"] == "success"
    assert resp.json()["output_file"].endswith(".xlsx")


def test_invoice_group_endpoints_reject_invalid_session() -> None:
    resp = client.get("/api/invoices", headers={"X-Session-Id": "missing"})
    assert resp.status_code == 400

    resp = client.post(
        "/api/invoice/invgrp::missing/preview",
        json={"seller": "GS PTE", "document": "INVOICE"},
        headers={"X-Session-Id": "missing"},
    )
    assert resp.status_code == 400


def test_export_batch_returns_single_zip(monkeypatch: MonkeyPatch, tmp_path: Path) -> None:
    captured: dict[str, Any] = {}

    def fake_export_document_groups(**kwargs: Any) -> GenerationResult:
        captured.update(kwargs)
        output = tmp_path / "RO-4500099999.zip"
        output.write_bytes(b"fake zip")
        return GenerationResult(
            status="success",
            files=(
                "GS_PTE-GS-PI-4500099999.xlsx",
                "EMAX_PTE-GS-INVOICE&PL-4500099999-INV-2603-001-P.xlsx",
            ),
            output_file=str(output),
        )

    monkeypatch.setattr("ro_workbench_api.app.export_document_groups", fake_export_document_groups)
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.post(
        "/api/po/4500099999/export-batch",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "groups": [
                {"seller": "GS PTE", "documents": ["PI"]},
                {
                    "seller": "EMAX PTE",
                    "documents": ["INVOICE_PL"],
                    "invoice_no": "INV-2603-001-P",
                },
            ],
        },
        headers={"X-Session-Id": sid},
    )

    assert resp.status_code == 200
    groups = captured["groups"]
    assert groups[0].seller == "GS PTE"
    assert groups[0].documents == ("PI",)
    assert groups[1].seller == "EMAX PTE"
    assert groups[1].documents == ("INVOICE", "PL")
    data = _response_json(resp)
    assert data["output_file"].endswith(".zip")
    assert data["files"] == [
        "GS_PTE-GS-PI-4500099999.xlsx",
        "EMAX_PTE-GS-INVOICE&PL-4500099999-INV-2603-001-P.xlsx",
    ]


# --- Session expiry ---


def test_cleanup_removes_expired_session() -> None:
    """Expired sessions are removed by cleanup (not by _get_session which doesn't check TTL)."""
    with _lock:
        _sessions["exp"] = SessionInfo(
            session_id="exp",
            base_file=str(FIXTURE),
            temp_dir="/tmp/fake-exp",
            created_at=0,
            last_access=0,
        )
    _cleanup_expired_sessions()
    with _lock:
        assert "exp" not in _sessions


def test_cleanup_keeps_fresh_sessions() -> None:
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]
    _cleanup_expired_sessions()
    with _lock:
        assert sid in _sessions


# --- Download path traversal protection ---


def test_download_cannot_read_another_session_file(tmp_path: Path) -> None:
    own_dir = tmp_path / "own"
    other_dir = tmp_path / "other"
    own_dir.mkdir()
    other_dir.mkdir()
    other_file = other_dir / "invoice.xlsx"
    other_file.write_bytes(b"not a real workbook; only auth boundary matters")

    with _lock:
        _sessions.clear()
        _sessions["own"] = SessionInfo(
            session_id="own",
            base_file="/tmp/own-base.xlsx",
            temp_dir=str(own_dir),
            created_at=time.time(),
            last_access=time.time(),
        )
        _sessions["other"] = SessionInfo(
            session_id="other",
            base_file="/tmp/other-base.xlsx",
            temp_dir=str(other_dir),
            created_at=time.time(),
            last_access=time.time(),
        )

    try:
        from ro_workbench_api.app import download_file

        with pytest.raises(HTTPException) as exc_info:
            response = download_file(path=str(other_file), x_session_id="own")
            assert isinstance(response, FileResponse)

        assert exc_info.value.status_code == 403
    finally:
        with _lock:
            _sessions.clear()
        shutil.rmtree(own_dir, ignore_errors=True)
        shutil.rmtree(other_dir, ignore_errors=True)


# --- check-path endpoint ---


def test_check_path_valid_file() -> None:
    resp = client.post("/api/check-path", json={"path": str(FIXTURE)})
    assert resp.status_code == 200
    data = _response_json(resp)
    assert data["ok"] is True
    assert "DATA BASE" in data["sheets"] or "PO record" in data["sheets"]
    assert data["size"] > 0


def test_check_path_nonexistent() -> None:
    resp = client.post("/api/check-path", json={"path": "/nonexistent/file.xlsx"})
    assert resp.status_code == 200
    data = _response_json(resp)
    assert data["ok"] is False
    assert "不存在" in data["error"]


def test_check_path_unsupported_extension(tmp_path: Path) -> None:
    f = tmp_path / "test.txt"
    f.write_text("hello")
    resp = client.post("/api/check-path", json={"path": str(f)})
    assert resp.status_code == 200
    data = _response_json(resp)
    assert data["ok"] is False
    assert "不支持" in data["error"]


def test_po_export_pdf_passthrough_and_download(
    monkeypatch: MonkeyPatch,
) -> None:
    captured: dict[str, Any] = {}

    def fake_generate(request: DocumentRequest) -> GenerationResult:
        captured["output_format"] = request.output_format
        output = Path(request.output_dir) / "GS-RO-INVOICE-4500099999-INV-001.pdf"
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_bytes(b"%PDF-1.4 fake")
        return GenerationResult(
            status="success",
            files=(output.name,),
            output_file=str(output),
        )

    monkeypatch.setattr("ro_workbench_api.app.generate", fake_generate)
    sid = _response_json(client.post("/api/session/open", json={"base_file": str(FIXTURE)}))[
        "session_id"
    ]

    resp = client.post(
        "/api/po/4500099999/export",
        json={
            "base_file": str(FIXTURE),
            "po_no": "4500099999",
            "seller": "GS PTE",
            "invoice_no": "INV-001",
            "documents": ["INVOICE"],
            "output_format": "pdf",
        },
        headers={"X-Session-Id": sid},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert captured["output_format"] == "pdf"
    assert body["output_file"].endswith(".pdf")

    dl = client.get(
        "/api/download",
        params={"path": body["output_file"]},
        headers={"X-Session-Id": sid},
    )
    assert dl.status_code == 200
    assert dl.headers["content-type"] == "application/pdf"


# --- Health ---


def test_health_returns_ok() -> None:
    resp = client.get("/api/health")
    assert resp.status_code == 200
    assert resp.json() == {"status": "ok"}


# --- edit endpoint ---


def test_edit_field_rejects_invalid_sheet(tmp_path: Path) -> None:
    base = tmp_path / "test_base.xlsx"
    shutil.copy(FIXTURE, base)
    try:
        resp = client.post(
            "/api/po/TEST/edit",
            json={
                "base_file": str(base),
                "sheet": "NonExistent",
                "row": 5,
                "field": "X",
                "value": "test",
            },
        )
        assert resp.status_code == 200
        data = _response_json(resp)
        assert data["ok"] is False
    finally:
        base.unlink(missing_ok=True)


def test_edit_field_writes_value(tmp_path: Path) -> None:
    import openpyxl

    base = tmp_path / "test_base.xlsx"
    shutil.copy(FIXTURE, base)
    try:
        # First open a session so cache is primed
        client.post("/api/session/open", json={"base_file": str(base)})

        # Get the actual header name from the PO record
        wb = openpyxl.load_workbook(str(base), data_only=True)
        ws = wb["PO record"]
        header_cell = ws.cell(row=4, column=5).value  # Column E, row 4
        wb.close()
        field_name = str(header_cell).strip().replace("\n", " ")

        resp = client.post(
            "/api/po/TEST/edit",
            json={
                "base_file": str(base),
                "sheet": "PO record",
                "row": 5,
                "field": field_name,
                "value": "TEST_VALUE",
            },
        )
        data = _response_json(resp)
        assert data["ok"] is True, f"Edit failed: {data.get('message')}"

        # Verify the value was actually written
        wb = openpyxl.load_workbook(str(base), data_only=True)
        ws = wb["PO record"]
        cell_val = (
            str(ws.cell(row=5, column=5).value)
            if ws.cell(row=5, column=5).value is not None
            else ""
        )
        wb.close()
        assert cell_val == "TEST_VALUE"
    finally:
        base.unlink(missing_ok=True)
