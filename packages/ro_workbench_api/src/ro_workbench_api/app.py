"""工作台后端 API：ro_generator 核心包的 FastAPI 薄包装层。

所有业务逻辑在 ro_generator 中，此层只负责：
- HTTP 协议（JSON 序列化 / 路由 / CORS）
- 请求到 DocumentRequest 的转换
- GenerationResult 到 JSON 响应的转换

禁止：在本层写任何校验、价格计算、业务判断。
"""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field

from ro_generator.errors import RoGeneratorError, WorkbookOpenError
from ro_generator.generator import generate
from ro_generator.models import DocumentRequest, GenerationResult, ValidationMessage
from ro_generator.resolver import resolve_po_lines
from ro_generator.schema import LEGAL_CHAIN_SEGMENTS
from ro_generator.source_index import SourceIndex
from ro_generator.validator import validate_workbook_structure
from ro_generator.workbook_reader import ROW_NUMBER_KEY, WorkbookReader

app = FastAPI(title="RO Workbench API", version="0.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# —————————————————————————————————————
# Request / Response 模型
# —————————————————————————————————————

class OpenSessionRequest(BaseModel):
    base_file: str


class PoListEntry(BaseModel):
    po_no: str
    status: str  # ready / partial / blocked
    chain_segments: list[dict[str, str]]
    line_count: int
    monthly_months: list[str]


class CellValue(BaseModel):
    value: object
    row: int
    col: str  # column letter
    field: str
    is_missing: bool = False
    is_formula_fallback: bool = False


class DataViewResponse(BaseModel):
    po_no: str
    headers: list[dict[str, object]]
    rows: list[dict[str, object]]


class DryRunRequest(BaseModel):
    base_file: str
    po_no: str
    seller: str
    buyer: str
    invoice_month: str | None = None
    invoice_no: str | None = None


class EditFieldRequest(BaseModel):
    base_file: str
    sheet: str
    row: int
    field: str
    value: object


class EditFieldResponse(BaseModel):
    ok: bool
    message: str = ""


# —————————————————————————————————————
# 辅助
# —————————————————————————————————————

def _result_to_dict(result: GenerationResult) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "status": result.status,
        "summary": result.summary,
        "files": list(result.files),
        "output_file": result.output_file,
        "errors": [_msg_to_dict(m) for m in result.errors],
        "warnings": [_msg_to_dict(m) for m in result.warnings],
        "missing_inputs": list(result.missing_inputs),
        "options": {k: list(v) for k, v in result.options.items()},
    }
    if isinstance(result.source_index, SourceIndex):
        payload["source_index"] = [
            {
                "doc_cell": cell,
                "source": {"sheet": loc.sheet, "row": loc.row, "field": loc.field, "is_computed": loc.is_computed},
            }
            for cell, loc in result.source_index
        ]
    else:
        payload["source_index"] = []
    return payload


def _msg_to_dict(m: ValidationMessage) -> dict[str, Any]:
    return asdict(m)


# —————————————————————————————————————
# 端点
# —————————————————————————————————————

@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/session/open")
def open_session(req: OpenSessionRequest) -> dict[str, Any]:
    """打开 base 文件，返回 PO 列表 + 状态摘要。

    PO 状态判定（产品方案 §8.2）：
    - ready: 全部行都有 SAP/QTY/价格，至少一个合法链段完整
    - partial: PI/PO 就绪但 Invoice 缺 INV# 等
    - blocked: 有阻断错误（缺 SAP、SAP 不在 DATA BASE 等）
    """
    try:
        reader = WorkbookReader(req.base_file)
    except WorkbookOpenError as exc:
        raise HTTPException(400, detail={"code": exc.code, "message": exc.message}) from exc

    try:
        struct = validate_workbook_structure(reader)
        if struct:
            return {"ok": False, "errors": [_msg_to_dict(m) for m in struct]}

        po_sheet = reader.read_sheet("PO record")
        po_map: dict[str, list[dict[str, object]]] = {}
        for row in po_sheet.rows:
            po_no = str(row.get("PO NO.", "")).strip()
            if not po_no:
                continue
            po_map.setdefault(po_no, []).append(row)

        po_list: list[dict[str, Any]] = []
        for po_no, rows in po_map.items():
            resolve_result = resolve_po_lines(reader, po_no)
            blocking = [m for m in resolve_result.messages if m.kind == "blocking_error"]

            priced_segments: list[dict[str, str]] = []
            for seg in LEGAL_CHAIN_SEGMENTS:
                if all(seg in line.prices for line in resolve_result.lines):
                    priced_segments.append({"seller": seg[0], "buyer": seg[1]})

            months: set[str] = set()
            for line in resolve_result.lines:
                months.update(line.monthly_shipments.keys())

            status = "blocked" if blocking else ("ready" if priced_segments else "partial")
            po_list.append({
                "po_no": po_no,
                "status": status,
                "chain_segments": priced_segments,
                "line_count": len(resolve_result.lines) or len(rows),
                "monthly_months": sorted(months),
                "blocking_count": len(blocking),
            })

        return {"ok": True, "po_list": po_list}
    finally:
        reader.close()


@app.get("/po/{po_no}")
def get_po_data(po_no: str, base_file: str = Query(...)) -> dict[str, Any]:
    """返回 PO 行数据视图（grid 数据）。

    返回每行的 dict，携带 __row_number__ 供前端溯源。
    """
    try:
        with WorkbookReader(base_file) as reader:
            sheet = reader.read_sheet("PO record")
    except WorkbookOpenError as exc:
        raise HTTPException(400, detail={"code": exc.code, "message": exc.message}) from exc

    rows = [dict(row) for row in sheet.rows if str(row.get("PO NO.", "")).strip() == po_no]
    return {
        "po_no": po_no,
        "headers": list(sheet.headers),
        "rows": rows,
    }


@app.post("/po/{po_no}/dry-run")
def dry_run(po_no: str, req: DryRunRequest) -> dict[str, Any]:
    """装配预览，返回数据摘要 + source_index。

    文件写入临时目录，通过 /download?path= 可访问。
    """
    import tempfile
    out_dir = tempfile.mkdtemp(prefix="ro-dry-run-")
    request = DocumentRequest(
        base_file=req.base_file,
        po_no=po_no,
        documents=("INVOICE",),
        seller=req.seller,
        buyer=req.buyer,
        invoice_month=req.invoice_month,
        invoice_no=req.invoice_no,
        output_dir=out_dir,
    )
    result = generate(request)
    return _result_to_dict(result)


@app.post("/po/{po_no}/edit")
def edit_field(po_no: str, req: EditFieldRequest) -> EditFieldResponse:
    """接受字段编辑，写回 base 文件。

    使用 openpyxl 直接修改单元格值（非 read_only 模式）。
    """
    import openpyxl

    try:
        path = Path(req.base_file)
        if not path.exists():
            return EditFieldResponse(ok=False, message=f"base 文件不存在：{path}")

        wb = openpyxl.load_workbook(str(path))
        if req.sheet not in wb.sheetnames:
            return EditFieldResponse(ok=False, message=f"sheet {req.sheet!r} 不存在")

        ws = wb[req.sheet]
        # 通过表头定位列索引
        header_row = 4
        col_map: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=c)
            if cell.value is not None:
                from ro_generator.schema import normalize_header

                norm = normalize_header(cell.value)
                if norm:
                    col_map[norm] = c

        col_idx = col_map.get(req.field)
        if col_idx is None:
            return EditFieldResponse(ok=False, message=f"表头中找不到字段 {req.field!r}")

        ws.cell(row=req.row, column=col_idx, value=req.value)
        wb.save(str(path))
        return EditFieldResponse(ok=True, message=f"已更新 {req.sheet} row={req.row} {req.field}")
    except Exception as exc:
        return EditFieldResponse(ok=False, message=str(exc))


@app.post("/export")
def export_documents(req: DryRunRequest) -> dict[str, Any]:
    """执行真实导出，返回生成的文件。

    成功时 output_file 指向导出的 xlsx 文件。
    """
    # 创建临时输出目录
    import tempfile

    out_dir = tempfile.mkdtemp(prefix="ro-export-")
    request = DocumentRequest(
        base_file=req.base_file,
        po_no=req.po_no,
        documents=("INVOICE",),
        seller=req.seller,
        buyer=req.buyer,
        invoice_month=req.invoice_month,
        invoice_no=req.invoice_no,
        output_dir=out_dir,
    )
    result = generate(request)
    return _result_to_dict(result)


@app.get("/download")
def download_file(path: str = Query(...)) -> FileResponse:
    p = Path(path)
    if not p.exists():
        raise HTTPException(404, detail=f"file not found: {path}")
    return FileResponse(str(p), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
