"""Renderer：把 DocumentModel 写入 Excel 模板，输出可下载的 .xlsx。

设计要点（核心来自 Phase 0 Spike A 的发现）：
- openpyxl `insert_rows()` 不平移 `row_dimensions`，必须先**倒序手动 += 1 行号**再调用
  insert_rows（详见 `_insert_styled_row` 实现）
- 样式复制使用 `copy.copy()` 处理 Font / Fill / Border / Alignment，行高从 style_source_row
  借用
- 模板里有样板数据（如 GS Invoice 行 18-24 有示例 SKU），渲染前必须先清掉
- totals 单元格的位置在数据行扩展时由 openpyxl 自动平移（公式如 =SUM(F16:F26) 也会
  跟着扩展），调用方拿到的最终行号会上移
- 渲染期间累积双向溯源索引（产品方案 §4.4），随 RenderResult 返回

输入：`DocumentModel` + `TemplateMapping` + 输出路径。
输出：RenderResult（文件路径 + 源索引）。
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any, Final, cast

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ro_generator.document_model import DocumentLine, DocumentModel
from ro_generator.errors import InternalError, TemplateError
from ro_generator.header_rules import (
    build_header_resolved_values,
    is_system_generated_header_field,
    resolve_header_field_spec,
)
from ro_generator.line_rules import (
    USD_NUMBER_FORMAT,
    LineFieldSpec,
    line_excel_number_format,
    resolve_line_field_spec,
    uses_po_record_row,
)
from ro_generator.profiles.runtime import current_schema
from ro_generator.schema import SHEET_PO_RECORD
from ro_generator.source_index import SourceIndex, SourceIndexBuilder, SourceLocation
from ro_generator.template_mapping import LineColumns, TemplateMapping, iter_line_columns
from ro_generator.totals_rules import total_spec_for_mapping_key, total_value_for_mapping_key

# —————————————————————————————————————
# 公开 API
# —————————————————————————————————————


@dataclass(frozen=True)
class RenderResult:
    """渲染输出。

    - `output_path`：写好的 .xlsx 绝对路径
    - `source_index`：装配单元格 ↔ base 字段的双向映射
    """

    output_path: Path
    source_index: SourceIndex


def render_document(
    model: DocumentModel,
    mapping: TemplateMapping,
    output_path: str | Path,
) -> RenderResult:
    """把 model 渲染到 mapping.template_path 上指定的模板，保存到 output_path。

    返回 RenderResult，含文件绝对路径与源索引。
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(filename=str(mapping.template_path))
    except Exception as exc:
        raise TemplateError(f"无法打开模板 {mapping.template_path}：{exc}") from exc

    builder = SourceIndexBuilder()
    try:
        _render_into_workbook(wb, model, mapping, builder)

        # 移除模板中未被使用的 sheet（如 GS PL 模板附带 INV sheet）。
        # 前端 SheetJS 默认读第一张 sheet，保留多余 sheet 会导致预览显示错误内容。
        for sn in list(wb.sheetnames):
            if sn != mapping.sheet:
                del wb[sn]

        wb.save(output_path)
    finally:
        wb.close()

    return RenderResult(
        output_path=output_path.resolve(),
        source_index=builder.build(),
    )


def render_document_bundle(
    items: tuple[tuple[DocumentModel, TemplateMapping], ...],
    output_path: str | Path,
) -> RenderResult:
    """把多个 DocumentModel 渲染到同一个 Excel 模板的多个 sheet。

    用于 SK/YM 的 Invoice + PL 合并导出。所有 mapping 必须指向同一个模板文件。
    """
    if not items:
        raise TemplateError("render_document_bundle 至少需要一个单据")

    template_path = items[0][1].template_path
    for _, mapping in items:
        if mapping.template_path != template_path:
            raise TemplateError("组合导出的 mapping 必须使用同一个模板文件")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = load_workbook(filename=str(template_path))
    except Exception as exc:
        raise TemplateError(f"无法打开模板 {template_path}：{exc}") from exc

    builder = SourceIndexBuilder()
    try:
        used_sheets = {mapping.sheet for _, mapping in items}
        for model, mapping in items:
            _render_into_workbook(wb, model, mapping, builder)

        for sn in list(wb.sheetnames):
            if sn not in used_sheets:
                del wb[sn]

        wb.save(output_path)
    finally:
        wb.close()

    return RenderResult(
        output_path=output_path.resolve(),
        source_index=builder.build(),
    )


def _render_into_workbook(
    wb: Workbook,
    model: DocumentModel,
    mapping: TemplateMapping,
    builder: SourceIndexBuilder,
) -> None:
    if mapping.sheet not in wb.sheetnames:
        raise TemplateError(f"模板 {mapping.template_path.name} 中找不到 sheet {mapping.sheet!r}")
    ws: Worksheet = wb[mapping.sheet]

    _write_styles(ws, mapping)
    _write_header(ws, model, mapping, builder)
    row_offset = _write_lines_and_totals(ws, model, mapping, builder)
    _write_cost_breakdown(ws, model, mapping, row_offset, builder)
    _write_notes(ws, model, mapping, row_offset, builder)
    _apply_print_layout(ws)


# —————————————————————————————————————
# Header
# —————————————————————————————————————

# 模板表头中属于固定文案的 key 集合。这类字段保留模板原值，不清除。
_PRESERVE_HEADER_KEYS: Final[set[str]] = {
    "payment_terms",
    "port_of_loading",
    "final_destination",
    "manufacturer",
    "manufacturer_name",
    "manufacturer_address",
    "manufacturer_address_2",
    "supplier",
    "bill_to",
    "bill_to_line2",
    "bill_to_line3",
    "shipped_per",
    "signature",
}

# 缺值占位符的中文标签（key = mapping key）
_PLACEHOLDER_LABELS: dict[str, str] = {
    "invoice_no": "需填: INV#",
    "ship_to": "需填: Ship To",
    "bill_to": "需填: Ship To",
    "po_no": "需填: PO#",
    "unit_price": "需填: 单价",
    "quantity": "需填: 数量",
    "ex_factory_date": "需填: 出厂日期",
}


def _write_styles(ws: Worksheet, mapping: TemplateMapping) -> None:
    """按 mapping.style 声明应用单元格样式。"""
    for addr in mapping.style.bold:
        cell = ws[addr]
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=True,
            italic=cell.font.italic,
            underline=cell.font.underline,
            color=cell.font.color,
        )
    for addr in mapping.style.underline:
        cell = ws[addr]
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=cell.font.bold,
            italic=cell.font.italic,
            underline="single",
            color=cell.font.color,
        )


def _write_header(
    ws: Worksheet,
    model: DocumentModel,
    mapping: TemplateMapping,
    builder: SourceIndexBuilder,
) -> None:
    """按 mapping.header 写表头字段。值为 None 时写 [需填: ...] 占位符。
    日期类字段写入当天日期；固定文案字段保留模板原值；其余未映射字段清除模板样本值。
    """
    written_addrs: set[str] = set()
    resolved_values = build_header_resolved_values(
        model,
        header_keys=mapping.header.keys(),
        header_fixed=mapping.header_fixed,
    )

    for mkey, cell_addr in mapping.header.items():
        value = resolved_values.get(mkey)
        spec = resolve_header_field_spec(
            mkey,
            seller=model.seller,
            document_type=model.document_type,
        )
        if value is not None:
            _safe_set_cell(ws, cell_addr, value)
            written_addrs.add(cell_addr)
            if (
                mkey not in mapping.header_fixed
                and spec is not None
                and spec.source_type == "base_field"
            ):
                sheet = spec.source_sheet if spec and spec.source_sheet else "PO record"
                field = spec.source_field if spec and spec.source_field else mkey
                builder.add(
                    cell_addr,
                    SourceLocation(sheet=sheet, row=None, field=field),
                )
            elif (
                mkey not in mapping.header_fixed
                and spec is not None
                and spec.source_type not in {"template_content", "base_field"}
            ):
                builder.add_computed(cell_addr, mkey)
            continue
        if is_system_generated_header_field(
            mkey,
            seller=model.seller,
            document_type=model.document_type,
        ):
            # 写入日期字符串。openpyxl 写 date 对象会把值序列化为 Excel 数字，
            # 而 SheetJS 预览读到的 raw value 仍是数字而非格式化日期。
            # 直接用字符串确保浏览器端显示为人类可读日期。
            _safe_set_cell(ws, cell_addr, date.today().strftime("%Y-%m-%d"))
        elif mkey in _PRESERVE_HEADER_KEYS and (
            mkey in mapping.header_fixed or spec is None or spec.source_type == "template_content"
        ):
            pass  # 保留模板原值
        else:
            # 清除模板样本值（如样本日期、注释占位符）
            _safe_set_cell(ws, cell_addr, None)


# —————————————————————————————————————
# Lines + Totals
# —————————————————————————————————————


def _write_lines_and_totals(
    ws: Worksheet,
    model: DocumentModel,
    mapping: TemplateMapping,
    builder: SourceIndexBuilder,
) -> int:
    """把订单行写到模板，超出预留空间则插入新行并复制样式。"""
    line_count = len(model.lines)
    if line_count == 0:
        # 没有行数据时只清理样板，写合计为 0
        _clear_sample_rows(ws, mapping)
        _write_totals(ws, model, mapping, totals_row_offset=0, builder=builder)
        return 0

    start_row = mapping.lines.start_row
    style_source_row = mapping.lines.style_source_row
    reserved_rows = _reserved_row_count(mapping)

    # 1. 统一预留数据区样式，再清掉模板里的样板数据。
    # 某些模板的预留行携带了脏 number_format（例如把单价列误设成日期）。
    # 统一继承 style_source_row，可保证导出结果与新增插入行的样式口径一致。
    _normalize_reserved_row_styles(ws, start_row, reserved_rows, style_source_row)
    _clear_sample_rows(ws, mapping)

    # 2. 如果实际行数 > 预留行数，插入额外的新行
    insertion_count = max(0, line_count - reserved_rows)
    if insertion_count > 0:
        # 在 reserved 区间末尾的下一行插入，避开样式源行
        insert_at = start_row + reserved_rows
        for _ in range(insertion_count):
            _insert_styled_row(ws, insert_at, style_src=style_source_row)

    # 3. 写每一行
    columns = mapping.lines.columns
    line_unit_label = mapping.lines.unit_label
    for offset, doc_line in enumerate(model.lines):
        row = start_row + offset
        _write_data_row(
            ws,
            row,
            doc_line,
            columns,
            line_unit_label,
            builder,
            document_type=model.document_type,
            seller=model.seller,
        )
        # Write row_fixed values (e.g., Country of The Origin = "China")
        for col_letter, fixed_val in mapping.lines.row_fixed.items():
            ws[f"{col_letter}{row}"] = fixed_val

    # 4. 写合计（位置已被 openpyxl 自动平移）
    _write_totals(ws, model, mapping, totals_row_offset=insertion_count, builder=builder)
    return insertion_count


def _reserved_row_count(mapping: TemplateMapping) -> int:
    """从 mapping 推断模板预留多少行数据空间。

    取 totals 单元格的最小行号 − start_row。如果 mapping.totals 没有任何单元格，
    退化为 1（只用 style_source_row 一行）。
    """
    if not mapping.totals:
        return 1
    totals_rows: list[int] = []
    for total_cell in mapping.totals.values():
        try:
            _, row = coordinate_from_string(total_cell.cell.strip())
        except ValueError as exc:
            raise InternalError(
                f"mapping totals 的单元格地址 {total_cell.cell!r} 无法解析"
            ) from exc
        totals_rows.append(row)
    min_totals_row = min(totals_rows)
    reserved = min_totals_row - mapping.lines.start_row
    return max(reserved, 1)


def _clear_sample_rows(ws: Worksheet, mapping: TemplateMapping) -> None:
    """清掉模板预留区间里的样板数据。

    清除 start_row 到 start_row + reserved - 1 之间所有映射列引用到的单元格，
    以及 row_fixed 列和 PL 装箱列。同时清除 start_row 上方和 totals 下方的注释/样板行。
    """
    reserved = _reserved_row_count(mapping)
    columns = mapping.lines.columns
    addrs = [v for _, v in iter_line_columns(columns)]
    # 追加 row_fixed 列
    for letter in mapping.lines.row_fixed:
        if letter not in addrs:
            addrs.append(letter)

    # 确定 header 区域的最末行（用于清除 header 与 data 之间的样板行）
    header_max_row = 0
    for addr in mapping.header.values():
        try:
            _, row = coordinate_from_string(addr.strip())
            if row > header_max_row:
                header_max_row = row
        except ValueError:
            pass

    # 清除 data 上方的样板行（从 header 结束到 start_row 之间的所有行）
    # 如模板显式声明了表格表头行，则保留该行，不再依赖启发式猜测。
    clear_above_start = max(header_max_row + 1, 1)
    for r in range(clear_above_start, mapping.lines.start_row):
        if r in mapping.table_header_row:
            continue
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None:
                cell.value = None

    # 清除 data 行中的样板数据
    for offset in range(reserved):
        row = mapping.lines.start_row + offset
        for letter in addrs:
            cell = ws[f"{letter}{row}"]
            cell.value = None

    # 清除 totals 值单元格，保留模板中的 label 文案（如 Total / Signature / Date）
    if mapping.totals:
        for total_cell in mapping.totals.values():
            try:
                col_letter, row = coordinate_from_string(total_cell.cell.strip())
            except ValueError:
                continue
            ws[f"{col_letter}{row}"].value = None


def _write_data_row(
    ws: Worksheet,
    row: int,
    doc_line: DocumentLine,
    columns: LineColumns,
    fixed_unit_label: str | None,
    builder: SourceIndexBuilder,
    *,
    document_type: str,
    seller: str,
) -> None:
    """数据驱动的行写入：遍历 YAML columns 中声明的所有列，按共享 line spec 写入。"""
    src_row = doc_line.source_row  # 可能为 None（合成数据 / 测试场景）

    for key, col_letter in iter_line_columns(columns):
        spec = resolve_line_field_spec(
            key,
            document_type=document_type,
            seller=seller,
            category=doc_line.category,
        )
        addr = f"{col_letter}{row}"

        # — fixed_value：固定文案（unit_label 等）
        if spec.fixed_value:
            if fixed_unit_label:
                ws[addr] = fixed_unit_label
            continue

        val = getattr(doc_line, key, None)

        # — computed：公式计算值（amount），只记 source 索引
        if spec.computed:
            if val is not None:
                _write_line_cell(ws[addr], val, spec)
            packing_sheet, packing_field = _packing_weight_location(doc_line, key)
            if packing_field:
                builder.add(
                    addr,
                    SourceLocation(
                        packing_sheet or current_schema().sheet("DATA BASE").name,
                        src_row
                        if packing_sheet == current_schema().sheet("PO record").name
                        else None,
                        packing_field,
                    ),
                )
            else:
                builder.add_computed(addr, key)
            continue

        source_field = spec.source_field or key
        if key == "quantity" and doc_line.quantity_source_field:
            source_field = doc_line.quantity_source_field
        source_sheet = spec.source_sheet or SHEET_PO_RECORD

        # — skip_if_none：空值跳过（PL 专属装箱字段）
        if spec.skip_if_none and val is None:
            continue

        # — 占位符（零值 / 空值）
        if isinstance(val, Decimal) and val == 0 and spec.zero_placeholder:
            ws[addr] = spec.zero_placeholder
        elif val is None and spec.none_placeholder:
            ws[addr] = spec.none_placeholder
        elif val is not None:
            _write_line_cell(ws[addr], val, spec)

        builder.add(
            addr,
            SourceLocation(
                source_sheet,
                src_row if uses_po_record_row(spec) else None,
                source_field,
            ),
        )


def _write_totals(
    ws: Worksheet,
    model: DocumentModel,
    mapping: TemplateMapping,
    totals_row_offset: int,
    builder: SourceIndexBuilder,
) -> None:
    """写合计单元格。位置随 insertion_count 平移。"""
    for field_name, total_cell in mapping.totals.items():
        try:
            col_letter, row = coordinate_from_string(total_cell.cell.strip())
        except ValueError as exc:
            raise InternalError(f"mapping totals 单元格地址 {total_cell.cell!r} 无法解析") from exc
        new_row = row + totals_row_offset
        cell_addr = f"{col_letter}{new_row}"
        if total_cell.value_mode == "fixed":
            ws[cell_addr] = total_cell.value
            builder.add_computed(cell_addr, f"totals.{field_name}")
            continue
        if total_cell.value_mode == "current_date":
            ws[cell_addr] = date.today().strftime("%Y-%m-%d")
            builder.add_computed(cell_addr, f"totals.{field_name}")
            continue
        if total_cell.value_mode == "model_date":
            if model.document_date is None:
                continue
            ws[cell_addr] = model.document_date.strftime("%Y-%m-%d")
            schema = current_schema()
            builder.add(
                cell_addr,
                SourceLocation(
                    sheet=schema.sheet("客户PO").name,
                    row=None,
                    field=schema.field("客户PO", "document_date"),
                ),
            )
            continue
        total_spec = total_spec_for_mapping_key(field_name)
        if total_spec is None:
            continue
        total_value = total_value_for_mapping_key(model, field_name)
        if total_value is not None:
            ws[cell_addr] = total_value
            if field_name == "amount":
                ws[cell_addr].number_format = USD_NUMBER_FORMAT
            # 合计是工作台计算得出，不指向某一行
            builder.add_computed(cell_addr, total_spec.preview_key)


def _write_cost_breakdown(
    ws: Worksheet,
    model: DocumentModel,
    mapping: TemplateMapping,
    row_offset: int,
    builder: SourceIndexBuilder,
) -> None:
    """写入可选的 Combo RODS/REELS 成本拆分表。"""

    section = mapping.cost_breakdown
    if section is None:
        return

    start_row = section.start_row + row_offset
    style_source_row = section.style_source_row + row_offset
    reserved_rows = section.reserved_rows
    columns = section.columns

    def clear_row(row: int) -> None:
        for column in columns.values():
            ws[f"{column}{row}"] = None

    for offset in range(reserved_rows):
        clear_row(start_row + offset)

    insertion_count = max(0, len(model.cost_breakdown) - reserved_rows)
    insert_at = start_row + reserved_rows
    for _ in range(insertion_count):
        _insert_styled_row(ws, insert_at, style_src=style_source_row)

    schema = current_schema()
    identity_fields = {
        "po_no": (schema.sheet("PO record").name, schema.field("PO record", "po_no")),
        "item_line_no": (
            schema.sheet("PO record").name,
            schema.field("PO record", "item_line"),
        ),
        "item_number": (schema.sheet("PO record").name, schema.field("PO record", "sap")),
        "description": (
            schema.sheet("PO record").name,
            schema.field("PO record", "description"),
        ),
    }
    for index, line in enumerate(model.cost_breakdown):
        row = start_row + index
        values = {
            "po_no": line.po_no,
            "item_line_no": line.item_line_no,
            "item_number": line.item_number,
            "description": line.description,
            "unit_price": line.unit_price,
        }
        for key, column in columns.items():
            value = values.get(key)
            if value is None:
                continue
            cell = ws[f"{column}{row}"]
            if key == "unit_price":
                _write_line_cell(
                    cell,
                    value,
                    LineFieldSpec(
                        rule=f'DATA BASE 的 "{line.source_field}" 列',
                        source_sheet=schema.sheet("DATA BASE").name,
                        source_field=line.source_field,
                        display_decimal_places=2,
                        display_prefix="$",
                    ),
                )
            else:
                cell.value = cast(Any, value)

            if key in identity_fields:
                source_sheet, source_field = identity_fields[key]
                builder.add(
                    f"{column}{row}",
                    SourceLocation(
                        sheet=source_sheet,
                        row=line.source_row,
                        field=source_field,
                    ),
                )
            elif key == "description":
                builder.add(
                    f"{column}{row}",
                    SourceLocation(
                        sheet=schema.sheet("PO record").name,
                        row=line.source_row,
                        field=schema.field("PO record", "description"),
                    ),
                )
            elif key == "unit_price":
                builder.add(
                    f"{column}{row}",
                    SourceLocation(
                        sheet=schema.sheet("DATA BASE").name,
                        row=None,
                        field=line.source_field or "component_price",
                    ),
                )

        # 组件标识是由 Profile 规则生成的说明，不再额外写入模板列。


def _write_notes(
    ws: Worksheet,
    model: DocumentModel,
    mapping: TemplateMapping,
    row_offset: int,
    builder: SourceIndexBuilder,
) -> None:
    """写模板底部说明字段，例如 PL 的 PACKED IN {CTNS}。"""
    for field_name, cell in mapping.notes.items():
        value = _note_value_for_mapping_key(model, field_name)
        if value is None:
            continue
        try:
            col_letter, row = coordinate_from_string(cell.strip())
        except ValueError as exc:
            raise InternalError(f"mapping notes 的单元格地址 {cell!r} 无法解析") from exc
        cell_addr = f"{col_letter}{row + row_offset}"
        ws[cell_addr] = value
        builder.add_computed(cell_addr, f"notes.{field_name}")


def _note_value_for_mapping_key(model: DocumentModel, field_name: str) -> str | None:
    if field_name in {"packed_in", "packed_in_ctns"}:
        if model.total_carton_count is None:
            return None
        return f"PACKED IN {_format_decimal_for_note(model.total_carton_count)} CTNS"
    return None


def _format_decimal_for_note(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(value.quantize(Decimal("1")))
    return format(value.normalize(), "f")


# —————————————————————————————————————
# Spike A 验证过的样式插入逻辑
# —————————————————————————————————————


def _insert_styled_row(ws: Worksheet, insert_at: int, style_src: int) -> None:
    """在 insert_at 处插入一行，并执行 openpyxl 不会自动做的事：

    - 平移 insert_at 之后所有 row_dimensions 一格（openpyxl 只移单元格内容，不移
      row_dimensions，不修复会导致行高错乱）
    - 把 style_src 行的样式复制到新插入的行

    详见 Phase 0 Spike A 的结论 (`docs/development/phase-0-spike-results.md`)。
    """
    # 1. 倒序处理，避免覆盖
    existing_rows = sorted(
        (r for r in ws.row_dimensions if r >= insert_at),
        reverse=True,
    )
    for orig_row in existing_rows:
        src_dim = ws.row_dimensions[orig_row]
        new_dim = ws.row_dimensions[orig_row + 1]
        new_dim.height = src_dim.height
        new_dim.hidden = src_dim.hidden
        new_dim.outlineLevel = src_dim.outlineLevel
    if insert_at in ws.row_dimensions:
        ws.row_dimensions[insert_at].height = None

    # 2. openpyxl 处理单元格内容、公式、合并区域的下移
    ws.insert_rows(insert_at)

    # 3. 把样板行样式复制到新行。style_source_row 位于插入点之后时，
    # openpyxl 已将它下移一行；必须使用下移后的行，否则会复制到新空行的默认样式。
    style_src_after_insert = style_src + 1 if style_src >= insert_at else style_src
    _copy_row_style(ws, src_row=style_src_after_insert, dst_row=insert_at, max_col=ws.max_column)


def _normalize_reserved_row_styles(
    ws: Worksheet,
    start_row: int,
    reserved_rows: int,
    style_source_row: int,
) -> None:
    """把模板预留数据区统一成 style_source_row 的样式。"""
    for offset in range(reserved_rows):
        row = start_row + offset
        if row == style_source_row:
            continue
        _copy_row_style(ws, src_row=style_source_row, dst_row=row, max_col=ws.max_column)


def _copy_row_style(ws: Worksheet, src_row: int, dst_row: int, max_col: int) -> None:
    """把 src_row 的每个单元格样式复制到 dst_row 的同列单元格。"""
    src_height = ws.row_dimensions[src_row].height
    if src_height is not None:
        ws.row_dimensions[dst_row].height = src_height
    for col_idx in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col_idx)
        dst_cell = ws.cell(row=dst_row, column=col_idx)
        # MergedCell 不可设置样式，跳过；MergedCell 的样式由所属合并区域统一控制
        if not isinstance(src_cell, Cell) or not isinstance(dst_cell, Cell):
            continue
        if src_cell.has_style:
            # openpyxl 的 stub 把 .font 等声明为 Serialisable，
            # 但运行时 copy(StyleProxy) 返回相应的 Font/Fill/... 实例，赋值有效（Spike A 验证过）。
            dst_cell.font = copy(src_cell.font)  # type: ignore[assignment]
            dst_cell.fill = copy(src_cell.fill)  # type: ignore[assignment]
            dst_cell.border = copy(src_cell.border)  # type: ignore[assignment]
            dst_cell.alignment = copy(src_cell.alignment)  # type: ignore[assignment]
            dst_cell.number_format = src_cell.number_format
            dst_cell.protection = copy(src_cell.protection)  # type: ignore[assignment]


# —————————————————————————————————————
# Helpers
# —————————————————————————————————————


def _safe_set_cell(ws: Worksheet, addr: str, value: object) -> None:
    """安全写单元格，跳过 MergedCell（合并区域非左上角单元格不可写）。"""
    cell = ws[addr]
    if not isinstance(cell, Cell):
        return  # MergedCell 不可写，跳过
    cell.value = cast(Any, value)


def _write_line_cell(cell: Cell, value: object, spec: LineFieldSpec) -> None:
    cell.value = cast(Any, value)
    number_format = line_excel_number_format(value, spec)
    if number_format is not None:
        cell.number_format = number_format


def _content_extent(ws: Worksheet) -> tuple[int, int]:
    """返回有内容（含合并区）的最大行、列，忽略模板里未使用的空列维度。"""

    max_row = 1
    max_col = 1
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value is None:
                continue
            max_row = max(max_row, int(cell.row))
            max_col = max(max_col, int(cell.column))
    for merged in ws.merged_cells.ranges:
        if ws.cell(merged.min_row, merged.min_col).value is None:
            continue
        max_row = max(max_row, int(merged.max_row))
        max_col = max(max_col, int(merged.max_col))
    return max_row, max_col


def _apply_print_layout(ws: Worksheet) -> None:
    """按实际内容重设打印区并缩放到一页，避免 LibreOffice PDF 分页切断表头。"""

    max_row, max_col = _content_extent(ws)
    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None


def _packing_weight_location(
    doc_line: DocumentLine,
    key: str,
) -> tuple[str | None, str | None]:
    if key == "net_weight":
        return doc_line.net_weight_source_sheet, doc_line.net_weight_source_field
    if key == "gross_weight":
        return doc_line.gross_weight_source_sheet, doc_line.gross_weight_source_field
    return None, None


__all__ = ["render_document", "render_document_bundle"]
