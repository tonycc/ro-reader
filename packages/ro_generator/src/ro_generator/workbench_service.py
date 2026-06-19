"""工作台查询服务：面向工作台 UI 的聚合查询接口。

所有工作台状态的业务推导（PO 状态、月份候选等）集中在此模块，
确保 API 层只做序列化，不做业务判断。

缓存路径：inspect_workbook() 和 get_po_data() 复用 WorkbookSnapshot，
避免每次请求重复读取 Excel。
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from ro_generator.errors import WorkbookOpenError
from ro_generator.generator import generate
from ro_generator.models import DocumentRequest, DocumentType, GenerationResult, ValidationMessage
from ro_generator.packager import build_zip_filename, package_zip
from ro_generator.resolver import resolve_po_rows
from ro_generator.workbook_cache import get_cache_manager
from ro_generator.workbook_snapshot import BuildSnapshotError, PoInspection


@dataclass(frozen=True)
class FileInspectionResult:
    ok: bool
    sheets: tuple[str, ...] = ()
    size: int = 0
    error: str = ""


def inspect_file_path(path: str) -> FileInspectionResult:
    """检查文件路径是否有效，返回 sheet 列表和文件大小。

    API 层不应直接使用 openpyxl——此函数封装了文件检测逻辑。
    """
    from openpyxl import load_workbook

    p = Path(path)
    if not p.exists():
        return FileInspectionResult(ok=False, error=f"文件不存在：{path}")
    if not p.is_file():
        return FileInspectionResult(ok=False, error=f"路径不是文件：{path}")
    if p.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
        return FileInspectionResult(ok=False, error=f"不支持的文件格式：{p.suffix}")

    try:
        wb = load_workbook(str(p), read_only=True)
        sheets = tuple(wb.sheetnames)
        wb.close()
        return FileInspectionResult(ok=True, sheets=sheets, size=p.stat().st_size)
    except Exception as e:
        return FileInspectionResult(ok=False, error=f"无法打开文件：{e}")


@dataclass(frozen=True)
class WorkbookInspectionResult:
    ok: bool
    po_list: tuple[PoInspection, ...]
    errors: tuple[ValidationMessage, ...] = ()


@dataclass(frozen=True)
class ExportDocumentGroup:
    seller: str
    documents: tuple[DocumentType, ...]
    invoice_no: str | None = None


def inspect_workbook(base_file: str) -> WorkbookInspectionResult:
    """返回 PO 列表和状态摘要。优先复用缓存。"""
    try:
        cache = get_cache_manager()
        snapshot = cache.get_snapshot(base_file)
        return WorkbookInspectionResult(ok=True, po_list=snapshot.po_summary)
    except (WorkbookOpenError, BuildSnapshotError, OSError) as exc:
        msg = str(exc)
        message = getattr(exc, "message", None)
        if isinstance(message, str):
            msg = message
        code = getattr(exc, "code", "WORKBOOK_OPEN_ERROR")
        return WorkbookInspectionResult(
            ok=False,
            po_list=(),
            errors=(ValidationMessage(kind="blocking_error", code=code, message=msg),),
        )


def get_po_data(base_file: str, po_no: str) -> dict[str, object]:
    """返回指定 PO 的行数据和表头。优先复用缓存。

    返回 {"po_no": str, "headers": list, "rows": list}。
    """
    cache = get_cache_manager()
    snapshot = cache.get_snapshot(base_file)
    rows = snapshot.po_rows_for_po(po_no)
    return {
        "po_no": po_no,
        "headers": list(snapshot.headers_po_record),
        "rows": [dict(r) for r in rows],  # shallow copy for safe serialization
    }


def get_customer_po_data(base_file: str, po_no: str) -> dict[str, object]:
    """返回指定 PO 号对应的客户PO数据。优先复用缓存。

    客户PO sheet 按 Purchasing Document 索引，PO record 的 PO NO. 列
    就是客户PO 的 Purchasing Document。
    """
    cache = get_cache_manager()
    snapshot = cache.get_snapshot(base_file)
    rows = snapshot.customer_po_rows_for_po(po_no)
    return {
        "po_no": po_no,
        "headers": list(snapshot.headers_customer_po),
        "rows": [dict(r) for r in rows],
    }


def _message_to_dict(message: ValidationMessage) -> dict[str, object]:
    return {
        "kind": message.kind,
        "code": message.code,
        "message": message.message,
        "sheet": message.sheet,
        "row": message.row,
        "field": message.field,
        "severity": message.severity,
    }


def get_po_issues(base_file: str, po_no: str) -> dict[str, object]:
    """返回指定 PO 的阻断和警告明细。

    与 PO 列表状态使用同一条 resolver 路径，避免工作台前端/API 层重复业务判断。
    """
    cache = get_cache_manager()
    snapshot = cache.get_snapshot(base_file)
    rows = snapshot.po_rows_for_po(po_no)
    customer_rows = snapshot.customer_po_rows_for_po(po_no)
    result = resolve_po_rows(
        rows,
        snapshot.product_index,
        po_no=po_no,
        customer_po_rows=customer_rows,
    )
    blocking = [m for m in result.messages if m.kind == "blocking_error"]
    warnings = [m for m in result.messages if m.kind == "warning"]
    return {
        "po_no": po_no,
        "blocking_count": len(blocking),
        "warnings_count": len(warnings),
        "blocking_errors": [_message_to_dict(m) for m in blocking],
        "warnings": [_message_to_dict(m) for m in warnings],
    }


def export_document_groups(
    *,
    base_file: str,
    po_no: str,
    output_dir: str,
    groups: tuple[ExportDocumentGroup, ...],
) -> GenerationResult:
    """按主体批量导出，并把所有可生成文件合并成一个 ZIP。

    这是工作台导出确认页的核心入口：API 只负责把前端选择转成 groups，
    真正的单据生成、阻断错误和 ZIP 打包都留在核心包。
    """
    if not groups:
        return GenerationResult(
            status="error",
            errors=(
                ValidationMessage(
                    kind="blocking_error",
                    code="NO_EXPORT_DOCUMENTS",
                    message="未选择需要导出的单据",
                ),
            ),
        )

    output_root = Path(output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    rendered_files: list[Path] = []
    warnings: list[ValidationMessage] = []

    for index, group in enumerate(groups, start=1):
        group_dir = output_root / f"group-{index}"
        request = DocumentRequest(
            base_file=base_file,
            po_no=po_no,
            seller=group.seller,
            invoice_no=group.invoice_no,
            documents=group.documents,
            output_format="xlsx",
            output_dir=str(group_dir),
        )
        result = generate(request)
        warnings.extend(result.warnings)

        if result.status != "success":
            return GenerationResult(
                status=result.status,
                errors=result.errors,
                warnings=tuple(warnings),
                missing_inputs=result.missing_inputs,
                options=result.options,
            )

        rendered_files.extend(_generated_file_paths(result))

    zip_path = package_zip(
        files=tuple(rendered_files),
        output_dir=output_root,
        zip_name=build_zip_filename(po_no=po_no),
    )
    return GenerationResult(
        status="success",
        summary={
            "po_no": po_no,
            "groups": [
                {"seller": group.seller, "documents": list(group.documents)} for group in groups
            ],
        },
        files=tuple(path.name for path in rendered_files),
        output_file=str(zip_path),
        warnings=tuple(warnings),
    )


def _generated_file_paths(result: GenerationResult) -> tuple[Path, ...]:
    if result.output_file is None:
        return ()
    output = Path(result.output_file)
    if output.is_dir():
        return tuple(output / filename for filename in result.files)
    return (output,)


__all__ = [
    "ExportDocumentGroup",
    "FileInspectionResult",
    "PoInspection",
    "WorkbookInspectionResult",
    "export_document_groups",
    "get_customer_po_data",
    "get_po_data",
    "get_po_issues",
    "inspect_file_path",
    "inspect_workbook",
]
