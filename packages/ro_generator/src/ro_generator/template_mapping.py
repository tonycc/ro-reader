"""Template mapping：把 YAML 字段映射加载成结构化对象，并校验对模板文件的引用。

设计边界：
- mapping 是模板的"字段位置说明书"。模板维护者只改 YAML，不改代码。
- mapping 加载时校验所有引用的单元格在模板中真实存在（产品方案 §13.2）。
  这道防线让"模板换了但 mapping 没改"的漂移立即暴露，而不是悄悄写错单元格。
- template_version 是必填字段。模板版本改了就必须更新 mapping，否则加载失败。
"""

from __future__ import annotations

from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Final, cast

import yaml
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet

from ro_generator.errors import MappingError, TemplateError
from ro_generator.models import DocumentType
from ro_generator.resources import find_profile_root

# —————————————————————————————————————
# 数据结构
# —————————————————————————————————————


@dataclass(frozen=True)
class LineColumns:
    """订单行各业务字段对应的模板列字母。"""

    quantity: str
    unit_price: str | None = None
    amount: str | None = None
    po_no: str | None = None
    item_number: str | None = None
    item_line_no: str | None = None
    sap: str | None = None
    description: str | None = None
    gs_model: str | None = None
    unit_label: str | None = None
    net_weight: str | None = None  # PL 专用
    gross_weight: str | None = None  # PL 专用
    carton_count: str | None = None  # PL 专用
    carton_from: str | None = None  # PL 箱号起
    carton_to: str | None = None  # PL 箱号止
    cbm: str | None = None  # PL 专用
    length: str | None = None  # PL 专用
    width: str | None = None  # PL 专用
    height: str | None = None  # PL 专用
    confirmed_ex_factory_date: str | None = None  # EMAX PI 专用
    extras: dict[str, str] = field(default_factory=dict)  # 未来扩展列，免改代码


@dataclass(frozen=True)
class LinesSection:
    start_row: int
    style_source_row: int
    columns: LineColumns
    unit_label: str | None = None  # 固定写到 unit_label 列的文案
    row_fixed: dict[str, str] = field(default_factory=dict)  # 每行固定值 {列字母: 值}


@dataclass(frozen=True)
class CostBreakdownSection:
    """可选的 Combo 成本拆分表区域。"""

    start_row: int
    style_source_row: int
    columns: dict[str, str]
    reserved_rows: int = 1


@dataclass(frozen=True)
class CellStyles:
    """可选的单元格样式声明（来自 mapping YAML 的 style 节）。"""

    bold: tuple[str, ...] = ()
    underline: tuple[str, ...] = ()


@dataclass(frozen=True)
class TotalCell:
    """totals 节中的单元格声明。

    - `model_total`: 取 DocumentModel 的标准合计字段（兼容旧写法 `amount: H27`）
    - `fixed`: 写固定文案
    - `current_date`: 写当天日期字符串
    - `model_date`: 写 DocumentModel.document_date
    """

    cell: str
    value_mode: str = "model_total"
    value: str | None = None


@dataclass(frozen=True)
class TemplateMapping:
    """单份 mapping 加载后的不可变结构。"""

    document: DocumentType
    template_version: str
    template_path: Path
    sheet: str
    header: dict[str, str]
    lines: LinesSection
    cost_breakdown: CostBreakdownSection | None
    totals: dict[str, TotalCell]
    notes: dict[str, str]
    table_header_row: list[int] = field(default_factory=list)
    header_fixed: dict[str, str] = field(default_factory=dict)
    style: CellStyles = field(default_factory=CellStyles)
    preview_content: dict[str, object] = field(default_factory=dict)
    preview_column_labels: tuple[tuple[str, str], ...] = ()
    preview_column_letters: dict[str, str] = field(default_factory=dict)
    # 可选的多行/合并表头结构，由模板的真实 merged_cells 解析得到。
    preview_header_rows: tuple[tuple[dict[str, object], ...], ...] = ()
    preview_header_labels: dict[str, str] = field(default_factory=dict)
    preview_static_values: dict[str, tuple[str, ...]] = field(default_factory=dict)


# —————————————————————————————————————
# 加载入口
# —————————————————————————————————————


VALID_DOCUMENT_TYPES: Final[set[str]] = {"PI", "PO", "INVOICE", "PL", "CI", "RO_PL"}
VALID_TOTAL_VALUE_MODES: Final[set[str]] = {
    "model_total",
    "fixed",
    "current_date",
    "model_date",
}
TEMPLATE_FOOTER_ROW_SLACK: Final[int] = 32
PREVIEW_CONTINUATION_HEADER_FIELDS: Final[frozenset[str]] = frozenset(
    {
        "bill_to_line2",
        "bill_to_line3",
        "ship_to_line2",
        "ship_to_line3",
        "manufacturer_address_2",
        "shipping_mark_2",
        "shipping_mark_3",
    }
)


def load_template_mapping(yaml_path: str | Path) -> TemplateMapping:
    """从 YAML 文件加载 mapping，并校验对应模板文件的所有引用。

    YAML 加载失败抛 `MappingError`；模板加载失败抛 `TemplateError`。
    引用校验失败抛 `MappingError`，错误消息会指出具体哪个字段错了。
    """
    yaml_path = Path(yaml_path)
    if not yaml_path.exists():
        raise MappingError(f"mapping 文件不存在：{yaml_path}")
    raw = _read_yaml(yaml_path)

    mapping = _parse_mapping(raw, yaml_path)
    (
        preview_column_labels,
        preview_column_letters,
        preview_header_rows,
        preview_header_labels,
        preview_static_values,
    ) = _validate_against_template(mapping)
    return replace(
        mapping,
        preview_column_labels=preview_column_labels,
        preview_column_letters=preview_column_letters,
        preview_header_rows=preview_header_rows,
        preview_header_labels=preview_header_labels,
        preview_static_values=preview_static_values,
    )


# —————————————————————————————————————
# YAML 解析
# —————————————————————————————————————


def _read_yaml(yaml_path: Path) -> dict[str, object]:
    try:
        with yaml_path.open(encoding="utf-8") as fp:
            data = yaml.safe_load(fp)
    except yaml.YAMLError as exc:
        raise MappingError(f"YAML 解析失败 {yaml_path}：{exc}") from exc
    if not isinstance(data, dict):
        raise MappingError(f"mapping 文件根节点必须是 dict：{yaml_path}")
    return data


def _parse_mapping(raw: dict[str, object], yaml_path: Path) -> TemplateMapping:
    document_raw = _require_str(raw, "document", yaml_path)
    document = _normalize_document_type(document_raw, yaml_path)
    template_version = _require_str(raw, "template_version", yaml_path)
    template_rel = _require_str(raw, "template", yaml_path)
    sheet_raw = raw.get("sheet")
    if not isinstance(sheet_raw, str) or not sheet_raw.strip():
        raise MappingError(f"mapping 缺少必填字段 'sheet' 或值非字符串：{yaml_path}")
    sheet = sheet_raw

    # 模板路径相对当前 Profile 根解析。绝对路径仍用于临时 mapping/测试 fixture；
    # 没有 profile.yaml 标记的临时相对 mapping 则相对 mapping 文件所在目录解析。
    template_value = Path(template_rel)
    if template_value.is_absolute():
        template_path = template_value.resolve()
    else:
        profile_root = find_profile_root(yaml_path)
        base_path = profile_root or yaml_path.parent
        template_path = (base_path / template_value).resolve()
    if not template_path.exists():
        raise MappingError(
            f"模板文件不存在：{template_rel}（相对 {yaml_path} 解析为 {template_path}）"
        )

    header = _require_dict_of_str(raw, "header", yaml_path)
    header_fixed = _optional_dict_of_str(raw, "header_fixed")
    lines = _parse_lines_section(raw, yaml_path)
    cost_breakdown = _parse_cost_breakdown_section(raw, yaml_path)
    totals = _parse_totals_section(raw, yaml_path)
    notes = _optional_dict_of_str(raw, "notes")
    table_header_row = _parse_table_header_rows(raw, "table_header_row", yaml_path)
    for r in table_header_row:
        if r >= lines.start_row:
            raise MappingError(
                f"table_header_row 每一项都必须小于 lines.start_row={lines.start_row}，但 {r} >= {lines.start_row}（{yaml_path}）"
            )
    style = _parse_style(raw.get("style"), yaml_path)
    preview_content = _parse_preview_content(raw.get("preview_content"), yaml_path)

    return TemplateMapping(
        document=document,
        template_version=template_version,
        template_path=template_path,
        sheet=sheet,
        header=header,
        lines=lines,
        cost_breakdown=cost_breakdown,
        totals=totals,
        notes=notes,
        table_header_row=table_header_row,
        header_fixed=header_fixed,
        style=style,
        preview_content=preview_content,
    )


def _normalize_document_type(value: str, yaml_path: Path) -> DocumentType:
    upper = value.strip().upper()
    if upper not in VALID_DOCUMENT_TYPES:
        raise MappingError(
            f"document 字段必须是 {sorted(VALID_DOCUMENT_TYPES)} 之一，得到 {value!r}（{yaml_path}）"
        )
    # mypy: 此处 upper 在白名单内，但 DocumentType 是 Literal，需要 cast
    return upper  # type: ignore[return-value]


def _parse_lines_section(raw: dict[str, object], yaml_path: Path) -> LinesSection:
    lines = raw.get("lines")
    if not isinstance(lines, dict):
        raise MappingError(f"mapping 必须有 lines 节（dict）：{yaml_path}")
    start_row = _require_int(lines, "start_row", yaml_path, ctx="lines")
    style_source_row = _require_int(lines, "style_source_row", yaml_path, ctx="lines")
    if start_row <= 0 or style_source_row <= 0:
        raise MappingError(f"lines.start_row / lines.style_source_row 必须为正整数（{yaml_path}）")

    columns_raw = lines.get("columns")
    if not isinstance(columns_raw, dict):
        raise MappingError(f"mapping 必须有 lines.columns 节（dict）：{yaml_path}")
    columns = _parse_columns(columns_raw, yaml_path)

    unit_label_default = lines.get("unit_label")
    unit_label = unit_label_default if isinstance(unit_label_default, str) else None

    # Parse row_fixed: per-row fixed values {column_letter: value}
    row_fixed: dict[str, str] = {}
    rf = lines.get("row_fixed")
    if isinstance(rf, dict):
        for col_letter, val in rf.items():
            if isinstance(col_letter, str) and isinstance(val, str):
                col_letter = col_letter.strip().upper()
                try:
                    column_index_from_string(col_letter)
                except (ValueError, KeyError) as exc:
                    raise MappingError(
                        f"lines.row_fixed 列字母 {col_letter!r} 不合法：{yaml_path}"
                    ) from exc
                row_fixed[col_letter] = val

    return LinesSection(
        start_row=start_row,
        style_source_row=style_source_row,
        columns=columns,
        unit_label=unit_label,
        row_fixed=row_fixed,
    )


def _parse_cost_breakdown_section(
    raw: dict[str, object],
    yaml_path: Path,
) -> CostBreakdownSection | None:
    section = raw.get("cost_breakdown")
    if section is None:
        return None
    if not isinstance(section, dict):
        raise MappingError(f"mapping.cost_breakdown 必须是 dict：{yaml_path}")
    start_row = _require_int(section, "start_row", yaml_path, ctx="cost_breakdown")
    style_source_row = _require_int(section, "style_source_row", yaml_path, ctx="cost_breakdown")
    if start_row <= 0 or style_source_row <= 0:
        raise MappingError(f"cost_breakdown 行号必须为正整数：{yaml_path}")
    reserved_rows = section.get("reserved_rows", 1)
    if not isinstance(reserved_rows, int) or isinstance(reserved_rows, bool) or reserved_rows < 1:
        raise MappingError(f"cost_breakdown.reserved_rows 必须是正整数：{yaml_path}")
    columns_raw = section.get("columns")
    if not isinstance(columns_raw, dict) or not columns_raw:
        raise MappingError(f"cost_breakdown.columns 必须是非空 dict：{yaml_path}")
    columns: dict[str, str] = {}
    for key, value in columns_raw.items():
        if not isinstance(key, str) or not isinstance(value, str):
            continue
        letter = value.strip().upper()
        try:
            column_index_from_string(letter)
        except (ValueError, KeyError) as exc:
            raise MappingError(
                f"cost_breakdown.columns.{key} {letter!r} 不是合法列字母：{yaml_path}"
            ) from exc
        columns[key] = letter
    return CostBreakdownSection(
        start_row=start_row,
        style_source_row=style_source_row,
        columns=columns,
        reserved_rows=reserved_rows,
    )


def _parse_columns(raw: dict[object, object], yaml_path: Path) -> LineColumns:
    required = {"quantity"}
    known = {
        "po_no",
        "item_number",
        "item_line_no",
        "description",
        "gs_model",
        "unit_label",
        "sap",
        "unit_price",
        "amount",
        "net_weight",
        "gross_weight",
        "carton_count",
        "carton_from",
        "carton_to",
        "cbm",
        "length",
        "width",
        "height",
        "confirmed_ex_factory_date",
    }

    parsed: dict[str, str] = {}
    extras: dict[str, str] = {}

    for key, value in raw.items():
        if not isinstance(key, str) or not isinstance(value, str):
            continue
        letter = value.strip().upper()
        try:
            column_index_from_string(letter)
        except (ValueError, KeyError) as exc:
            raise MappingError(
                f"lines.columns.{key} {letter!r} 不是合法列字母：{yaml_path}"
            ) from exc
        if key in known or key in required:
            parsed[key] = letter
        else:
            extras[key] = letter

    for key in required:
        if key not in parsed:
            raise MappingError(f"lines.columns 缺少必填项 {key!r}：{yaml_path}")

    return LineColumns(extras=extras, **parsed)


def _parse_totals_section(raw: dict[str, object], yaml_path: Path) -> dict[str, TotalCell]:
    totals_raw = raw.get("totals")
    if not isinstance(totals_raw, dict):
        raise MappingError(f"mapping 缺少必填节 'totals'（必须是 dict）：{yaml_path}")

    parsed: dict[str, TotalCell] = {}
    for key, value in totals_raw.items():
        if not isinstance(key, str):
            raise MappingError(f"totals 的键必须是字符串：{yaml_path}")
        if isinstance(value, str):
            parsed[key] = TotalCell(cell=value.strip())
            continue
        if not isinstance(value, dict):
            raise MappingError(
                f"totals.{key} 必须是字符串单元格地址，或包含 cell/value_mode 的 dict：{yaml_path}"
            )

        cell_raw = value.get("cell")
        if not isinstance(cell_raw, str) or not cell_raw.strip():
            raise MappingError(f"totals.{key}.cell 必须是非空字符串：{yaml_path}")

        mode_raw = value.get("value_mode", "model_total")
        if not isinstance(mode_raw, str) or mode_raw not in VALID_TOTAL_VALUE_MODES:
            raise MappingError(
                f"totals.{key}.value_mode 必须是 {sorted(VALID_TOTAL_VALUE_MODES)} 之一：{yaml_path}"
            )

        fixed_value: str | None = None
        if mode_raw == "fixed":
            value_raw = value.get("value")
            if not isinstance(value_raw, str):
                raise MappingError(
                    f"totals.{key}.value_mode=fixed 时必须提供字符串 value：{yaml_path}"
                )
            fixed_value = value_raw

        parsed[key] = TotalCell(
            cell=cell_raw.strip(),
            value_mode=mode_raw,
            value=fixed_value,
        )
    return parsed


# —————————————————————————————————————
# 模板引用校验（产品方案 §13.2）
# —————————————————————————————————————


def _validate_against_template(
    mapping: TemplateMapping,
) -> tuple[
    tuple[tuple[str, str], ...],
    dict[str, str],
    tuple[tuple[dict[str, object], ...], ...],
    dict[str, str],
    dict[str, tuple[str, ...]],
]:
    """打开模板，校验所有 mapping 引用的单元格在模板中存在且可写。"""
    try:
        wb = load_workbook(filename=str(mapping.template_path), read_only=False)
    except Exception as exc:
        raise TemplateError(f"无法打开模板 {mapping.template_path}：{exc}") from exc

    try:
        if mapping.sheet not in wb.sheetnames:
            raise MappingError(
                f"模板 {mapping.template_path.name} 中找不到 sheet {mapping.sheet!r}"
            )
        ws = wb[mapping.sheet]
        max_row = ws.max_row
        max_col = ws.max_column

        # 1. header 单元格
        for field_name, addr in mapping.header.items():
            _check_cell_reference(addr, max_row, max_col, f"header.{field_name}")

        # 2. lines.style_source_row 必须 ≤ max_row
        if mapping.lines.style_source_row > max_row:
            raise MappingError(
                f"lines.style_source_row={mapping.lines.style_source_row} 超过模板行数 {max_row}"
            )
        # start_row 可以 ≤ style_source_row（先填后插），不强制 ≤ max_row。
        # 但 start_row 列上的样式参考还是要在模板里存在 → 通过 style_source_row 间接保证。

        # 3. 行列字母在模板列范围内
        for col_name, col_letter in iter_line_columns(mapping.lines.columns):
            col_idx = column_index_from_string(col_letter)
            if col_idx > max_col:
                raise MappingError(
                    f"lines.columns.{col_name} 列 {col_letter} (idx={col_idx}) "
                    f"超过模板列数 {max_col}"
                )

        # 3. 可选 Combo 成本拆分表
        if mapping.cost_breakdown is not None:
            if mapping.cost_breakdown.style_source_row > max_row:
                raise MappingError(
                    "cost_breakdown.style_source_row="
                    f"{mapping.cost_breakdown.style_source_row} 超过模板行数 {max_row}"
                )
            for col_name, col_letter in mapping.cost_breakdown.columns.items():
                col_idx = column_index_from_string(col_letter)
                if col_idx > max_col:
                    raise MappingError(
                        f"cost_breakdown.columns.{col_name} 列 {col_letter} "
                        f"(idx={col_idx}) 超过模板列数 {max_col}"
                    )

        # 4. totals 单元格
        for field_name, spec in mapping.totals.items():
            _check_cell_reference(
                spec.cell,
                max_row,
                max_col,
                f"totals.{field_name}",
                row_slack=TEMPLATE_FOOTER_ROW_SLACK,
            )

        # 5. notes 单元格
        for field_name, addr in mapping.notes.items():
            _check_cell_reference(addr, max_row, max_col, f"notes.{field_name}")

        # 6. 可选表格表头行
        for r in mapping.table_header_row:
            if r > max_row:
                raise MappingError(
                    f"table_header_row 每一项都不能超过模板行数 {max_row}，但 {r} > {max_row}"
                )
        (
            preview_column_labels,
            preview_column_letters,
            preview_header_rows,
        ) = _resolve_preview_columns(mapping, ws)
        return (
            preview_column_labels,
            preview_column_letters,
            preview_header_rows,
            _resolve_preview_header_labels(mapping, ws),
            _resolve_preview_static_values(mapping, ws, max_row=max_row, max_col=max_col),
        )
    finally:
        wb.close()


def _resolve_preview_header_labels(
    mapping: TemplateMapping,
    ws: Worksheet,
) -> dict[str, str]:
    """从每个 header 值单元格左侧读取模板实际显示的字段标签。"""

    labels: dict[str, str] = {}
    for field_name, address in mapping.header.items():
        if field_name in PREVIEW_CONTINUATION_HEADER_FIELDS:
            labels[field_name] = ""
            continue
        column_letter, row = coordinate_from_string(address.strip())
        value_column = column_index_from_string(column_letter)
        label = ""
        for column in range(value_column - 1, 0, -1):
            raw = _template_cell_value(ws, row, column)
            if raw is None:
                continue
            text = " ".join(str(raw).split())
            if text:
                label = text
                break
        labels[field_name] = label
    return labels


def _resolve_preview_static_values(
    mapping: TemplateMapping,
    ws: Worksheet,
    *,
    max_row: int,
    max_col: int,
) -> dict[str, tuple[str, ...]]:
    """按坐标读取标题和公司抬头等模板固定文本，避免在 YAML 复制文案。"""

    configured = mapping.preview_content.get("template_fields")
    if configured is None:
        return {}
    if not isinstance(configured, dict):
        raise MappingError(f"preview_content.template_fields 必须是 dict：{mapping.template_path}")

    resolved: dict[str, tuple[str, ...]] = {}
    for raw_key, raw_addresses in configured.items():
        if not isinstance(raw_key, str) or not raw_key.strip():
            raise MappingError(
                f"preview_content.template_fields 的键必须是非空字符串：{mapping.template_path}"
            )
        if isinstance(raw_addresses, str):
            addresses = [raw_addresses]
        elif (
            isinstance(raw_addresses, list)
            and raw_addresses
            and all(isinstance(item, str) for item in raw_addresses)
        ):
            addresses = cast(list[str], raw_addresses)
        else:
            raise MappingError(
                "preview_content.template_fields 的值必须是单元格地址或非空地址列表："
                f"{mapping.template_path}"
            )

        values: list[str] = []
        for address in addresses:
            _check_cell_reference(
                address,
                max_row,
                max_col,
                f"preview_content.template_fields.{raw_key}",
            )
            column_letter, row = coordinate_from_string(address.strip())
            raw = _template_cell_value(
                ws,
                row,
                column_index_from_string(column_letter),
            )
            text = " ".join(str(raw).split()) if raw is not None else ""
            if not text:
                raise MappingError(
                    f"preview_content.template_fields.{raw_key} 引用空单元格 {address}："
                    f"{mapping.template_path}"
                )
            values.append(text)
        resolved[raw_key.strip()] = tuple(values)
    return resolved


def _resolve_preview_columns(
    mapping: TemplateMapping,
    ws: Worksheet,
) -> tuple[
    tuple[tuple[str, str], ...],
    dict[str, str],
    tuple[tuple[dict[str, object], ...], ...],
]:
    """以模板表头为预览列文案事实源，同时保留 YAML 声明的列选择和顺序。"""

    configured = mapping.preview_content.get("column_labels")
    if not isinstance(configured, dict) or not configured:
        return (), {}, ()

    header_rows = _preview_column_label_rows(mapping)
    line_columns = dict(iter_line_columns(mapping.lines.columns))
    preview_letters: dict[str, str] = {}
    unresolved_keys: list[str] = []

    for key in configured:
        if not isinstance(key, str):
            continue
        if key in line_columns:
            preview_letters[key] = line_columns[key]
        elif key in mapping.lines.row_fixed:
            preview_letters[key] = key
        else:
            unresolved_keys.append(key)

    fixed_columns = [
        column for column in mapping.lines.row_fixed if column not in preview_letters.values()
    ]
    if len(unresolved_keys) == 1 and len(fixed_columns) == 1:
        preview_letters[unresolved_keys.pop()] = fixed_columns[0]
    if unresolved_keys:
        joined = ", ".join(unresolved_keys)
        raise MappingError(
            f"preview_content.column_labels 中的键无法对应到模板明细列：{joined} "
            f"（{mapping.template_path}）"
        )

    resolved: list[tuple[str, str]] = []
    for raw_key, raw_fallback in configured.items():
        if not isinstance(raw_key, str):
            continue
        fallback = raw_fallback if isinstance(raw_fallback, str) else raw_key
        column = preview_letters[raw_key]
        label = _template_header_label(ws, column, header_rows)
        resolved.append((raw_key, label if header_rows else fallback))
    header_rows_structure = (
        _resolve_preview_header_rows(
            ws,
            header_rows,
            preview_letters,
            configured,
        )
        if mapping.preview_content.get("merged_headers") is True
        else ()
    )
    return tuple(resolved), preview_letters, header_rows_structure


def _resolve_preview_header_rows(
    ws: Worksheet,
    header_rows: list[int],
    preview_letters: dict[str, str],
    configured: dict[object, object],
) -> tuple[tuple[dict[str, object], ...], ...]:
    """把模板真实合并表头转换为前端可渲染的 rowspan/colspan 结构。

    `column_labels` 仍保留叶子列和来源索引契约；本结构只负责表头布局，
    因而不会把 Excel 中的 `D18:E19`、`J18:L18` 等合并关系拆开。
    """

    if not header_rows:
        return ()

    ordered_keys = [key for key in configured if isinstance(key, str) and key in preview_letters]
    if not ordered_keys:
        return ()
    ordered_columns = [
        (key, column_index_from_string(preview_letters[key])) for key in ordered_keys
    ]
    merged_ranges = tuple(ws.merged_cells.ranges)
    occupied: set[tuple[int, int]] = set()
    output: list[tuple[dict[str, object], ...]] = []

    for row_index, row in enumerate(header_rows):
        cells: list[dict[str, object]] = []
        for column_index, (key, column) in enumerate(ordered_columns):
            if (row_index, column_index) in occupied:
                continue

            merged = next(
                (
                    cell_range
                    for cell_range in merged_ranges
                    if cell_range.min_row <= row <= cell_range.max_row
                    and cell_range.min_col <= column <= cell_range.max_col
                ),
                None,
            )
            if merged is not None:
                mapped_positions = [
                    index
                    for index, (_, mapped_column) in enumerate(ordered_columns)
                    if merged.min_col <= mapped_column <= merged.max_col
                ]
                mapped_row_indices = [
                    index
                    for index, mapped_row in enumerate(header_rows)
                    if merged.min_row <= mapped_row <= merged.max_row
                ]
                if not mapped_positions or not mapped_row_indices:
                    continue
                if column_index != mapped_positions[0] or row_index != mapped_row_indices[0]:
                    continue

                for occupied_row in mapped_row_indices:
                    for occupied_column in mapped_positions:
                        occupied.add((occupied_row, occupied_column))
                cell: dict[str, object] = {
                    "label": _template_cell_text(ws, merged.min_row, merged.min_col),
                    "colspan": len(mapped_positions),
                    "rowspan": len(mapped_row_indices),
                }
                if len(mapped_positions) == 1 and len(mapped_row_indices) == 1:
                    cell["key"] = key
                cells.append(cell)
                continue

            cells.append(
                {
                    "key": key,
                    "label": _template_cell_text(ws, row, column),
                    "colspan": 1,
                    "rowspan": 1,
                }
            )
        output.append(tuple(cells))
    return tuple(output)


def _template_cell_text(ws: Worksheet, row: int, column: int) -> str:
    raw = _template_cell_value(ws, row, column)
    return " ".join(str(raw).split()) if raw is not None else ""


def _preview_column_label_rows(mapping: TemplateMapping) -> list[int]:
    raw = mapping.preview_content.get("column_label_rows")
    if raw is None:
        return list(mapping.table_header_row)
    if isinstance(raw, int) and not isinstance(raw, bool):
        rows = [raw]
    elif (
        isinstance(raw, list)
        and raw
        and all(isinstance(item, int) and not isinstance(item, bool) for item in raw)
    ):
        rows = cast(list[int], raw)
    else:
        raise MappingError(
            f"preview_content.column_label_rows 必须是正整数或非空正整数列表："
            f"{mapping.template_path}"
        )
    if any(row <= 0 for row in rows):
        raise MappingError(
            f"preview_content.column_label_rows 必须全部为正整数：{mapping.template_path}"
        )
    undeclared = [row for row in rows if row not in mapping.table_header_row]
    if undeclared:
        raise MappingError(
            "preview_content.column_label_rows 必须是 table_header_row 的子集，"
            f"未声明行：{undeclared}（{mapping.template_path}）"
        )
    return rows


def _template_header_label(ws: Worksheet, column: str, rows: list[int]) -> str:
    column_index = column_index_from_string(column)
    parts: list[str] = []
    for row in rows:
        value = _template_cell_value(ws, row, column_index)
        if value is None:
            continue
        text = " ".join(str(value).split())
        if text and (not parts or parts[-1] != text):
            parts.append(text)
    return "\n".join(parts)


def _template_cell_value(ws: Worksheet, row: int, column: int) -> object | None:
    return ws.cell(row=row, column=column).value


def _check_cell_reference(
    addr: str,
    max_row: int,
    max_col: int,
    ctx: str,
    *,
    row_slack: int = 0,
) -> None:
    try:
        col_letter, row = coordinate_from_string(addr.strip())
    except ValueError as exc:
        raise MappingError(f"{ctx} 单元格地址不合法：{addr!r}") from exc
    col_idx = column_index_from_string(col_letter)
    allowed_max_row = max_row + row_slack
    if row > allowed_max_row or col_idx > max_col:
        raise MappingError(
            f"{ctx} 单元格 {addr} 超出模板范围（max_row={allowed_max_row}, max_col={max_col}）"
        )


def iter_line_columns(cols: LineColumns) -> list[tuple[str, str]]:
    """按固定顺序迭代所有列（已知键 + extras），返回 [(key, column_letter), ...]."""
    pairs: list[tuple[str, str]] = []
    for key in (
        "po_no",
        "item_number",
        "item_line_no",
        "description",
        "gs_model",
        "sap",
        "quantity",
        "unit_price",
        "amount",
        "unit_label",
        "carton_from",
        "carton_to",
        "net_weight",
        "gross_weight",
        "carton_count",
        "cbm",
        "length",
        "width",
        "height",
        "confirmed_ex_factory_date",
    ):
        v = getattr(cols, key)
        if isinstance(v, str):
            pairs.append((key, v))
    # 追加 YAML 中声明的扩展列（这些列在 LineColumns 中没有独立字段）
    for key, col_letter in cols.extras.items():
        pairs.append((key, col_letter))
    return pairs


# —————————————————————————————————————
# 小工具
# —————————————————————————————————————


def _require_str(raw: dict[str, object], key: str, yaml_path: Path) -> str:
    v = raw.get(key)
    if not isinstance(v, str) or not v.strip():
        raise MappingError(f"mapping 缺少必填字段 {key!r} 或值非字符串：{yaml_path}")
    return v.strip()


def _require_int(raw: dict[object, object], key: str, yaml_path: Path, ctx: str) -> int:
    v = raw.get(key)
    if not isinstance(v, int) or isinstance(v, bool):
        raise MappingError(f"{ctx}.{key} 必须为整数：{yaml_path}")
    return v


def _optional_positive_int(raw: dict[str, object], key: str, yaml_path: Path) -> int | None:
    value = raw.get(key)
    if value is None:
        return None
    if not isinstance(value, int) or isinstance(value, bool) or value <= 0:
        raise MappingError(f"{key} 必须为正整数（如有）：{yaml_path}")
    return value


def _parse_table_header_rows(raw: dict[str, object], key: str, yaml_path: Path) -> list[int]:
    """解析 table_header_row，支持单行 int、多行 list[int] 或省略（返回空列表）。"""
    value = raw.get(key)
    if value is None:
        return []
    if isinstance(value, int) and not isinstance(value, bool):
        if value <= 0:
            raise MappingError(f"{key} 必须为正整数：{yaml_path}")
        return [value]
    if isinstance(value, list):
        result: list[int] = []
        for i, item in enumerate(value):
            if not isinstance(item, int) or isinstance(item, bool) or item <= 0:
                raise MappingError(f"{key}[{i}] 必须为正整数：{yaml_path}")
            result.append(item)
        if not result:
            raise MappingError(f"{key} 不能为空列表（如有）：{yaml_path}")
        return result
    raise MappingError(f"{key} 必须为正整数或正整数列表（如有）：{yaml_path}")


def _require_dict_of_str(raw: dict[str, object], key: str, yaml_path: Path) -> dict[str, str]:
    v = raw.get(key)
    if not isinstance(v, dict):
        raise MappingError(f"mapping 缺少必填节 {key!r}（必须是 dict）：{yaml_path}")
    out: dict[str, str] = {}
    for k, val in v.items():
        if not isinstance(k, str) or not isinstance(val, str):
            raise MappingError(f"{key}.{k!r} 的键和值都必须是字符串：{yaml_path}")
        out[k] = val.strip()
    return out


def _parse_preview_content(raw: object, yaml_path: Path) -> dict[str, object]:
    """解析 preview_content 节，在 load 阶段与其余 mapping 一同校验。

    preview_content 由 document_preview.build_preview() 消费，
    不再需要单独重新读取 YAML 文件。
    """
    if raw is None:
        return {}
    if not isinstance(raw, dict):
        raise MappingError(f"preview_content 必须是 dict：{yaml_path}")
    return cast(dict[str, object], raw)


def _parse_style(raw: object, yaml_path: Path) -> CellStyles:
    """解析可选的 style 节。格式：
    style:
      bold: [A1, B4]        # 加粗的单元格地址
      underline: [H6, F6]   # 下划线的单元格地址
    """
    if not isinstance(raw, dict):
        return CellStyles()
    bold = _parse_cell_list(raw.get("bold"), yaml_path)
    underline = _parse_cell_list(raw.get("underline"), yaml_path)
    return CellStyles(bold=tuple(bold), underline=tuple(underline))


def _parse_cell_list(raw: object, yaml_path: Path) -> list[str]:
    if raw is None:
        return []
    if not isinstance(raw, list):
        raise MappingError(f"style 中的值必须是列表：{raw!r} ({yaml_path})")
    result: list[str] = []
    for item in raw:
        if not isinstance(item, str):
            raise MappingError(f"style 列表项必须是字符串：{item!r} ({yaml_path})")
        result.append(item.strip().upper())
    return result


def _optional_dict_of_str(raw: dict[str, object], key: str) -> dict[str, str]:
    v = raw.get(key)
    if v is None:
        return {}
    if not isinstance(v, dict):
        raise MappingError(f"mapping 中 {key!r} 节必须是 dict（如有）")
    out: dict[str, str] = {}
    for k, val in v.items():
        if not isinstance(k, str) or not isinstance(val, str):
            raise MappingError(f"{key}.{k!r} 的键和值都必须是字符串")
        out[k] = val.strip()
    return out


__all__ = [
    "VALID_DOCUMENT_TYPES",
    "LineColumns",
    "LinesSection",
    "TemplateMapping",
    "TotalCell",
    "load_template_mapping",
]
