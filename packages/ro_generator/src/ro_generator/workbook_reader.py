"""Workbook reader：把 base.xlsx 解析成结构化的 sheet 数据。

设计原则：
- 只读，绝不向 base 文件写回。
- `data_only=True`：公式列返回最近一次保存时的缓存值。读到 None 时由 resolver
  按产品方案 §10.4 的公式回退现算，reader 这一层不做计算。
- 不做语义校验：reader 关心结构（sheet 是否能打开、表头第几行、空白行如何跳过），
  validator 关心规则（必需表头是否齐、PO 是否存在等）。
- 行号一律 **1-based**（与 openpyxl 一致），dict 里以 `__row_number__` 暴露给下游。

每行返回 dict 而非 tuple，是为了把"列号"这个 Excel 细节限制在 reader 之内：
下游用规范化表头名取数据，例如 `row["SAP Number"]`，迁移列布局时只需改 schema。
"""

from __future__ import annotations

import re
from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path
from types import TracebackType
from typing import Final

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ro_generator.base_schema import base_schema
from ro_generator.errors import WorkbookOpenError
from ro_generator.schema import FIRST_DATA_ROW, HEADER_ROW, normalize_header

ROW_NUMBER_KEY: Final = "__row_number__"
CELL_NUMBER_FORMATS_KEY: Final = "__cell_number_formats__"
_BASE_SCHEMA = base_schema()


@dataclass(frozen=True)
class SheetHeaders:
    """单个 sheet 的表头解析结果，不读取数据行。"""

    sheet_name: str
    headers: tuple[str, ...]
    header_columns: dict[str, int]
    max_data_col: int


@dataclass(frozen=True)
class SheetData:
    """单个 sheet 的解析结果。

    `headers` 与 `header_columns.keys()` 是同一份信息的两种视图，前者保留顺序。
    重复表头时**首次出现的列号胜出**（resolver 层若需要重复检查，可比较 length）。
    """

    sheet_name: str
    headers: tuple[str, ...]
    header_columns: dict[str, int]  # normalized header → 1-based column index
    rows: tuple[dict[str, object], ...]


class WorkbookReader:
    """对 openpyxl 的薄包装，按 base 文件的固定布局读三张 sheet。

    用法：

        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
            for row in data.rows:
                print(row["SAP Number"], row["FINALQTY"])

    也可以不用 context manager，手动 `close()`。
    """

    def __init__(self, path: str | Path) -> None:
        self._path = Path(path)
        if not self._path.exists():
            raise WorkbookOpenError(f"base 文件不存在：{self._path}")
        if not self._path.is_file():
            raise WorkbookOpenError(f"base 路径不是文件：{self._path}")

        # read_only=True 流式读取，对 50K+ 行的 workbook 很重要。
        # data_only=True 让公式单元格返回缓存值。
        try:
            self._wb: Workbook = load_workbook(
                filename=str(self._path),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:  # openpyxl 抛出的异常种类杂，统一包装
            raise WorkbookOpenError(f"无法打开 base 文件 {self._path}：{exc}") from exc

    # —————————————————————————————————————
    # Context manager
    # —————————————————————————————————————

    def __enter__(self) -> WorkbookReader:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        self.close()

    def close(self) -> None:
        self._wb.close()

    # —————————————————————————————————————
    # Sheet 元信息
    # —————————————————————————————————————

    def sheet_names(self) -> tuple[str, ...]:
        return tuple(self._wb.sheetnames)

    def has_sheet(self, name: str) -> bool:
        return name in self._wb.sheetnames

    # —————————————————————————————————————
    # 主接口
    # —————————————————————————————————————

    def read_headers(
        self,
        sheet_name: str,
        header_row: int = HEADER_ROW,
    ) -> SheetHeaders:
        """只读取一个 sheet 的表头，不消费数据行。"""
        if sheet_name not in self._wb.sheetnames:
            raise WorkbookOpenError(f"workbook 中找不到 sheet：{sheet_name!r}")

        ws: Worksheet = self._wb[sheet_name]
        headers, header_columns, max_data_col = self._read_headers(ws, sheet_name, header_row)
        return SheetHeaders(
            sheet_name=sheet_name,
            headers=headers,
            header_columns=header_columns,
            max_data_col=max_data_col,
        )

    def read_sheet(
        self,
        sheet_name: str,
        header_row: int = HEADER_ROW,
        first_data_row: int = FIRST_DATA_ROW,
    ) -> SheetData:
        """读取一个 sheet 并返回结构化结果。

        - `header_row`、`first_data_row` 默认值来自 schema.py，正常调用不需要传。
        - sheet 不存在抛 `WorkbookOpenError`（结构性问题，validator 不必再处理）。
        - 完全空白的数据行被跳过；只要任意一列非空就视为有效行。
        """
        header_data = self.read_headers(sheet_name, header_row=header_row)
        ws: Worksheet = self._wb[sheet_name]
        rows = tuple(
            self._read_data_rows(
                ws,
                header_data.header_columns,
                first_data_row,
                header_data.max_data_col,
            )
        )

        return SheetData(
            sheet_name=sheet_name,
            headers=header_data.headers,
            header_columns=header_data.header_columns,
            rows=rows,
        )

    # —————————————————————————————————————
    # internals
    # —————————————————————————————————————

    @staticmethod
    def _read_headers(
        ws: Worksheet,
        sheet_name: str,
        header_row: int,
    ) -> tuple[tuple[str, ...], dict[str, int], int]:
        """从指定行读出表头并规范化。返回 (headers, header_columns, max_data_col)。

        max_data_col 是 header 行中最右侧有效列的 1-based 索引。当 Excel 维度被污染
        （如声明了 16379 列但只有 30 列有数据），用它限制后续行的读取范围，避免
        openpyxl 在每行创建上万个空值。
        """
        # ws.iter_rows in read_only mode is forward-only; advance to header_row
        header_cells: tuple[object, ...] = ()
        for cur_row, row in enumerate(ws.iter_rows(), start=1):
            if cur_row == header_row:
                header_cells = tuple(row)
                break

        # 找到 header 行中最右侧非 None 列 —— 超过此列的数据全部是维度污染
        max_data_col = 0
        for i in range(len(header_cells) - 1, -1, -1):
            if _cell_value(header_cells[i]) is not None:
                max_data_col = i + 1  # 1-based
                break

        ordered_headers: list[str] = []
        seen: dict[str, int] = {}
        for col_idx, raw in enumerate(header_cells, start=1):
            normalized = normalize_header(_cell_value(raw))
            normalized = _BASE_SCHEMA.canonical_header(sheet_name, normalized)
            if not normalized:
                continue
            if normalized not in seen:
                seen[normalized] = col_idx
                ordered_headers.append(normalized)
        return tuple(ordered_headers), seen, max_data_col

    @classmethod
    def _read_data_rows(
        cls,
        ws: Worksheet,
        header_columns: dict[str, int],
        first_data_row: int,
        max_data_col: int = 0,
    ) -> Iterator[dict[str, object]]:
        """从 first_data_row 开始迭代非空行。

        max_data_col 限制列范围，避免 Excel 维度污染导致每行创建大量空值。
        """
        # 重新建立一次迭代器：read_only 工作表里读两次需要重新 iter_rows
        row_iter = ws.iter_rows(max_col=max_data_col) if max_data_col > 0 else ws.iter_rows()
        for cur_row, row_cells in enumerate(row_iter, start=1):
            if cur_row < first_data_row:
                continue
            if _is_blank_row(row_cells):
                continue
            row_dict: dict[str, object] = {ROW_NUMBER_KEY: cur_row}
            number_formats: dict[str, str] = {}
            for header, col_idx in header_columns.items():
                cell = row_cells[col_idx - 1] if col_idx - 1 < len(row_cells) else None
                value = _cell_value(cell)
                row_dict[header] = value
                number_format = _cell_number_format(cell)
                if number_format:
                    number_formats[header] = number_format
            row_dict[CELL_NUMBER_FORMATS_KEY] = number_formats
            yield row_dict


def _is_blank_row(row_values: tuple[object, ...]) -> bool:
    """所有列都是 None 或空字符串才算空白行。

    不去除空格再判断：比如 `" "` 是用户故意留的占位，按非空对待，让 validator
    在数值字段上去捕捉这种异常。
    """
    return all(_cell_value(v) is None or _cell_value(v) == "" for v in row_values)


def row_cell_number_format(row: dict[str, object], header: str) -> str | None:
    raw = row.get(CELL_NUMBER_FORMATS_KEY)
    if not isinstance(raw, dict):
        return None
    value = raw.get(header)
    return value if isinstance(value, str) and value else None


def row_decimal_places(row: dict[str, object], header: str) -> int | None:
    return number_format_decimal_places(row_cell_number_format(row, header))


def number_format_decimal_places(number_format: str | None) -> int | None:
    if not number_format:
        return None
    primary = _primary_number_format_section(number_format)
    if not primary or _looks_like_date_format(primary):
        return None
    normalized = primary.replace("\\", "")
    match = re.search(r"[0#]+\.([0#]+)", normalized)
    if match is None:
        return 0 if re.search(r"[0#]", normalized) else None
    return len(match.group(1))


def _primary_number_format_section(number_format: str) -> str:
    return number_format.split(";", maxsplit=1)[0].strip()


def _looks_like_date_format(number_format: str) -> bool:
    lowered = number_format.lower()
    return any(token in lowered for token in ("yy", "dd", "mm", "hh", "ss"))


def _cell_value(cell: object) -> object:
    return getattr(cell, "value", cell)


def _cell_number_format(cell: object) -> str | None:
    value = getattr(cell, "number_format", None)
    return value if isinstance(value, str) and value else None


__all__ = [
    "CELL_NUMBER_FORMATS_KEY",
    "ROW_NUMBER_KEY",
    "SheetData",
    "SheetHeaders",
    "WorkbookReader",
    "number_format_decimal_places",
    "row_cell_number_format",
    "row_decimal_places",
]
