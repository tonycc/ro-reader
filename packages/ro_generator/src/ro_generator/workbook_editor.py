"""工作簿编辑：对 base 文件的字段级写回。

供工作台 inline 编辑使用。所有 openpyxl 直接操作集中在此模块，
API 层只调用公开接口。
"""

from __future__ import annotations

import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any, cast

from openpyxl import load_workbook

from ro_generator.base_schema import BaseSchema
from ro_generator.profiles.base import CustomerProfile
from ro_generator.profiles.runtime import current_schema
from ro_generator.schema import normalize_header

_file_locks: dict[str, threading.Lock] = {}
_file_locks_guard = threading.Lock()


@dataclass(frozen=True)
class EditResult:
    ok: bool
    message: str = ""


def _lock_for_file(path: str) -> threading.Lock:
    with _file_locks_guard:
        if path not in _file_locks:
            _file_locks[path] = threading.Lock()
        return _file_locks[path]


def edit_workbook_cell(
    base_file: str,
    sheet: str,
    row: int,
    field: str,
    value: object,
    *,
    header_row: int | None = None,
    schema: BaseSchema | None = None,
    profile: CustomerProfile | None = None,
) -> EditResult:
    """将值写回 base 文件的指定单元格。

    通过表头行定位列（默认来自 base_schema.yaml），再按 (row, col) 写入。
    使用 per-file 互斥锁防止并发写入导致覆盖。
    """
    path = Path(base_file)
    if not path.exists():
        return EditResult(ok=False, message=f"base 文件不存在：{path}")

    resolved = str(path.resolve())
    active_schema = profile.schema if profile is not None else current_schema(schema)
    logical_sheet = active_schema.logical_sheet_name(sheet)
    if logical_sheet is None:
        return EditResult(ok=False, message=f"sheet {sheet!r} 不存在")
    physical_sheet = active_schema.sheet(logical_sheet).name
    if header_row is None:
        sheet_config = active_schema.sheet(logical_sheet)
        header_row = sheet_config.header_row

    with _lock_for_file(resolved):
        try:
            wb = load_workbook(str(path))
        except Exception as exc:
            return EditResult(ok=False, message=f"无法打开工作簿：{exc}")

        try:
            if physical_sheet not in wb.sheetnames:
                return EditResult(ok=False, message=f"sheet {physical_sheet!r} 不存在")

            ws = wb[physical_sheet]
            col_map: dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=c)
                if cell.value is not None:
                    norm = normalize_header(cell.value)
                    if norm:
                        col_map[norm] = c

            field_candidates = {
                normalize_header(field),
                normalize_header(active_schema.field(logical_sheet, field)),
            }
            col_idx = next(
                (col_map[candidate] for candidate in field_candidates if candidate in col_map), None
            )
            if col_idx is None:
                return EditResult(ok=False, message=f"表头中找不到字段 {field!r}")

            ws.cell(row=row, column=col_idx, value=cast(Any, value))
            wb.save(str(path))
            return EditResult(ok=True, message=f"已更新 {physical_sheet} row={row} {field}")
        except Exception as exc:
            return EditResult(ok=False, message=str(exc))
        finally:
            wb.close()
