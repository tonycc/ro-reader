"""Generator 流水线集成测试 — 新 base 文件数据模型。"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from ro_generator.generator import (
    CODE_MAPPING_NOT_FOUND,
    INPUT_INVOICE_NO,
    INPUT_SELLER,
    build_document_model,
    export_invoice_group_from_snapshot,
    generate,
    preview,
    preview_invoice_group_from_snapshot,
)
from ro_generator.models import DocumentRequest
from ro_generator.resolver import resolve_po_lines
from ro_generator.source_index import SourceIndex
from ro_generator.workbook_reader import WorkbookReader
from ro_generator.workbook_snapshot import build_workbook_snapshot

DATA_BASE_HEADER = [
    "SAP",
    "Material Description",
    "Category",
    "GS MODEL",
    "GS-SK/YM COMBO FOB 2026",
    "GS-SK/YM YM ROD FOB 2026",
    "GS-SK/YM SK REEL FOB 2026",
    "EMAX-GS PTE COMBO FOB 2026",
    "EMAX PTE COMBO FOB 2026",
    "round value",
    "L",
    "W",
    "H",
]

PO_RECORD_HEADER = [
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "FINALQTY",
    "GS-SK/YM USD FOB",
    "EMAX-GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "SHIP QTY",
    "SK/YM INVOICE NO.",
    "CTNS",
    "TOTAL CBM",
    "外箱(最终出口装箱率)",
    "N/W",
    "G/W",
    "L",
    "W",
    "H",
    "FINAL EX-FACTORY DATE",
    "E10 PO",
    "YM PO",
    "CATEGORY",
]

CUSTOMER_PO_HEADER = [
    "Purchasing Document",
    "Item",
    "Material",
    "ship to",
    "Order Quantity",
    "ship DATE",
    "manufacturer",
    "final destination",
]


def _write_sheet(ws, headers, rows, header_row=4, first_data_row=5):
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            if header in row and row[header] is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=row[header])


def _default_customer_po_rows(po_record_rows):
    rows = []
    for row in po_record_rows:
        po_no = row.get("PO NO.")
        material = row.get("SAP Number")
        if po_no is None or material is None:
            continue
        rows.append(
            {
                "Purchasing Document": po_no,
                "Item": str(row.get("ITEM LINE#", "10")),
                "Material": material,
                "ship to": "Customer PO Ship To",
                "Order Quantity": row.get("FINALQTY", 100),
                "ship DATE": row.get("FINAL EX-FACTORY DATE"),
            }
        )
    return rows


def make_base_file(
    tmp_path, *, data_base_rows, po_record_rows, customer_po_rows=None, name="base.xlsx"
):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(ws_db, DATA_BASE_HEADER, data_base_rows)
    ws_po = wb.create_sheet("PO record")
    _write_sheet(ws_po, PO_RECORD_HEADER, po_record_rows)
    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(
        ws_cp,
        CUSTOMER_PO_HEADER,
        customer_po_rows
        if customer_po_rows is not None
        else _default_customer_po_rows(po_record_rows),
        header_row=1,
        first_data_row=2,
    )
    path = tmp_path / name
    wb.save(path)
    return path


COMBO_PRODUCT = {
    "SAP": "21-44640",
    "Material Description": "CB2500.B2",
    "Category": 1,
    "GS MODEL": "Q1",
    "GS-SK/YM COMBO FOB 2026": Decimal("28.0"),
    "EMAX-GS PTE COMBO FOB 2026": Decimal("32.8"),
    "EMAX PTE COMBO FOB 2026": Decimal("38.0"),
    "round value": 24,
    "L": 60,
    "W": 40,
    "H": 30,
}


def basic_po_row(**overrides):
    base = {
        "PO NO.": "4500030844",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2",
        "FINALQTY": 100,
        "GS-SK/YM USD FOB": Decimal("28.0"),
        "EMAX-GS PTE FOB": Decimal("32.8"),
        "EMAX PTE": Decimal("38.0"),
        "INV#": "INV-001",
        "SHIP QTY": 100,
        "SK/YM INVOICE NO.": "SKYM-INV-001",
        "外箱(最终出口装箱率)": 24,
        "CTNS": 5,
        "TOTAL CBM": Decimal("0.36"),
        "N/W": Decimal("8.5"),
        "G/W": Decimal("10.1"),
        "L": 48,
        "W": 31,
        "H": 35,
        "FINAL EX-FACTORY DATE": date(2026, 3, 15),
        "CATEGORY": 1,
    }
    base.update(overrides)
    return base


class TestSuccessPath:
    def test_zip_output_with_single_workbook_is_still_packaged(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_format="zip",
            output_dir=str(tmp_path / "out"),
        )

        result = generate(request)

        assert result.status == "success", result.errors
        assert result.output_file is not None
        assert result.output_file.endswith(".zip")
        with ZipFile(result.output_file) as archive:
            assert archive.namelist() == list(result.files)

    def test_invoice_with_seller_and_invoice_no(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors
        assert result.output_file is not None
        assert Path(result.output_file).exists()
        assert "INVOICE" in str(result.files[0])
        assert "INV-001" in str(result.files[0])
        assert isinstance(result.source_index, SourceIndex)
        assert len(result.source_index) > 0
        assert result.summary["seller"] == "GS PTE"
        assert result.summary["line_count"] == 1

    def test_output_file_contains_real_data(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        wb = load_workbook(result.output_file)
        ws = wb["INV"]
        assert ws["H6"].value == "INV-001"
        assert ws["H7"].value is not None  # invoice_date written

    @pytest.mark.parametrize(
        ("seller", "field", "expected"),
        [
            ("SK", "E10 PO", "E10-PO-001"),
            ("YM", "YM PO", "YM-PO-001"),
        ],
    )
    def test_sk_ym_pi_uses_factory_po_number_when_present(self, tmp_path, seller, field, expected):
        category = 3 if seller == "SK" else 1
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{field: expected, "CATEGORY": category})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller=seller,
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)

        assert result.status == "success", result.errors
        assert result.preview is not None
        assert result.preview.pi_no == expected

    def test_sk_pi_export_ignores_unselected_factory_rows_missing_customer_po(self, tmp_path):
        reel_product = {
            **COMBO_PRODUCT,
            "SAP": "21-REEL",
            "Material Description": "REEL ITEM",
            "Category": 3,
            "GS-SK/YM SK REEL FOB 2026": Decimal("33.0"),
        }
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT, reel_product],
            po_record_rows=[
                basic_po_row(
                    **{
                        "ITEM LINE#": "10",
                        "SAP Number": "21-44640",
                        "DESCRIPTION": "YM ITEM",
                        "CATEGORY": 1,
                        "YM PO": "YM-PI-001",
                    }
                ),
                basic_po_row(
                    **{
                        "ITEM LINE#": "20",
                        "SAP Number": "21-REEL",
                        "DESCRIPTION": "SK ITEM",
                        "CATEGORY": 3,
                        "E10 PO": "SK-PI-001",
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "20",
                    "Material": "21-REEL",
                    "Order Quantity": 30,
                },
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="SK",
            output_dir=str(tmp_path / "out"),
        )

        result = generate(request)

        assert result.status == "success", result.errors
        wb = load_workbook(result.output_file)
        ws = wb["Standard Invoice format"]
        assert ws["B6"].value == "SK-PI-001"
        assert ws["D20"].value == "21-REEL"


class TestNeedsInput:
    def test_multiple_invoices_needs_input(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(FINALQTY=100, **{"INV#": "INV-001"}),
                basic_po_row(**{"ITEM LINE#": "20", "FINALQTY": 100, "INV#": "INV-002"}),
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "needs_input"
        assert INPUT_INVOICE_NO in result.missing_inputs

    def test_single_invoice_auto_selects(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors

    def test_segment_undecided_returns_needs_input(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "needs_input"
        assert INPUT_SELLER in result.missing_inputs


class TestErrorPath:
    def test_unknown_po(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="9999999",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "error"
        assert any(m.code == "PO_NOT_FOUND" for m in result.errors)

    def test_missing_invoice_no_blocks_generation(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"INV#": None})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "error"

    def test_empty_invoice_no_without_selection_warns_and_generates(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"INV#": None})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success"
        assert any(m.code == "INVOICE_NO_MISSING" for m in result.warnings)

    def test_multi_doc_generates_both(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI", "INVOICE"),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors
        assert len(result.files) == 2

    @pytest.mark.parametrize(
        ("seller", "invoice_no", "expected_file", "invoice_sheet"),
        [
            ("GS PTE", "INV-001", "GS_PTE-RO-INVOICE&PL-4500030844-INV-001.xlsx", "INV"),
            ("EMAX PTE", "INV-001-P", "EMAX_PTE-RO-INVOICE&PL-4500030844-INV-001-P.xlsx", "CI"),
        ],
    )
    def test_non_factory_invoice_pl_generates_single_workbook_with_two_sheets(
        self,
        tmp_path,
        seller,
        invoice_no,
        expected_file,
        invoice_sheet,
    ):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE", "PL"),
            seller=seller,
            invoice_no=invoice_no,
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)

        assert result.status == "success", result.errors
        assert result.files == (expected_file,)
        assert result.output_file is not None
        assert result.output_file.endswith(".xlsx")

        wb = load_workbook(result.output_file)
        assert invoice_sheet in wb.sheetnames
        assert "PL" in wb.sheetnames

    def test_sk_ym_po_blocked(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PO",),
            seller="SK",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "error"
        assert any(m.code == CODE_MAPPING_NOT_FOUND for m in result.errors)

    @pytest.mark.parametrize(
        ("seller", "field"),
        [
            ("SK", "E10 PO"),
            ("YM", "YM PO"),
        ],
    )
    def test_sk_ym_pi_requires_factory_po_number(self, tmp_path, seller, field):
        category = 3 if seller == "SK" else 1
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{field: None, "CATEGORY": category})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller=seller,
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)

        assert result.status == "error"
        assert any(m.code == "PI_NO_MISSING" and m.field == field for m in result.errors)

    def test_missing_workbook_returns_error(self, tmp_path):
        request = DocumentRequest(
            base_file=str(tmp_path / "nope.xlsx"),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "error"
        assert any(m.code == "WORKBOOK_OPEN_ERROR" for m in result.errors)

    def test_missing_sheet_returns_error(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            wb.remove(ws)
        wb.create_sheet("DATA BASE")
        path = tmp_path / "incomplete.xlsx"
        wb.save(path)
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "error"
        assert any(m.code == "SHEET_MISSING" for m in result.errors)


class TestWarningsPropagation:
    def test_formula_fallback_warning_in_result(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=240, **{"CTNS": None, "TOTAL CBM": None})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors
        assert any(m.code == "FORMULA_FALLBACK" for m in result.warnings)


class TestConflictStrategy:
    def test_overwrite_replaces_existing(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
            on_conflict="overwrite",
        )
        first = generate(request)
        assert first.status == "success"
        second = generate(request)
        assert second.status == "success"
        assert first.output_file == second.output_file


@pytest.mark.parametrize("_label", ["smoke"])
def test_generation_result_immutable(tmp_path, _label):
    path = make_base_file(tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()])
    request = DocumentRequest(
        base_file=str(path),
        po_no="4500030844",
        documents=("INVOICE",),
        seller="GS PTE",
        invoice_no="INV-001",
        output_dir=str(tmp_path / "out"),
    )
    import dataclasses

    result = generate(request)
    with pytest.raises(dataclasses.FrozenInstanceError):
        result.status = "error"


class TestBuildDocumentModel:
    def test_builds_invoice_model(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        with WorkbookReader(str(path)) as reader:
            resolved = resolve_po_lines(reader, "4500030844")
        build = build_document_model(
            resolved.lines,
            seller="GS PTE",
            buyer="EMAX PTE",
            po_no="4500030844",
            invoice_no="INV-001",
            doc_type="INVOICE",
        )
        assert build.model is not None
        assert build.mapping is not None
        assert build.model.document_type == "INVOICE"
        assert build.model.seller == "GS PTE"
        assert build.model.invoice_no == "INV-001"
        assert len(build.model.lines) == 1

    def test_builds_pi_model(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        with WorkbookReader(str(path)) as reader:
            resolved = resolve_po_lines(reader, "4500030844")
        build = build_document_model(
            resolved.lines,
            seller="GS PTE",
            buyer="EMAX PTE",
            po_no="4500030844",
            invoice_no=None,
            doc_type="PI",
        )
        assert build.model is not None
        assert build.model.document_type == "PI"
        assert build.model.total_quantity == Decimal("100")
        assert str(build.model.ex_factory_date) == "2026-03-15"
        assert build.model.ship_to == "Customer PO Ship To"

    def test_sk_po_blocked(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        with WorkbookReader(str(path)) as reader:
            resolved = resolve_po_lines(reader, "4500030844")
        build = build_document_model(
            resolved.lines,
            seller="SK",
            buyer="GS PTE",
            po_no="4500030844",
            invoice_no=None,
            doc_type="PO",
        )
        assert build.model is None
        assert any(m.code == CODE_MAPPING_NOT_FOUND for m in build.messages)


class TestPreviewFunction:
    def test_preview_returns_structured_data(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        assert result.preview is not None

        p = result.preview
        assert getattr(p, "document_type", "") == "INVOICE"
        assert getattr(p, "title", "") != ""
        assert getattr(p, "seller", "") == "GS PTE"
        assert getattr(p, "buyer", "") == "EMAX PTE"
        assert getattr(p, "po_no", "") == "4500030844"
        assert getattr(p, "invoice_no", None) == "INV-001"
        assert len(getattr(p, "lines", [])) == 1
        assert len(getattr(p, "source_entries", [])) > 0

    def test_preview_no_output_file(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        # Preview should NOT produce an output_file
        assert not hasattr(result, "output_file")

    def test_preview_needs_input_for_multiple_invoices(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(FINALQTY=100, **{"INV#": "INV-001"}),
                basic_po_row(**{"ITEM LINE#": "20", "FINALQTY": 100, "INV#": "INV-002"}),
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "needs_input"
        assert INPUT_INVOICE_NO in result.missing_inputs

    def test_preview_error_for_missing_po(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="9999999",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "error"

    def test_preview_contains_seller_info(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        p = result.preview
        seller_info = getattr(p, "seller_info", [])
        assert len(seller_info) > 0
        assert any("GLOBALSINO" in line for line in seller_info)

    def test_preview_contains_terms(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        p = result.preview
        terms = getattr(p, "terms", {})
        assert isinstance(terms, dict)
        assert terms == {
            "term": "T/T 75 DAYS AFTER BL DATE",
            "from": "QINGDAO, CHINA",
            "to": "KANSAS CITY, MO, USA",
        }

    def test_preview_terms_merge_resolved_header_fields_and_static_terms(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        terms = getattr(p, "terms", {})
        assert terms == {
            "payment_terms": "Net 90 days",
            "port_of_loading": "China",
            "final_destination": "USA",
            "incoterm": "FOB GUANGDONG",
        }
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["payment_terms"] == "Net 90 days"
        assert resolved_values["port_of_loading"] == "China"
        assert resolved_values["final_destination"] == "USA"
        assert resolved_values["bill_to"] == "209 Stoneridge Drive"
        assert resolved_values["bill_to_line2"] == "Columbia, South Carolina 29210"
        assert resolved_values["bill_to_line3"] == "United States"
        layout = getattr(p, "layout", {})
        info = layout.get("info", {}) if isinstance(layout, dict) else {}
        left = info.get("left", []) if isinstance(info, dict) else []
        assert "bill_to" in left
        assert "bill_to_line2" in left
        assert "bill_to_line3" in left

    @pytest.mark.parametrize(("seller", "pi_field"), [("SK", "E10 PO"), ("YM", "YM PO")])
    def test_sk_ym_pi_preview_terms_reuse_export_header_fixed_values(
        self,
        tmp_path,
        seller,
        pi_field,
    ):
        category = 3 if seller == "SK" else 1
        po_row = basic_po_row(**{pi_field: "PI-4500030844", "CATEGORY": category})
        path = make_base_file(tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[po_row])
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller=seller,
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        terms = getattr(p, "terms", {})
        assert terms == {
            "incoterm": "FOB Qingdao",
            "payment_terms": "Net 75 days",
            "port_of_loading": "Qingdao, China",
            "final_destination": "USA",
        }
        resolved_values = getattr(p, "resolved_values", {})
        assert terms["incoterm"] == resolved_values["incoterm"]
        assert terms["port_of_loading"] == resolved_values["port_of_loading"]

    def test_preview_document_date_source_entry_is_system_generated(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        entries = getattr(p, "source_entries", [])
        document_date_entry = next(e for e in entries if e["preview_field"] == "document_date")
        assert document_date_entry["source_type"] == "system_generated"
        assert document_date_entry["value"]
        assert "当天日期" in document_date_entry["rule"]

    def test_emax_pi_ex_factory_date_source_entry_comes_from_po_record(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"FINAL EX-FACTORY DATE": date(2026, 3, 15)})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                    "ship DATE": date(2026, 4, 20),
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["ex_factory_date"] == "2026-03-15"

        entries = getattr(p, "source_entries", [])
        ex_factory_entry = next(e for e in entries if e["preview_field"] == "ex_factory_date")
        assert ex_factory_entry["source_type"] == "base_field"
        assert ex_factory_entry["sheet"] == "PO record"
        assert ex_factory_entry["field"] == "FINAL EX-FACTORY DATE"
        assert ex_factory_entry["value"] == "2026-03-15"
        assert "FINAL EX-FACTORY DATE" in ex_factory_entry["rule"]

    def test_gs_pi_ship_to_uses_header_fixed_over_customer_po(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["ship_to"] == "E MAX SPORT PTE. LTD."

        # pi_no comes from customer PO Purchasing Document, not header_fixed
        entries = getattr(p, "source_entries", [])
        pi_no_entry = next(e for e in entries if e["preview_field"] == "pi_no")
        assert pi_no_entry["source_type"] == "base_field"
        assert pi_no_entry["sheet"] == "客户PO"
        assert pi_no_entry["field"] == "Purchasing Document"
        assert pi_no_entry["value"] == "4500030844"

        ship_to_entry = next(e for e in entries if e["preview_field"] == "ship_to")
        assert ship_to_entry["source_type"] == "template_content"
        assert ship_to_entry["sheet"] is None
        assert ship_to_entry["field"] is None
        assert ship_to_entry["value"] == "E MAX SPORT PTE. LTD."

    def test_emax_pi_preview_ship_to_uses_yaml_multiline_fields(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "209 Stoneridge Drive, Columbia, South Carolina 29210, United States",
                    "Order Quantity": 100,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["ship_to"] == "209 Stoneridge Drive"
        assert resolved_values["ship_to_line2"] == "Columbia, South Carolina 29210"
        assert resolved_values["ship_to_line3"] == "United States"

        right = p.layout["info"]["right"]
        assert "ship_to" in right
        assert "ship_to_line2" in right
        assert "ship_to_line3" in right

        entries = getattr(p, "source_entries", [])
        ship_to_line2 = next(e for e in entries if e["preview_field"] == "ship_to_line2")
        ship_to_line3 = next(e for e in entries if e["preview_field"] == "ship_to_line3")
        assert ship_to_line2["sheet"] == "客户PO"
        assert ship_to_line2["field"] == "ship to"
        assert ship_to_line2["value"] == "Columbia, South Carolina 29210"
        assert ship_to_line3["sheet"] == "客户PO"
        assert ship_to_line3["field"] == "ship to"
        assert ship_to_line3["value"] == "United States"

    def test_emax_pi_preview_ship_to_does_not_repeat_compact_company_address(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": (
                        "Rather Outdoors Corporation,\n"
                        "40 Industrial Road,Dauphin, MB R7N 2V2\n"
                        "Rather Outdoors Corporation, 40 Industrial Road,Dauphin, MB R7N 2V2"
                    ),
                    "Order Quantity": 100,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["ship_to"] == "Rather Outdoors Corporation,"
        assert resolved_values["ship_to_line2"] == "40 Industrial Road,Dauphin, MB R7N 2V2"
        assert "ship_to_line3" not in resolved_values

        export_request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        export_result = generate(export_request)
        assert export_result.status == "success", export_result.errors
        wb = load_workbook(export_result.output_file)
        ws = wb["Standard Invoice format"]
        assert ws["G9"].value == "Rather Outdoors Corporation,"
        assert ws["G10"].value == "40 Industrial Road,Dauphin, MB R7N 2V2"
        assert ws["G11"].value is None

    @pytest.mark.parametrize("doc_type", ["PI", "PO"])
    def test_gs_pi_po_manufacturer_fields_use_customer_po_y(self, tmp_path, doc_type):
        manufacturer = "ACME FACTORY LIMITED\nNO. 1 INDUSTRIAL ROAD\nQINGYUAN, GUANGDONG, CHINA"
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                    "manufacturer": manufacturer,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=(doc_type,),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )

        preview_result = preview(request)
        assert preview_result.status == "success", preview_result.errors
        assert preview_result.preview is not None
        resolved_values = preview_result.preview.resolved_values
        assert resolved_values["manufacturer"] == "ACME FACTORY LIMITED"
        assert resolved_values["manufacturer_address"] == "NO. 1 INDUSTRIAL ROAD"
        assert resolved_values["manufacturer_address_2"] == "QINGYUAN, GUANGDONG, CHINA"

        entries = getattr(preview_result.preview, "source_entries", [])
        for field_name in ("manufacturer", "manufacturer_address", "manufacturer_address_2"):
            entry = next(e for e in entries if e["preview_field"] == field_name)
            assert entry["sheet"] == "客户PO"
            assert entry["field"] == "manufacturer"

        export_result = generate(request)
        assert export_result.status == "success", export_result.errors
        wb = load_workbook(export_result.output_file)
        ws = wb["Sheet1"]
        assert ws["G14"].value == "ACME FACTORY LIMITED"
        assert ws["G15"].value == "NO. 1 INDUSTRIAL ROAD"
        assert ws["G16"].value == "QINGYUAN, GUANGDONG, CHINA"

    def test_emax_pi_item_no_comes_from_customer_po_material(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"ITEM LINE#": "10"})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "CP-ITEM-001",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        assert p.lines[0]["item_line_no"] == "CP-ITEM-001"

        entries = getattr(p, "source_entries", [])
        item_no_entry = next(e for e in entries if e["preview_field"] == "line[0].item_line_no")
        assert item_no_entry["sheet"] == "客户PO"
        assert item_no_entry["field"] == "item"
        assert item_no_entry["row"] is None
        assert item_no_entry["value"] == "CP-ITEM-001"

    def test_gs_po_item_number_uses_matching_customer_po_material(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"ITEM LINE#": "10"})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                    "ship DATE": date(2026, 3, 15),
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PO",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        assert p.lines[0]["item_number"] == "21-44640"

        entries = getattr(p, "source_entries", [])
        item_number_entry = next(e for e in entries if e["preview_field"] == "line[0].item_number")
        assert item_number_entry["sheet"] == "客户PO"
        assert item_number_entry["field"] == "material"
        assert item_number_entry["row"] is None
        assert item_number_entry["value"] == "21-44640"

    def test_emax_pi_line_sources_follow_data_base_and_customer_po_rules(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {
                    **COMBO_PRODUCT,
                    "Material Description": "DB Description",
                    "EMAX PTE COMBO FOB 2026": Decimal("88.8"),
                }
            ],
            po_record_rows=[
                basic_po_row(
                    **{
                        "DESCRIPTION": "PO Description",
                        "FINALQTY": 100,
                        "FINAL EX-FACTORY DATE": date(2026, 3, 15),
                    }
                )
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 240,
                    "ship DATE": date(2026, 4, 20),
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        first_line = p.lines[0]
        assert first_line["description"] == "DB Description"
        assert first_line["unit_price"] == "$88.80"
        assert first_line["quantity"] == "240"
        assert first_line["confirmed_ex_factory_date"] == date(2026, 3, 15)

        entries = getattr(p, "source_entries", [])
        description_entry = next(e for e in entries if e["preview_field"] == "line[0].description")
        unit_price_entry = next(e for e in entries if e["preview_field"] == "line[0].unit_price")
        quantity_entry = next(e for e in entries if e["preview_field"] == "line[0].quantity")
        ex_factory_entry = next(
            e for e in entries if e["preview_field"] == "line[0].confirmed_ex_factory_date"
        )
        assert description_entry["sheet"] == "DATA BASE"
        assert description_entry["field"] == "Material Description"
        assert unit_price_entry["sheet"] == "DATA BASE"
        assert unit_price_entry["field"] == "EMAX PTE COMBO FOB 2026"
        assert quantity_entry["sheet"] == "客户PO"
        assert quantity_entry["field"] == "Order Quantity"
        assert quantity_entry["value"] == "240"
        assert ex_factory_entry["sheet"] == "PO record"
        assert ex_factory_entry["field"] == "FINAL EX-FACTORY DATE"

    def test_emax_pi_preview_formats_usd_unit_price_and_amount_with_dollar(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {
                    **COMBO_PRODUCT,
                    "EMAX PTE COMBO FOB 2026": Decimal("88.8"),
                }
            ],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 240,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        first_line = p.lines[0]
        assert first_line["unit_price"] == "$88.80"
        assert first_line["amount"] == "$21,312.00"
        assert p.totals["total_amount"] == "$21,312.00"

        entries = getattr(p, "source_entries", [])
        unit_price_entry = next(e for e in entries if e["preview_field"] == "line[0].unit_price")
        amount_entry = next(e for e in entries if e["preview_field"] == "line[0].amount")
        assert unit_price_entry["value"] == "$88.80"
        assert amount_entry["value"] == "$21,312.00"

    def test_emax_po_unit_price_uses_emax_pte_fob_column(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {
                    **COMBO_PRODUCT,
                    "EMAX-GS PTE COMBO FOB 2026": Decimal("32.8"),
                    "EMAX PTE COMBO FOB 2026": Decimal("88.8"),
                }
            ],
            po_record_rows=[basic_po_row(**{"FINALQTY": 100})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 240,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PO",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        first_line = p.lines[0]
        assert first_line["unit_price"] == "$88.80"

        entries = getattr(p, "source_entries", [])
        unit_price_entry = next(e for e in entries if e["preview_field"] == "line[0].unit_price")
        assert unit_price_entry["sheet"] == "DATA BASE"
        assert unit_price_entry["field"] == "EMAX PTE COMBO FOB 2026"

    def test_gs_po_unit_price_uses_gs_sk_ym_fob_column(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {
                    **COMBO_PRODUCT,
                    "GS-SK/YM COMBO FOB 2026": Decimal("28.0"),
                    "EMAX-GS PTE COMBO FOB 2026": Decimal("32.8"),
                }
            ],
            po_record_rows=[basic_po_row(**{"FINALQTY": 100})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 240,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PO",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        first_line = p.lines[0]
        assert first_line["unit_price"] == "$28.00"

        entries = getattr(p, "source_entries", [])
        unit_price_entry = next(e for e in entries if e["preview_field"] == "line[0].unit_price")
        assert unit_price_entry["sheet"] == "DATA BASE"
        assert unit_price_entry["field"] == "GS-SK/YM COMBO FOB 2026"

    @pytest.mark.parametrize(
        ("doc_type", "invoice_no"),
        [("PI", None), ("INVOICE", "SKYM-INV-001")],
    )
    def test_ym_pi_and_invoice_unit_price_always_use_g_column(
        self,
        tmp_path,
        doc_type,
        invoice_no,
    ):
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {
                    **COMBO_PRODUCT,
                    "Category": 2,
                    "GS-SK/YM COMBO FOB 2026": Decimal("28.0"),
                    "GS-SK/YM YM ROD FOB 2026": Decimal("99.0"),
                }
            ],
            po_record_rows=[
                basic_po_row(
                    **{
                        "FINALQTY": 100,
                        "SHIP QTY": 40,
                        "SK/YM INVOICE NO.": "SKYM-INV-001",
                        "YM PO": "YM-PI-001",
                    }
                )
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 240,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=(doc_type,),
            seller="YM",
            invoice_no=invoice_no,
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        first_line = p.lines[0]
        assert first_line["unit_price"] == "$28.00"

        entries = getattr(p, "source_entries", [])
        unit_price_entry = next(e for e in entries if e["preview_field"] == "line[0].unit_price")
        assert unit_price_entry["sheet"] == "DATA BASE"
        assert unit_price_entry["field"] == "GS-SK/YM COMBO FOB 2026"

        export_result = generate(request)
        assert export_result.status == "success", export_result.errors
        assert export_result.output_file is not None
        wb = load_workbook(export_result.output_file)
        ws = wb["SHEET1"] if doc_type == "PI" else wb["Standard Invoice format"]
        unit_price_cell = "F20" if doc_type == "PI" else "E15"
        assert ws[unit_price_cell].value == Decimal("28.0")

    def test_gs_po_header_and_line_ex_factory_dates_use_different_sources(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "FINAL EX-FACTORY DATE": date(2026, 3, 15),
                    }
                )
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 240,
                    "ship DATE": date(2026, 4, 20),
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PO",),
            seller="GS PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None

        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["ex_factory_date"] == "2026-03-15"
        assert p.lines[0]["confirmed_ex_factory_date"] == date(2026, 4, 20)

        entries = getattr(p, "source_entries", [])
        header_entry = next(e for e in entries if e["preview_field"] == "ex_factory_date")
        line_entry = next(
            e for e in entries if e["preview_field"] == "line[0].confirmed_ex_factory_date"
        )
        assert header_entry["sheet"] == "PO record"
        assert header_entry["field"] == "FINAL EX-FACTORY DATE"
        assert line_entry["sheet"] == "客户PO"
        assert line_entry["field"] == "ship DATE"

    def test_emax_invoice_number_appends_p_suffix(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"INV#": "EMAX20260710"})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        assert p.invoice_no == "EMAX20260710-P"

        entries = getattr(p, "source_entries", [])
        invoice_entry = next(e for e in entries if e["preview_field"] == "invoice_no")
        assert invoice_entry["sheet"] == "PO record"
        assert invoice_entry["field"] == "INV#"
        assert invoice_entry["value"] == "EMAX20260710-P"
        assert '"-P"' in invoice_entry["rule"]

    @pytest.mark.parametrize("seller", ["SK", "YM"])
    def test_sk_ym_invoice_and_pl_export_as_two_sheet_workbook(self, tmp_path, seller):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "SHIP QTY": 40,
                        "YM PO": "YM-PI-001",
                        "CATEGORY": 3 if seller == "SK" else 1,
                    }
                )
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE", "PL"),
            seller=seller,
            invoice_no="SKYM-GS-001",
            output_dir=str(tmp_path / "out"),
        )

        result = generate(request)

        assert result.status == "success", result.errors
        assert result.output_file is not None
        assert len(result.files) == 1
        assert "INVOICE&PL" in result.files[0]
        wb = load_workbook(result.output_file)
        assert wb.sheetnames == ["Standard Invoice format", "PL"]
        assert wb["Standard Invoice format"]["H6"].value == "SKYM-GS-001"
        assert wb["Standard Invoice format"]["F15"].value == Decimal("40")
        assert wb["PL"]["K5"].value == "SKYM-GS-001"
        assert wb["PL"]["E9"].value == Decimal("40")

    @pytest.mark.parametrize(
        ("seller", "expected_sap", "unexpected_sap"),
        [("YM", "21-44640", "21-REEL"), ("SK", "21-REEL", "21-44640")],
    )
    def test_sk_ym_invoice_and_pl_export_uses_selected_seller_workbook(
        self,
        tmp_path,
        seller,
        expected_sap,
        unexpected_sap,
    ):
        reel_product = {
            **COMBO_PRODUCT,
            "SAP": "21-REEL",
            "Material Description": "REEL ITEM",
            "Category": 3,
            "GS-SK/YM SK REEL FOB 2026": Decimal("33.0"),
        }
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT, reel_product],
            po_record_rows=[
                basic_po_row(
                    **{
                        "ITEM LINE#": "10",
                        "SAP Number": "21-44640",
                        "DESCRIPTION": "COMBO ITEM",
                        "SHIP QTY": 40,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CATEGORY": 1,
                    }
                ),
                basic_po_row(
                    **{
                        "ITEM LINE#": "20",
                        "SAP Number": "21-REEL",
                        "DESCRIPTION": "REEL ITEM",
                        "SHIP QTY": 30,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CATEGORY": 3,
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
                {
                    "Purchasing Document": "4500030844",
                    "Item": "20",
                    "Material": "21-REEL",
                    "Order Quantity": 30,
                },
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE", "PL"),
            seller=seller,
            invoice_no="SKYM-GS-001",
            output_dir=str(tmp_path / "out"),
        )

        result = generate(request)

        assert result.status == "success", result.errors
        assert len(result.files) == 1
        assert result.output_file is not None
        assert result.output_file.endswith(".xlsx")
        assert result.files[0].startswith(f"{seller}-RO-INVOICE&PL-")
        wb = load_workbook(result.output_file)
        assert wb.sheetnames == ["Standard Invoice format", "PL"]
        assert wb["Standard Invoice format"]["D15"].value == expected_sap
        assert wb["PL"]["D9"].value == expected_sap
        assert wb["Standard Invoice format"]["D16"].value != unexpected_sap

    @pytest.mark.parametrize(
        ("seller", "expected_sap"),
        [("YM", "21-44640"), ("SK", "21-REEL")],
    )
    def test_sk_ym_preview_filters_lines_by_po_record_category(
        self,
        tmp_path,
        seller,
        expected_sap,
    ):
        reel_product = {
            **COMBO_PRODUCT,
            "SAP": "21-REEL",
            "Material Description": "REEL ITEM",
            "Category": 3,
            "GS-SK/YM SK REEL FOB 2026": Decimal("33.0"),
        }
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT, reel_product],
            po_record_rows=[
                basic_po_row(
                    **{
                        "ITEM LINE#": "10",
                        "SAP Number": "21-44640",
                        "DESCRIPTION": "COMBO ITEM",
                        "SHIP QTY": 40,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CATEGORY": 1,
                    }
                ),
                basic_po_row(
                    **{
                        "ITEM LINE#": "20",
                        "SAP Number": "21-REEL",
                        "DESCRIPTION": "REEL ITEM",
                        "SHIP QTY": 30,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CATEGORY": 3,
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
                {
                    "Purchasing Document": "4500030844",
                    "Item": "20",
                    "Material": "21-REEL",
                    "Order Quantity": 30,
                },
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller=seller,
            invoice_no="SKYM-GS-001",
            output_dir=str(tmp_path / "out"),
        )

        result = preview(request)

        assert result.status == "success"
        p = result.preview
        assert p is not None
        assert len(p.lines) == 1
        assert p.lines[0]["sap"] == expected_sap

    @pytest.mark.parametrize("doc_type", ["INVOICE", "PL"])
    def test_sk_ym_gs_invoice_number_uses_sk_ym_invoice_no(self, tmp_path, doc_type):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "INV#": "RAW-INV-001",
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CATEGORY": 3,
                    }
                )
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=(doc_type,),
            seller="SK",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        assert p.invoice_no == "SKYM-GS-001"

        entries = getattr(p, "source_entries", [])
        invoice_entry = next(e for e in entries if e["preview_field"] == "invoice_no")
        assert invoice_entry["sheet"] == "PO record"
        assert invoice_entry["field"] == "SK/YM INVOICE NO."
        assert invoice_entry["value"] == "SKYM-GS-001"

    @pytest.mark.parametrize("seller", ["SK", "YM"])
    def test_sk_ym_invoice_to_uses_customer_po_final_destination(self, tmp_path, seller):
        final_destination = "LOS ANGELES, CA, USA"
        category = 3 if seller == "SK" else 1
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"CATEGORY": category})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "ship to": "Customer PO Warehouse",
                    "Order Quantity": 100,
                    "ship DATE": date(2026, 4, 20),
                    "final destination": final_destination,
                }
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller=seller,
            invoice_no="SKYM-INV-001",
            output_dir=str(tmp_path / "out"),
        )
        preview_result = preview(request)
        assert preview_result.status == "success", preview_result.errors
        assert preview_result.preview is not None
        assert preview_result.preview.resolved_values["to"] == final_destination

        entries = getattr(preview_result.preview, "source_entries", [])
        to_entry = next(e for e in entries if e["preview_field"] == "to")
        assert to_entry["sheet"] == "客户PO"
        assert to_entry["field"] == "final destination"
        assert to_entry["value"] == final_destination

        export_result = generate(request)
        assert export_result.status == "success", export_result.errors
        wb = load_workbook(export_result.output_file)
        ws = wb["Standard Invoice format"]
        assert ws["A12"].value == final_destination

    def test_preview_invoice_uses_po_record_description_rule(self, tmp_path):
        po_row = basic_po_row(DESCRIPTION="PO Record Description")
        db_row = dict(COMBO_PRODUCT)
        db_row["Material Description"] = "DB Description"
        path = make_base_file(tmp_path, data_base_rows=[db_row], po_record_rows=[po_row])
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("INVOICE",),
            seller="GS PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        assert p.lines[0]["description"] == "PO Record Description"

        entries = getattr(p, "source_entries", [])
        description_entry = next(e for e in entries if e["preview_field"] == "line[0].description")
        assert description_entry["sheet"] == "PO record"
        assert description_entry["field"] == "DESCRIPTION"
        assert description_entry["value"] == "PO Record Description"

    def test_preview_pl_totals_keep_shared_labels_and_values(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="GS PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None
        totals = getattr(p, "totals", {})
        labels = totals.get("_labels", {})
        assert totals["total_quantity"] == "100"
        assert totals["total_net_weight"] == "8.5"
        assert totals["total_gross_weight"] == "10.1"
        assert totals["total_cbm"] == "0.36"
        assert totals["total_carton_count"] == "5"
        assert labels["total_quantity"] == "Total Qty"
        assert labels["total_net_weight"] == "Total N/W (KGS)"
        assert labels["total_gross_weight"] == "Total G/W (KGS)"
        assert labels["total_cbm"] == "Total CBM"
        assert labels["total_carton_count"] == "Total CTNS"

        footer_items = totals.get("_footer_items", [])
        assert footer_items == [
            {"key": "quantity", "label": "Total Qty", "value": "100 PCS"},
            {"key": "net_weight", "label": "Total N/W (KGS)", "value": "8.5"},
            {"key": "gross_weight", "label": "Total G/W (KGS)", "value": "10.1"},
            {"key": "cbm", "label": "Total CBM", "value": "0.36 CBM"},
            {"key": "carton_count", "label": "Total CTNS", "value": "5"},
        ]
        layout = getattr(p, "layout", {})
        info = layout.get("info", {}) if isinstance(layout, dict) else {}
        top = layout.get("top", {}) if isinstance(layout, dict) else {}
        assert info.get("left", []) == ["shipping_mark", "shipping_mark_2"]
        assert info.get("right", []) == ["invoice_no"]
        assert top.get("left", []) == ["seller_info"]

    @pytest.mark.parametrize(
        ("seller", "invoice_no"),
        [
            ("GS PTE", "INV-001"),
            ("EMAX PTE", "INV-001-P"),
        ],
    )
    def test_preview_gs_emax_pl_includes_ctns_column(self, tmp_path, seller, invoice_no):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"CTNS": Decimal("7")})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller=seller,
            invoice_no=invoice_no,
        )
        result = preview(request)
        assert result.status == "success", result.errors

        p = result.preview
        assert p is not None
        labels = {c["key"]: c["label"] for c in p.column_labels}
        assert labels["carton_count"] == "CTNS"
        assert p.lines[0]["carton_count"] == "7"

        entries = getattr(p, "source_entries", [])
        ctns_entry = next(e for e in entries if e["preview_field"] == "line[0].carton_count")
        assert ctns_entry["sheet"] == "PO record"
        assert ctns_entry["field"] == "CTNS"
        assert 'PO record AD列 "CTNS"' in ctns_entry["rule"]

    @pytest.mark.parametrize(
        ("seller", "category", "sap"),
        [("YM", 1, "21-44640"), ("SK", 3, "21-REEL")],
    )
    def test_preview_sk_ym_pl_includes_ctns_column(self, tmp_path, seller, category, sap):
        reel_product = {
            **COMBO_PRODUCT,
            "SAP": "21-REEL",
            "Material Description": "REEL ITEM",
            "Category": 3,
            "GS-SK/YM SK REEL FOB 2026": Decimal("33.0"),
        }
        product = reel_product if seller == "SK" else COMBO_PRODUCT
        path = make_base_file(
            tmp_path,
            data_base_rows=[product],
            po_record_rows=[
                basic_po_row(
                    **{
                        "SAP Number": sap,
                        "CATEGORY": category,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CTNS": Decimal("7"),
                    }
                )
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller=seller,
            invoice_no="SKYM-GS-001",
        )
        result = preview(request)
        assert result.status == "success", result.errors

        p = result.preview
        assert p is not None
        labels = {c["key"]: c["label"] for c in p.column_labels}
        assert labels["carton_count"] == "CTNS"
        assert p.lines[0]["carton_count"] == "7"

    @pytest.mark.parametrize(
        ("seller", "category", "sap"),
        [("YM", 1, "21-44640"), ("SK", 3, "21-REEL")],
    )
    def test_export_sk_ym_pl_writes_ctns_column(self, tmp_path, seller, category, sap):
        reel_product = {
            **COMBO_PRODUCT,
            "SAP": "21-REEL",
            "Material Description": "REEL ITEM",
            "Category": 3,
            "GS-SK/YM SK REEL FOB 2026": Decimal("33.0"),
        }
        product = reel_product if seller == "SK" else COMBO_PRODUCT
        path = make_base_file(
            tmp_path,
            data_base_rows=[product],
            po_record_rows=[
                basic_po_row(
                    **{
                        "SAP Number": sap,
                        "CATEGORY": category,
                        "SK/YM INVOICE NO.": "SKYM-GS-001",
                        "CTNS": Decimal("7"),
                    }
                )
            ],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller=seller,
            invoice_no="SKYM-GS-001",
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors

        wb = load_workbook(result.output_file)
        ws = wb["PL"]
        assert ws["M9"].value == 7
        assert ws["M12"].value == 7
        assert ws["A16"].value == "PACKED IN 7 CTNS"

    @pytest.mark.parametrize(
        ("seller", "invoice_no", "sheet", "line_cell", "total_cell"),
        [
            ("GS PTE", "INV-001", "PL", "M9", "M15"),
            ("EMAX PTE", "INV-001-P", "PL", "M10", "M16"),
        ],
    )
    def test_export_gs_emax_pl_writes_ctns_column(
        self,
        tmp_path,
        seller,
        invoice_no,
        sheet,
        line_cell,
        total_cell,
    ):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"CTNS": Decimal("7")})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller=seller,
            invoice_no=invoice_no,
            output_dir=str(tmp_path / "out"),
        )
        result = generate(request)
        assert result.status == "success", result.errors

        wb = load_workbook(result.output_file)
        ws = wb[sheet]
        assert ws[line_cell].value == 7
        assert ws[total_cell].value == 7

    def test_preview_emax_pl_includes_shipping_mark_headers(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="EMAX PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        resolved_values = getattr(p, "resolved_values", {})
        assert resolved_values["shipping_mark"] == "WEIHAI"
        assert resolved_values["shipping_mark_2"] == "C/T#"
        assert resolved_values["shipping_mark_3"] == "MADE IN CHINA"

        layout = getattr(p, "layout", {})
        info = layout.get("info", {}) if isinstance(layout, dict) else {}
        left = info.get("left", []) if isinstance(info, dict) else []
        right = info.get("right", []) if isinstance(info, dict) else []
        assert left == ["shipping_mark", "shipping_mark_2", "shipping_mark_3"]
        assert right == ["invoice_no"]

    def test_preview_emax_pl_uses_po_record_description_rule(self, tmp_path):
        po_row = basic_po_row(DESCRIPTION="PO Record Description")
        db_row = dict(COMBO_PRODUCT)
        db_row["Material Description"] = "DB Description"
        path = make_base_file(tmp_path, data_base_rows=[db_row], po_record_rows=[po_row])
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="EMAX PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        assert p.lines[0]["description"] == "PO Record Description"

        entries = getattr(p, "source_entries", [])
        description_entry = next(e for e in entries if e["preview_field"] == "line[0].description")
        po_no_entry = next(e for e in entries if e["preview_field"] == "line[0].po_no")
        assert description_entry["sheet"] == "PO record"
        assert description_entry["field"] == "DESCRIPTION"
        assert description_entry["value"] == "PO Record Description"
        assert po_no_entry["sheet"] == "客户PO"
        assert po_no_entry["field"] == "Purchasing Document"

    def test_preview_emax_pl_includes_fixed_unit_columns(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="EMAX PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        assert [c["key"] for c in p.column_labels] == [
            "po_no",
            "description",
            "sap",
            "quantity",
            "F",
            "net_weight",
            "H",
            "gross_weight",
            "J",
            "cbm",
            "L",
            "carton_count",
        ]
        labels = {c["key"]: c["label"] for c in p.column_labels}
        assert labels["F"] == ""
        assert labels["H"] == ""
        assert labels["J"] == ""
        assert labels["L"] == ""
        assert labels["carton_count"] == "CTNS"
        first_line = p.lines[0]
        assert first_line["F"] == "PCS"
        assert first_line["H"] == "KGS"
        assert first_line["J"] == "KGS"
        assert first_line["L"] == "CBM"
        assert first_line["carton_count"] == "5"

    def test_preview_emax_pl_uses_po_record_total_cbm_rule(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"TOTAL CBM": Decimal("1.2")})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="EMAX PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        assert p.lines[0]["cbm"] == "1.20"

        entries = getattr(p, "source_entries", [])
        cbm_entry = next(e for e in entries if e["preview_field"] == "line[0].cbm")
        assert cbm_entry["sheet"] == "PO record"
        assert cbm_entry["field"] == "TOTAL CBM"
        assert cbm_entry["value"] == "1.20"
        assert 'PO record AJ列 "TOTAL CBM"' in cbm_entry["rule"]

    def test_preview_emax_pl_preserves_source_cbm_decimal_places(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"TOTAL CBM": 1.2})],
        )
        wb = load_workbook(path)
        ws = wb["PO record"]
        total_cbm_col = PO_RECORD_HEADER.index("TOTAL CBM") + 1
        ws.cell(row=5, column=total_cbm_col).number_format = "0.0000"
        wb.save(path)
        wb.close()

        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller="EMAX PTE",
            invoice_no="INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        assert p.lines[0]["cbm"] == "1.2000"

        entries = getattr(p, "source_entries", [])
        cbm_entry = next(e for e in entries if e["preview_field"] == "line[0].cbm")
        assert cbm_entry["value"] == "1.2000"

    @pytest.mark.parametrize(
        "seller, expected_right", [("SK", ["invoice_no", "invoice_date"]), ("YM", ["invoice_no"])]
    )
    def test_preview_sk_ym_pl_layout_uses_supported_header_fields(
        self, tmp_path, seller, expected_right
    ):
        category = 3 if seller == "SK" else 1
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"CATEGORY": category})],
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PL",),
            seller=seller,
            invoice_no="SKYM-INV-001",
        )
        result = preview(request)
        assert result.status == "success"

        p = result.preview
        assert p is not None
        layout = getattr(p, "layout", {})
        info = layout.get("info", {}) if isinstance(layout, dict) else {}
        top = layout.get("top", {}) if isinstance(layout, dict) else {}
        assert info.get("left", []) == ["shipping_mark", "shipping_mark_2"]
        assert info.get("right", []) == expected_right
        assert top.get("left", []) == ["seller_info"]

    def test_preview_emax_pi_includes_custom_totals_from_mapping(self, tmp_path):
        path = make_base_file(
            tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[basic_po_row()]
        )
        request = DocumentRequest(
            base_file=str(path),
            po_no="4500030844",
            documents=("PI",),
            seller="EMAX PTE",
            output_dir=str(tmp_path / "out"),
        )
        result = preview(request)
        assert result.status == "success"
        p = result.preview
        assert p is not None

        totals = getattr(p, "totals", {})
        assert totals["signature"] == "Joyce"
        assert totals["Date"] == date.today().strftime("%Y-%m-%d")

        extra_items = totals.get("_extra_items", [])
        assert {
            "key": "signature",
            "label": "Signature",
            "value": "Joyce",
            "source_type": "template_content",
            "rule": "mapping.totals 固定值",
        } in extra_items
        assert {
            "key": "Date",
            "label": "Date",
            "value": date.today().strftime("%Y-%m-%d"),
            "source_type": "system_generated",
            "rule": "系统生成当前日期",
        } in extra_items

        footer_items = totals.get("_footer_items", [])
        assert footer_items == [
            {"key": "amount", "label": "Total Amount", "value": "$3,800.00"},
            {"key": "signature", "label": "Signature", "value": "Joyce"},
            {"key": "Date", "label": "Date", "value": date.today().strftime("%Y-%m-%d")},
        ]

        entries = getattr(p, "source_entries", [])
        amount_entry = next(e for e in entries if e["preview_field"] == "totals.amount")
        signature_entry = next(e for e in entries if e["preview_field"] == "totals.signature")
        date_entry = next(e for e in entries if e["preview_field"] == "totals.Date")
        assert amount_entry["value"] == "$3,800.00"
        assert amount_entry["source_type"] == "computed"
        assert signature_entry["value"] == "Joyce"
        assert signature_entry["source_type"] == "template_content"
        assert date_entry["value"] == date.today().strftime("%Y-%m-%d")
        assert date_entry["source_type"] == "system_generated"


class TestInvoiceGroupPreview:
    def test_preview_combines_rows_from_multiple_pos(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(**{"PO NO.": "PO-1", "SHIP QTY": 100}),
                basic_po_row(**{"PO NO.": "PO-2", "ITEM LINE#": "20", "SHIP QTY": 50}),
            ],
        )
        snapshot = build_workbook_snapshot(str(path))
        group = snapshot.invoice_summary[0]

        result = preview_invoice_group_from_snapshot(
            snapshot,
            group.invoice_group_key,
            seller="GS PTE",
            document="INVOICE",
        )

        assert result.status == "success", result.errors
        assert result.preview is not None
        assert result.preview.po_no == "PO-1, PO-2"
        assert result.preview.invoice_no == "INV-001"
        assert len(result.preview.lines) == 2

    def test_preview_allows_cross_po_ship_to_differences(self, tmp_path):
        rows = [
            basic_po_row(**{"PO NO.": "PO-1"}),
            basic_po_row(**{"PO NO.": "PO-2", "ITEM LINE#": "20"}),
        ]
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=rows,
            customer_po_rows=[
                {
                    "Purchasing Document": "PO-1",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                    "ship to": "Destination A",
                },
                {
                    "Purchasing Document": "PO-2",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                    "ship to": "Destination B",
                },
            ],
        )
        snapshot = build_workbook_snapshot(str(path))
        group = snapshot.invoice_summary[0]

        result = preview_invoice_group_from_snapshot(
            snapshot,
            group.invoice_group_key,
            seller="GS PTE",
            document="INVOICE",
        )

        assert result.status == "success"

    def test_export_invoice_group_returns_combined_xlsx(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(**{"PO NO.": "PO-1", "SHIP QTY": 100}),
                basic_po_row(**{"PO NO.": "PO-2", "ITEM LINE#": "20", "SHIP QTY": 50}),
            ],
        )
        snapshot = build_workbook_snapshot(str(path))
        group = snapshot.invoice_summary[0]

        result = export_invoice_group_from_snapshot(
            snapshot,
            group.invoice_group_key,
            seller="GS PTE",
            documents=("INVOICE", "PL"),
            output_dir=str(tmp_path / "invoice-export"),
        )

        assert result.status == "success", result.errors
        assert result.output_file is not None
        assert Path(result.output_file).suffix == ".xlsx"
        assert result.files == ("GS_PTE-RO-INVOICE&PL-INV-001.xlsx",)
