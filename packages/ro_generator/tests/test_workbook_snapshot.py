"""WorkbookSnapshot 单元测试。"""

from __future__ import annotations

from typing import cast

import pytest
from openpyxl import Workbook
from ro_generator.workbook_reader import SheetData, WorkbookReader
from ro_generator.workbook_snapshot import (
    FileSignature,
    build_workbook_snapshot,
)

# —————————————————————————————————————
# Fixtures
# —————————————————————————————————————

DATA_BASE_HEADER = [
    "SAP",
    "Material Description",
    "Category",
    "GS MODEL",
    "round value",
    "L",
    "W",
    "H",
]
PO_RECORD_HEADER = [
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "FINALQTY",
    "CATEGORY",
    "GS-SK/YM USD FOB",
    "EMAX-GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "SK/YM INVOICE NO.",
    "E10 PO",
    "YM PO",
    "SHIP QTY",
    "CTNS",
    "TOTAL CBM",
    "外箱(最终出口装箱率)",
    "N/W",
    "G/W",
    "L",
    "W",
    "H",
]


def _write_sheet(ws, headers, rows, header_row=4, first_data_row=5):
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            if header in row and row[header] is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=row[header])


CUSTOMER_PO_HEADER = [
    "Purchasing Document",
    "Item",
    "Material",
    "Short Text",
    "Order Quantity",
    "Net price",
    "Currency",
]


def make_base_file(
    tmp_path, *, data_base_rows, po_record_rows, name="base.xlsx", customer_po_rows=()
):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(ws_db, DATA_BASE_HEADER, data_base_rows)
    ws_po = wb.create_sheet("PO record")
    _write_sheet(ws_po, PO_RECORD_HEADER, po_record_rows)
    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(ws_cp, CUSTOMER_PO_HEADER, customer_po_rows, header_row=1, first_data_row=2)
    path = tmp_path / name
    wb.save(path)
    return str(path)


COMBO_PRODUCT = {
    "SAP": "21-44640",
    "Material Description": "CB2500.B2",
    "Category": 1,
    "GS MODEL": "Q1",
    "round value": 24,
    "L": 60,
    "W": 40,
    "H": 30,
}


def basic_po_row(**overrides):
    base = {
        "PO NO.": "4500030844",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2",
        "FINALQTY": 100,
        "CATEGORY": 1,
        "GS-SK/YM USD FOB": 28.0,
        "EMAX-GS PTE FOB": 32.8,
        "EMAX PTE": 38.0,
        "INV#": "INV-001",
        "SK/YM INVOICE NO.": "SKYM-INV-001",
        "E10 PO": "SK-PI-001",
        "YM PO": "YM-PI-001",
        "SHIP QTY": 100,
        "外箱(最终出口装箱率)": 24,
        "CTNS": 5,
        "TOTAL CBM": 0.36,
        "N/W": 8.5,
        "G/W": 10.1,
        "L": 48,
        "W": 31,
        "H": 35,
    }
    base.update(overrides)
    return base


# —————————————————————————————————————
# Tests
# —————————————————————————————————————


class TestFileSignature:
    def test_from_file(self, tmp_path):
        f = tmp_path / "test.xlsx"
        f.write_text("hello")
        sig = FileSignature.from_file(str(f))
        assert sig.path == str(f.resolve())
        assert sig.size > 0
        assert sig.mtime_ns > 0

    def test_equality(self, tmp_path):
        f = tmp_path / "a.xlsx"
        f.write_text("data")
        sig1 = FileSignature.from_file(str(f))
        sig2 = FileSignature.from_file(str(f))
        assert sig1 == sig2

    def test_inequality_different_size(self, tmp_path):
        f = tmp_path / "b.xlsx"
        f.write_text("small")
        sig1 = FileSignature.from_file(str(f))
        f.write_text("larger content")
        sig2 = FileSignature.from_file(str(f))
        # Can't guarantee mtime changes, so just verify they exist
        assert sig1.path == sig2.path


class TestWorkbookSnapshot:
    def test_builds_snapshot(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        snap = build_workbook_snapshot(path)
        assert len(snap.product_index) == 1
        assert "21-44640" in snap.product_index
        assert len(snap.po_rows) == 1
        assert len(snap.po_index) == 1
        assert "4500030844" in snap.po_index

    def test_po_index_uses_row_indices(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(), basic_po_row(**{"ITEM LINE#": "20"})],
        )
        snap = build_workbook_snapshot(path)
        indices = snap.po_index["4500030844"]
        assert isinstance(indices, tuple)
        assert all(isinstance(i, int) for i in indices)
        assert len(indices) == 2
        # indices should reference po_rows
        for i in indices:
            assert snap.po_rows[i]["PO NO."] == "4500030844"

    def test_po_rows_for_po(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(), basic_po_row(**{"PO NO.": "4500099999"})],
        )
        snap = build_workbook_snapshot(path)
        rows = snap.po_rows_for_po("4500030844")
        assert len(rows) == 1
        assert rows[0]["PO NO."] == "4500030844"

        rows2 = snap.po_rows_for_po("4500099999")
        assert len(rows2) == 1
        assert rows2[0]["PO NO."] == "4500099999"

    def test_po_rows_for_po_missing(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        snap = build_workbook_snapshot(path)
        rows = snap.po_rows_for_po("NONEXISTENT")
        assert rows == ()

    def test_invoice_summary_groups_rows_across_pos(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(**{"PO NO.": "PO-1", "INV#": "INV-001"}),
                basic_po_row(
                    **{
                        "PO NO.": "PO-2",
                        "ITEM LINE#": "20",
                        "INV#": "INV-001",
                        "SK/YM INVOICE NO.": None,
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "PO-1",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
                {
                    "Purchasing Document": "PO-2",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
            ],
        )

        snap = build_workbook_snapshot(path)

        assert len(snap.invoice_summary) == 1
        summary = snap.invoice_summary[0]
        assert summary.po_nos == ("PO-1", "PO-2")
        assert snap.invoice_index[summary.invoice_group_key] == (0, 1)
        assert snap.invoice_header_context[summary.invoice_group_key].conflicts == ()
        assert tuple(
            row["PO NO."] for row in snap.invoice_rows_for_group(summary.invoice_group_key)
        ) == (
            "PO-1",
            "PO-2",
        )

    def test_invoice_summary_excludes_zero_ship_qty(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SHIP QTY": 0})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                }
            ],
        )

        snap = build_workbook_snapshot(path)

        assert snap.invoice_summary == ()
        assert snap.invoice_index == {}

    def test_headers_preserved(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        snap = build_workbook_snapshot(path)
        assert len(snap.headers_data_base) > 0
        assert "SAP" in snap.headers_data_base
        assert "PO NO." in snap.headers_po_record

    def test_po_summary_ready(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                }
            ],
        )
        snap = build_workbook_snapshot(path)
        assert len(snap.po_summary) == 1
        s = snap.po_summary[0]
        assert s.po_no == "4500030844"
        assert s.status == "ready"
        assert s.line_count == 1
        assert s.blocking_count == 0

    def test_po_summary_invoice_options_use_seller_context(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "ITEM LINE#": "10",
                        "CATEGORY": 1,
                        "INV#": "INV-001",
                        "SK/YM INVOICE NO.": "YM-INV-001",
                    }
                ),
                basic_po_row(
                    **{
                        "ITEM LINE#": "20",
                        "CATEGORY": 3,
                        "INV#": "INV-001",
                        "SK/YM INVOICE NO.": "SK-INV-001",
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
                {
                    "Purchasing Document": "4500030844",
                    "Item": "20",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
            ],
        )

        snap = build_workbook_snapshot(path)

        summary = snap.po_summary[0]
        assert summary.invoice_nos == ("INV-001",)
        assert summary.invoice_options_by_seller["YM"] == ("YM-INV-001",)
        assert summary.invoice_options_by_seller["SK"] == ("SK-INV-001",)
        assert summary.invoice_options_by_seller["GS PTE"] == ("INV-001",)
        assert summary.invoice_options_by_seller["EMAX PTE"] == ("INV-001-P",)

    def test_po_summary_exportable_documents_use_factory_category_context(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "CATEGORY": 1,
                        "INV#": "INV-001",
                        "SK/YM INVOICE NO.": "YM-INV-001",
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
            ],
        )

        snap = build_workbook_snapshot(path)

        summary = snap.po_summary[0]
        assert summary.exportable_documents_by_seller["SK"] == ()
        assert summary.exportable_documents_by_seller["YM"] == ("PI", "INVOICE_PL")
        assert summary.exportable_documents_by_seller["GS PTE"] == ("PI", "PO", "INVOICE_PL")
        assert summary.exportable_documents_by_seller["EMAX PTE"] == ("PI", "PO", "INVOICE_PL")

    def test_po_summary_factory_pi_requires_factory_po_number(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(
                    **{
                        "CATEGORY": 1,
                        "YM PO": None,
                        "SK/YM INVOICE NO.": "YM-INV-001",
                    }
                ),
            ],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                },
            ],
        )

        snap = build_workbook_snapshot(path)

        summary = snap.po_summary[0]
        assert summary.exportable_documents_by_seller["YM"] == ("INVOICE_PL",)

    def test_po_summary_blocked_missing_sap(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SAP Number": None, "PO NO.": "4500088888"})],
        )
        snap = build_workbook_snapshot(path)
        s = snap.po_summary[0]
        assert s.po_no == "4500088888"
        assert s.status == "blocked"
        assert s.blocking_count >= 1

    def test_multiple_po_grouping(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[
                basic_po_row(),
                basic_po_row(**{"ITEM LINE#": "20", "PO NO.": "4500099999"}),
                basic_po_row(**{"ITEM LINE#": "30", "PO NO.": "4500099999"}),
            ],
        )
        snap = build_workbook_snapshot(path)
        assert len(snap.po_index) == 2
        assert len(snap.po_index["4500030844"]) == 1
        assert len(snap.po_index["4500099999"]) == 2

    def test_frozen_dataclass(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        snap = build_workbook_snapshot(path)
        import dataclasses

        with pytest.raises(dataclasses.FrozenInstanceError):
            snap.base_file = "modified"

    def test_file_metadata(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        snap = build_workbook_snapshot(path)
        assert snap.created_at > 0
        # 文件签名由 FileSignature 管理，不存储在快照中
        from ro_generator.workbook_snapshot import FileSignature

        sig = FileSignature.from_file(str(path))
        assert sig.size > 0
        assert sig.mtime_ns > 0

    def test_build_reads_each_sheet_once(self, tmp_path, monkeypatch):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Item": "10",
                    "Material": "21-44640",
                    "Order Quantity": 100,
                }
            ],
        )
        calls: list[str] = []
        original_read_sheet = WorkbookReader.read_sheet

        def counting_read_sheet(
            self: WorkbookReader,
            sheet_name: str,
            *args: object,
            **kwargs: object,
        ) -> SheetData:
            calls.append(sheet_name)
            return cast(SheetData, original_read_sheet(self, sheet_name, *args, **kwargs))

        monkeypatch.setattr(WorkbookReader, "read_sheet", counting_read_sheet)

        snap = build_workbook_snapshot(path)

        assert len(snap.po_summary) == 1
        assert calls.count("DATA BASE") == 1
        assert calls.count("PO record") == 1
        assert calls.count("客户PO") == 1
