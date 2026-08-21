"""Template mapping 测试：YAML 解析 + 引用校验。"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from ro_generator.errors import MappingError, TemplateError
from ro_generator.template_mapping import (
    LineColumns,
    TemplateMapping,
    TotalCell,
    load_template_mapping,
)


def _preview_info_left(mapping: TemplateMapping) -> object:
    layout = mapping.preview_content.get("layout")
    assert isinstance(layout, dict)
    info = layout.get("info")
    assert isinstance(info, dict)
    return info.get("left")


# ————————————————————————————————————————
# 真实 GS Invoice mapping 加载
# ————————————————————————————————————————


REPO_ROOT = Path(__file__).resolve().parents[3]
RO_TEMPLATES = REPO_ROOT / "customer_profiles" / "ro" / "templates"
PF_TEMPLATES = REPO_ROOT / "customer_profiles" / "pf" / "templates"
GS_INVOICE_MAPPING = RO_TEMPLATES / "gs" / "mappings" / "invoice.yaml"
GS_PI_MAPPING = RO_TEMPLATES / "gs" / "mappings" / "pi.yaml"
GS_PO_MAPPING = RO_TEMPLATES / "gs" / "mappings" / "po.yaml"
EMAX_PI_MAPPING = RO_TEMPLATES / "emax" / "mappings" / "pi.yaml"
EMAX_PO_MAPPING = RO_TEMPLATES / "emax" / "mappings" / "po.yaml"
EMAX_INVOICE_MAPPING = RO_TEMPLATES / "emax" / "mappings" / "invoice.yaml"
EMAX_PL_MAPPING = RO_TEMPLATES / "emax" / "mappings" / "pl.yaml"
SK_PI_MAPPING = RO_TEMPLATES / "sk" / "mappings" / "pi.yaml"
SK_INVOICE_MAPPING = RO_TEMPLATES / "sk" / "mappings" / "invoice.yaml"
SK_PL_MAPPING = RO_TEMPLATES / "sk" / "mappings" / "pl.yaml"
YM_PI_MAPPING = RO_TEMPLATES / "ym" / "mappings" / "pi.yaml"
YM_INVOICE_MAPPING = RO_TEMPLATES / "ym" / "mappings" / "invoice.yaml"
YM_PL_MAPPING = RO_TEMPLATES / "ym" / "mappings" / "pl.yaml"
PF_GS_PL_MAPPING = PF_TEMPLATES / "gs" / "mappings" / "pl.yaml"
PF_EMAX_PL_MAPPING = PF_TEMPLATES / "emax" / "mappings" / "pl.yaml"
PF_GS_PI_MAPPING = PF_TEMPLATES / "gs" / "mappings" / "pi.yaml"
PF_SK_PI_MAPPING = PF_TEMPLATES / "sk" / "mappings" / "pi.yaml"
PF_YM_PI_MAPPING = PF_TEMPLATES / "ym" / "mappings" / "pi.yaml"


def _font_is_red(font: Font) -> bool:
    color = font.color
    if color is None:
        return False
    if color.type == "rgb" and str(color.rgb).upper() in {"FFFF0000", "FF0000", "00FF0000"}:
        return True
    return bool(color.type == "indexed" and color.indexed == 10)


class TestRealGsInvoiceMapping:
    """对仓库内实际 mapping 做端到端验证。"""

    def test_loads_without_error(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        assert isinstance(mapping, TemplateMapping)

    def test_top_level_fields(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        assert mapping.document == "INVOICE"
        assert mapping.template_version == "2026.06"
        assert mapping.sheet == "INV"
        assert mapping.template_path.name == "invoice&pl.xlsx"
        assert mapping.template_path.exists()

    def test_header_section(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        assert mapping.header["invoice_no"] == "H6"
        assert mapping.header["invoice_date"] == "H7"

    def test_lines_section(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        assert mapping.lines.start_row == 15
        assert mapping.lines.style_source_row == 15
        cols = mapping.lines.columns
        assert isinstance(cols, LineColumns)
        # 新版 GS Invoice 列布局 (2026.06 模板)
        assert cols.po_no == "B"
        assert cols.description == "C"
        assert cols.sap == "D"
        assert cols.unit_price == "E"
        assert cols.quantity == "F"
        assert cols.amount == "H"

    def test_totals_section(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        assert mapping.totals["quantity"] == TotalCell(cell="F20")
        assert mapping.totals["amount"] == TotalCell(cell="H20")


class TestRoInvoicePreviewHeaders:
    """RO Invoice 预览抬头必须跟 Excel 模板单元格一致，不能抄成 COMMERCIAL INVOICE。"""

    def test_gs_invoice_header_comes_from_template(self) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)

        assert mapping.preview_static_values["title"] == ("INVOICE",)
        assert mapping.preview_static_values["seller_info"] == (
            "GLOBALSINO PTE.LTD.",
            "10 KAKI BUKIT ROAD 2, #01-37, FIRST EAST CENTRE, SINGAPORE 417868",
        )
        assert mapping.header["bill_to"] == "A6"
        assert mapping.header_fixed["bill_to"] == "TO:E MAX SPORT PTE. LTD."
        assert mapping.header_fixed["bill_to_line2"] == (
            "8 KAKI BUKIT AVENUE 4, #08-32,PREMIER @ KAKI BUKIT,"
        )
        assert mapping.header_fixed["bill_to_line3"] == "SINGAPORE 415875"
        assert mapping.header_fixed["shipped_per"] == "Shipped per ss/mv"
        assert _preview_info_left(mapping) == [
            "bill_to",
            "bill_to_line2",
            "bill_to_line3",
            "shipped_per",
            "from",
            "to",
        ]

    @pytest.mark.parametrize(
        ("mapping_path", "expected_company"),
        [
            (
                SK_INVOICE_MAPPING,
                "GUANGDONG GLOBALSINO OUTDOOR SPORTS EQUIPMENT LIMITED",
            ),
            (YM_INVOICE_MAPPING, "WEIHAI E-MAX SPORT APPARATUS CO.,LTD"),
        ],
    )
    def test_sk_ym_invoice_header_comes_from_template(
        self,
        mapping_path: Path,
        expected_company: str,
    ) -> None:
        mapping = load_template_mapping(mapping_path)

        assert mapping.preview_static_values["title"] == ("INVOICE",)
        assert mapping.preview_static_values["seller_info"][0] == expected_company
        assert mapping.header_fixed["bill_to"] == "TO:GLOBALSINO PTE.LTD."
        assert mapping.header_fixed["bill_to_line2"] == (
            "10 KAKI BUKIT ROAD 2, #01-37, FIRST EAST CENTRE,"
        )
        assert mapping.header_fixed["bill_to_line3"] == "SINGAPORE 417868"
        assert mapping.header_fixed["shipped_per"] == "Shipped per ss/mv"
        assert _preview_info_left(mapping) == [
            "bill_to",
            "bill_to_line2",
            "bill_to_line3",
            "shipped_per",
            "from",
            "to",
        ]


class TestRealMappingsTableHeaderRows:
    @pytest.mark.parametrize(
        ("mapping_path", "expected_rows"),
        [
            (GS_PI_MAPPING, [19]),
            (GS_PO_MAPPING, [19]),
            (EMAX_INVOICE_MAPPING, [12, 13, 14]),
            (EMAX_PI_MAPPING, [17]),
            (EMAX_PO_MAPPING, [19]),
            (EMAX_PL_MAPPING, [9]),
            (SK_PI_MAPPING, [19]),
            (SK_PL_MAPPING, [8]),
            (YM_PI_MAPPING, [19]),
            (YM_PL_MAPPING, [8]),
        ],
    )
    def test_declares_table_header_row(self, mapping_path: Path, expected_rows: list[int]) -> None:
        mapping = load_template_mapping(mapping_path)
        assert mapping.table_header_row == expected_rows


class TestPreviewColumnHeaders:
    @pytest.mark.parametrize("mapping_path", [SK_PI_MAPPING, YM_PI_MAPPING])
    def test_sk_ym_pi_uses_exact_template_header(self, mapping_path: Path) -> None:
        mapping = load_template_mapping(mapping_path)

        assert dict(mapping.preview_column_labels)["A"] == "Country of The Origin"

    def test_emax_pi_uses_exact_item_header(self) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)

        assert dict(mapping.preview_column_labels)["item_number"] == "Item Number"

    def test_multiline_invoice_headers_come_from_declared_template_rows(self) -> None:
        mapping = load_template_mapping(EMAX_INVOICE_MAPPING)

        labels = dict(mapping.preview_column_labels)
        assert labels["description"] == "GOODS\nDESCRIPT"
        assert labels["unit_price"] == "REEL\nFOB, USD"
        assert labels["unit_label"] == ""
        assert labels["amount"] == "TOTAL\nAMOUNT"
        assert mapping.preview_column_letters["unit_label"] == "G"

    def test_merged_pf_pl_headers_are_flattened_with_line_breaks(self) -> None:
        mapping = load_template_mapping(PF_GS_PL_MAPPING)

        labels = dict(mapping.preview_column_labels)
        assert labels["sap"] == "DESCRIPTION OF GOODS\nITEM#"
        assert labels["description"] == "DESCRIPTION"
        assert labels["quantity"] == "QUANTITY\n(PCS)"
        assert labels["width"] == "W"
        assert labels["carton_from"] == "CTN#\nFr"
        assert labels["carton_to"] == "To"

    @pytest.mark.parametrize(
        "mapping_path",
        [PF_GS_PL_MAPPING, PF_EMAX_PL_MAPPING],
    )
    def test_pf_pl_exposes_merged_header_rows(self, mapping_path: Path) -> None:
        mapping = load_template_mapping(mapping_path)

        assert len(mapping.preview_header_rows) == 3
        first_row = list(mapping.preview_header_rows[0])
        assert {cell["label"] for cell in first_row} >= {
            "CTN#",
            "PO#",
            "DESCRIPTION OF GOODS",
            "MEASUREMENT",
        }
        description_group = next(
            cell for cell in first_row if cell["label"] == "DESCRIPTION OF GOODS"
        )
        measurement_group = next(cell for cell in first_row if cell["label"] == "MEASUREMENT")
        assert description_group["colspan"] == 2
        assert description_group["rowspan"] == 2
        assert measurement_group["colspan"] == 3


class TestPreviewDocumentHeaders:
    @pytest.mark.parametrize(
        ("mapping_path", "expected_company"),
        [
            (
                PF_SK_PI_MAPPING,
                "GUANGDONG GLOBALSINO OUTDOOR SPORTS EQUIPMENT LIMITED",
            ),
            (PF_YM_PI_MAPPING, "WEIHAI E-MAX SPORT APPARATUS CO.,LTD"),
        ],
    )
    def test_pf_sk_ym_pi_header_comes_from_template(
        self,
        mapping_path: Path,
        expected_company: str,
    ) -> None:
        mapping = load_template_mapping(mapping_path)

        assert mapping.preview_static_values["title"] == ("PROFORMA INVOICE",)
        assert mapping.preview_static_values["seller_info"][0] == expected_company
        assert mapping.preview_header_labels["pi_no"] == "PI Number:"
        assert mapping.preview_header_labels["etd_baseline"] == "ETD (Baseline Date for FOB Term):"
        assert mapping.preview_header_labels["manufacturer"] == ("Actual Manufacturer Company Name")
        assert mapping.preview_content["layout"]

    @pytest.mark.parametrize("mapping_path", [PF_SK_PI_MAPPING, PF_YM_PI_MAPPING])
    def test_pf_sk_ym_pi_hash_addresses_are_not_truncated(self, mapping_path: Path) -> None:
        mapping = load_template_mapping(mapping_path)

        assert mapping.header_fixed["bill_to_line2"] == (
            "10 KAKI BUKIT ROAD 2, #01-37, FIRST EAST CENTRE"
        )
        assert mapping.header_fixed["ship_to_line2"] == (
            "10 KAKI BUKIT ROAD 2, #01-37, FIRST EAST CENTRE"
        )

    def test_pf_continuation_rows_do_not_borrow_the_opposite_column_label(self) -> None:
        mapping = load_template_mapping(PF_GS_PI_MAPPING)

        assert mapping.preview_header_labels["bill_to_line2"] == ""
        assert mapping.preview_header_labels["ship_to_line2"] == ""
        assert mapping.preview_header_labels["manufacturer_address_2"] == ""

    def test_pf_gs_pi_declares_material_column_and_customer_creation_date(self) -> None:
        mapping = load_template_mapping(PF_GS_PI_MAPPING)

        assert mapping.template_version == "pf_2026.2"
        assert mapping.lines.columns.item_number == "D"
        assert mapping.lines.columns.sap is None
        assert dict(mapping.preview_column_labels)["item_number"] == "Item Number"
        assert mapping.header_fixed["ex_factory_date"] == "SEE BELOW"
        assert mapping.header_fixed["manufacturer_code"] == "134102"
        assert "manufacturer" not in mapping.header_fixed
        assert "manufacturer_address" not in mapping.header_fixed
        assert mapping.totals["signature_date"].value_mode == "model_date"
        assert mapping.lines.start_row == 20
        assert mapping.totals["subtotal"].cell == "H21"
        assert mapping.totals["signature"].cell == "G24"

    def test_pf_gs_pi_template_keeps_reserved_line_row_without_source_notes(self) -> None:
        workbook = load_workbook(PF_TEMPLATES / "gs" / "pi.xlsx")
        sheet = workbook["Sheet1"]
        try:
            assert sheet["F21"].value is not None and "Sub-Total" in str(sheet["F21"].value)
            assert sheet["G22"].value is not None and "TOTAL EXCLUDING EXCISE TAX" in str(
                sheet["G22"].value
            )
            assert sheet["F24"].value == "Signature:"
            assert sheet["F25"].value == "Date:"
            assert sheet["G15"].value in (None, "")
            assert sheet["G16"].value in (None, "")
            assert sheet["G17"].value in (None, "")
            assert sheet["A20"].border.left.style == "thin"
            assert sheet["F20"].number_format == "#,##0.00"
            assert sheet["I20"].number_format == "dd/mmm/yy"
            leaked = [
                f"{cell.coordinate}={cell.value!r}"
                for row in sheet.iter_rows(min_row=1, max_row=30, max_col=13)
                for cell in row
                if isinstance(cell.value, str)
                and any(
                    marker in cell.value
                    for marker in ("new PO template", "固定值", "空着", "DATA BASE BK")
                )
            ]
            assert leaked == []
        finally:
            workbook.close()

    def test_pf_emax_pi_declares_screenshot_header_rules(self) -> None:
        mapping = load_template_mapping(PF_TEMPLATES / "emax" / "mappings" / "pi.yaml")

        assert mapping.header_fixed["ex_factory_date"] == "SEE BELOW"
        assert mapping.totals["signature_date"].value_mode == "model_date"

    def test_pf_ym_pi_manufacturer_address_matches_updated_template(self) -> None:
        mapping = load_template_mapping(PF_YM_PI_MAPPING)
        workbook = load_workbook(PF_TEMPLATES / "ym" / "pi.xlsx")
        sheet = workbook["SHEET1"]
        try:
            assert mapping.template_version == "pf_2026.2"
            assert mapping.header_fixed["manufacturer"] == ("WEIHAI E-MAX SPORT APPARATUS CO.LTD")
            assert mapping.header_fixed["manufacturer_address"] == "NO.25 TONGYI NORTH ROAD,"
            assert mapping.header_fixed["manufacturer_address_2"] == (
                "HUANCUI DISTRICT, WEIHAI, SHANDONG, CHINA."
            )
            assert sheet["G16"].value == "NO.25 TONGYI NORTH ROAD,"
            assert sheet["G17"].value == "HUANCUI DISTRICT, WEIHAI, SHANDONG, CHINA."
            assert sheet["F21"].value is not None and "Sub-Total" in str(sheet["F21"].value)
            assert sheet["F24"].value == "Signature:"
        finally:
            workbook.close()

    def test_pf_gs_invoice_body_and_cost_breakdown_use_arial_9(self) -> None:
        mapping = load_template_mapping(PF_TEMPLATES / "gs" / "mappings" / "invoice.yaml")
        workbook = load_workbook(PF_TEMPLATES / "gs" / "invoice.xlsx")
        sheet = workbook["Sheet1"]
        try:
            assert mapping.template_version == "pf_2026.2"
            for addr in ("E20", "A29", "B30", "C30", "F30", "B31", "E31", "F31"):
                font = sheet[addr].font
                assert font.name == "Arial", addr
                assert font.size == 9, addr
        finally:
            workbook.close()

    @pytest.mark.parametrize("mapping_path", [PF_GS_PL_MAPPING, PF_EMAX_PL_MAPPING])
    def test_pf_pl_style_and_total_rows_use_black_font(self, mapping_path: Path) -> None:
        mapping = load_template_mapping(mapping_path)
        workbook = load_workbook(mapping.template_path)
        sheet = workbook[mapping.sheet]
        try:
            assert mapping.template_version == "pf_2026.3"
            columns = mapping.lines.columns
            addrs = [
                f"{letter}{mapping.lines.style_source_row}"
                for letter in (
                    columns.carton_from,
                    columns.carton_to,
                    columns.po_no,
                    columns.sap,
                    columns.description,
                    columns.quantity,
                    columns.carton_count,
                    columns.net_weight,
                    columns.gross_weight,
                    columns.length,
                    columns.width,
                    columns.height,
                    columns.cbm,
                )
                if letter is not None
            ]
            addrs.extend(total.cell for total in mapping.totals.values())
            for addr in addrs:
                assert not _font_is_red(sheet[addr].font), addr
        finally:
            workbook.close()

    def test_pf_gs_po_declares_material_column_and_customer_creation_date(self) -> None:
        mapping = load_template_mapping(PF_TEMPLATES / "gs" / "mappings" / "po.yaml")

        assert mapping.lines.columns.item_number == "D"
        assert mapping.lines.columns.sap is None
        assert dict(mapping.preview_column_labels)["item_number"] == "Item Number"
        assert mapping.header_fixed["incoterm"] == "FOB Qingdao"
        assert mapping.header_fixed["ex_factory_date"] == "SEE BELOW"
        assert mapping.totals["signature_date"].value_mode == "model_date"


# ————————————————————————————————————————
# 合成 fixture：可控的错误注入测试
# ————————————————————————————————————————


def make_template(tmp_path: Path, *, max_row: int = 30, max_col: int = 10) -> Path:
    """生成一个最小可识别的"模板"文件（足够覆盖默认 mapping 的引用）。"""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    # 通过写入边界单元格定义 max_row / max_col
    ws.cell(row=max_row, column=max_col, value="x")
    path = tmp_path / "template.xlsx"
    wb.save(path)
    return path


def write_yaml(path: Path, content: str) -> Path:
    path.write_text(content, encoding="utf-8")
    return path


def good_yaml_content(template_rel: str) -> str:
    return f"""\
document: invoice
template_version: "v1"
template: {template_rel}
sheet: Sheet1
header:
  invoice_no: H6
lines:
  start_row: 18
  style_source_row: 19
  columns:
    sap: D
    quantity: F
    unit_price: E
    amount: H
totals:
  amount: H27
"""


@pytest.fixture
def template_file(tmp_path: Path) -> Path:
    return make_template(tmp_path)


@pytest.fixture
def good_yaml(tmp_path: Path, template_file: Path) -> Path:
    yaml_path = tmp_path / "mapping.yaml"
    return write_yaml(yaml_path, good_yaml_content(str(template_file)))


# ————————————————————————————————————————
# 加载与基础解析
# ————————————————————————————————————————


class TestLoadBasic:
    def test_loads_synthetic_mapping(self, good_yaml: Path) -> None:
        mapping = load_template_mapping(good_yaml)
        assert mapping.document == "INVOICE"
        assert mapping.template_version == "v1"
        assert mapping.totals["amount"] == TotalCell(cell="H27")

    def test_loads_declared_total_value_mode(self, tmp_path: Path, template_file: Path) -> None:
        yaml_path = write_yaml(
            tmp_path / "mapping.yaml",
            f"""\
document: invoice
template_version: "v1"
template: {template_file}
sheet: Sheet1
header:
  invoice_no: H6
lines:
  start_row: 18
  style_source_row: 19
  columns:
    sap: D
    quantity: F
    unit_price: E
    amount: H
totals:
  amount: H27
  signature:
    cell: H28
    value_mode: fixed
    value: Joyce
  Date:
    cell: H29
    value_mode: current_date
""",
        )
        mapping = load_template_mapping(yaml_path)
        assert mapping.totals["signature"] == TotalCell(
            cell="H28",
            value_mode="fixed",
            value="Joyce",
        )
        assert mapping.totals["Date"] == TotalCell(cell="H29", value_mode="current_date")

    def test_loads_optional_table_header_row(self, tmp_path: Path, template_file: Path) -> None:
        yaml_path = write_yaml(
            tmp_path / "mapping.yaml",
            f"""\
document: invoice
template_version: "v1"
template: {template_file}
sheet: Sheet1
header:
  invoice_no: H6
table_header_row: 17
lines:
  start_row: 18
  style_source_row: 19
  columns:
    sap: D
    quantity: F
    unit_price: E
    amount: H
totals:
  amount: H27
""",
        )
        mapping = load_template_mapping(yaml_path)
        assert mapping.table_header_row == [17]

    def test_yaml_path_must_exist(self, tmp_path: Path) -> None:
        with pytest.raises(MappingError, match="mapping 文件不存在"):
            load_template_mapping(tmp_path / "nope.yaml")

    def test_yaml_root_must_be_dict(self, tmp_path: Path) -> None:
        path = write_yaml(tmp_path / "mapping.yaml", "- just\n- a list\n")
        with pytest.raises(MappingError, match="根节点必须是 dict"):
            load_template_mapping(path)

    def test_invalid_yaml_syntax(self, tmp_path: Path) -> None:
        path = write_yaml(tmp_path / "mapping.yaml", "key: : invalid")
        with pytest.raises(MappingError, match="YAML 解析失败"):
            load_template_mapping(path)


# ————————————————————————————————————————
# 必填字段校验
# ————————————————————————————————————————


class TestRequiredFields:
    def test_missing_template_version(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file))
        content = content.replace('template_version: "v1"\n', "")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="template_version"):
            load_template_mapping(path)

    def test_missing_template_path_field(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file))
        content = content.replace(f"template: {template_file}\n", "")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="template"):
            load_template_mapping(path)

    def test_template_file_not_found(self, tmp_path: Path) -> None:
        path = write_yaml(
            tmp_path / "mapping.yaml",
            good_yaml_content("/path/to/nonexistent.xlsx"),
        )
        with pytest.raises(MappingError, match="模板文件不存在"):
            load_template_mapping(path)

    def test_unknown_document_type(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "document: invoice", "document: receipt"
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="document 字段必须是"):
            load_template_mapping(path)

    def test_fixed_total_requires_value(self, tmp_path: Path, template_file: Path) -> None:
        path = write_yaml(
            tmp_path / "mapping.yaml",
            f"""\
document: invoice
template_version: "v1"
template: {template_file}
sheet: Sheet1
header:
  invoice_no: H6
lines:
  start_row: 18
  style_source_row: 19
  columns:
    sap: D
    quantity: F
    unit_price: E
    amount: H
totals:
  signature:
    cell: H28
    value_mode: fixed
""",
        )
        with pytest.raises(MappingError, match="value_mode=fixed"):
            load_template_mapping(path)

    def test_table_header_row_must_be_positive_int(
        self, tmp_path: Path, template_file: Path
    ) -> None:
        path = write_yaml(
            tmp_path / "mapping.yaml",
            good_yaml_content(str(template_file)).replace(
                "sheet: Sheet1\n",
                "sheet: Sheet1\ntable_header_row: 0\n",
            ),
        )
        with pytest.raises(
            MappingError, match=r"table_header_row 必须为正整数|table_header_row\[0\] 必须为正整数"
        ):
            load_template_mapping(path)


class TestLinesSection:
    def test_missing_lines_section(self, tmp_path: Path, template_file: Path) -> None:
        content = f"""\
document: invoice
template_version: "v1"
template: {template_file}
sheet: Sheet1
header:
  invoice_no: H6
totals:
  amount: H27
"""
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="lines"):
            load_template_mapping(path)

    def test_negative_start_row(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace("start_row: 18", "start_row: -1")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="必须为正整数"):
            load_template_mapping(path)

    def test_missing_required_column(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace("    quantity: F\n", "")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match=r"lines\.columns 缺少必填项"):
            load_template_mapping(path)

    def test_invalid_column_letter(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace("    sap: D", "    sap: '@@'")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="不是合法列字母"):
            load_template_mapping(path)

    def test_table_header_row_must_be_before_start_row(
        self, tmp_path: Path, template_file: Path
    ) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "sheet: Sheet1\n",
            "sheet: Sheet1\ntable_header_row: 18\n",
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match=r"都必须小于 lines\.start_row"):
            load_template_mapping(path)

    def test_table_header_row_list_form(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "sheet: Sheet1\n",
            "sheet: Sheet1\ntable_header_row:\n  - 16\n  - 17\n",
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        mapping = load_template_mapping(path)
        assert mapping.table_header_row == [16, 17]

    def test_table_header_row_empty_list_rejected(
        self, tmp_path: Path, template_file: Path
    ) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "sheet: Sheet1\n",
            "sheet: Sheet1\ntable_header_row: []\n",
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="不能为空列表"):
            load_template_mapping(path)

    def test_table_header_row_list_bad_item_rejected(
        self, tmp_path: Path, template_file: Path
    ) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "sheet: Sheet1\n",
            "sheet: Sheet1\ntable_header_row:\n  - 17\n  - 0\n",
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match=r"table_header_row\[1\] 必须为正整数"):
            load_template_mapping(path)


# ————————————————————————————————————————
# 模板引用校验（核心防漂移）
# ————————————————————————————————————————


class TestReferenceValidation:
    def test_header_cell_out_of_range_raises(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "  invoice_no: H6", "  invoice_no: H999"
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match=r"header\.invoice_no.*超出模板范围"):
            load_template_mapping(path)

    def test_totals_cell_out_of_range_raises(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace("  amount: H27", "  amount: H999")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match=r"totals\.amount.*超出模板范围"):
            load_template_mapping(path)

    def test_style_source_row_beyond_max(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "style_source_row: 19", "style_source_row: 999"
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="style_source_row"):
            load_template_mapping(path)

    def test_line_column_letter_beyond_max(self, tmp_path: Path, template_file: Path) -> None:
        # 模板默认 max_col=10 (J)。把列改到 Z 即超界
        content = good_yaml_content(str(template_file)).replace("    amount: H", "    amount: Z")
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="超过模板列数"):
            load_template_mapping(path)

    def test_sheet_not_in_template(self, tmp_path: Path, template_file: Path) -> None:
        content = good_yaml_content(str(template_file)).replace(
            "sheet: Sheet1", "sheet: NonExistent"
        )
        path = write_yaml(tmp_path / "mapping.yaml", content)
        with pytest.raises(MappingError, match="找不到 sheet"):
            load_template_mapping(path)

    def test_corrupted_template(self, tmp_path: Path) -> None:
        bogus = tmp_path / "bogus.xlsx"
        bogus.write_bytes(b"not a real xlsx")
        path = write_yaml(
            tmp_path / "mapping.yaml",
            good_yaml_content(str(bogus)),
        )
        with pytest.raises(TemplateError, match="无法打开模板"):
            load_template_mapping(path)


# ————————————————————————————————————————
# 不可变性
# ————————————————————————————————————————


def test_mapping_is_frozen(good_yaml: Path) -> None:
    import dataclasses

    mapping = load_template_mapping(good_yaml)
    with pytest.raises(dataclasses.FrozenInstanceError):
        mapping.template_version = "v2"  # type: ignore[misc]
    with pytest.raises(dataclasses.FrozenInstanceError):
        mapping.lines.start_row = 99  # type: ignore[misc]
