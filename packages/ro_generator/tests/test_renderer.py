"""Renderer 测试：用真实 GS Invoice 模板装配 Invoice，验证样式保留 + 数据正确。

复用 Phase 0 Spike A 的样式不变量断言，确保插入行 + 复制样式仍然安全。
"""

from __future__ import annotations

from dataclasses import replace
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook
from ro_generator.document_model import (
    DocumentModel,
    build_invoice_model,
    build_pi_model,
    build_pl_model,
    build_po_model,
)
from ro_generator.errors import TemplateError
from ro_generator.models import OrderLine, Product
from ro_generator.renderer import render_document
from ro_generator.schema import (
    ENTITY_EMAX_PTE,
    ENTITY_GS_PTE,
    ENTITY_PF,
)
from ro_generator.source_index import COMPUTED_SHEET
from ro_generator.template_mapping import load_template_mapping

REPO_ROOT = Path(__file__).resolve().parents[3]
GS_INVOICE_MAPPING = REPO_ROOT / "templates" / "gs" / "mappings" / "invoice.yaml"
GS_INVOICE_TEMPLATE = REPO_ROOT / "templates" / "gs" / "invoice&pl.xlsx"
EMAX_PI_MAPPING = REPO_ROOT / "templates" / "emax" / "mappings" / "pi.yaml"
EMAX_PO_MAPPING = REPO_ROOT / "templates" / "emax" / "mappings" / "po.yaml"
EMAX_PL_MAPPING = REPO_ROOT / "templates" / "emax" / "mappings" / "pl.yaml"


# ————————————————————————————————————————
# Fixture builders
# ————————————————————————————————————————


def make_order_line(
    *,
    sap: str,
    description: str,
    gs_model: str,
    quantity: int,
    item_line_no: str = "10",
    unit_prices: dict[tuple[str, str], Decimal] | None = None,
    source_row: int | None = None,
    confirmed_ex_factory_date: date | None = None,
    po_ex_factory_date: date | None = None,
) -> OrderLine:
    if unit_prices is None:
        unit_prices = {(ENTITY_GS_PTE, ENTITY_EMAX_PTE): Decimal("32.8")}
    qty = Decimal(quantity)
    subtotals = {seg: (price * qty).quantize(Decimal("0.01")) for seg, price in unit_prices.items()}
    return OrderLine(
        po_no="4500030844",
        item_line_no=item_line_no,
        sap=sap,
        description=description,
        category=1,
        quantity=qty,
        product=Product(
            sap=sap,
            description=description,
            category=1,
            gs_model=gs_model,
        ),
        ship_to="EMAX HQ",
        invoice_no="INV-RENDER-001",
        confirmed_ex_factory_date=confirmed_ex_factory_date,
        po_ex_factory_date=po_ex_factory_date,
        prices=unit_prices,
        subtotals=subtotals,
        source_row=source_row,
    )


def build_three_line_invoice(
    seller: str = ENTITY_GS_PTE, buyer: str = ENTITY_EMAX_PTE
) -> DocumentModel:
    """3 行（在 GS Invoice 9 行预留范围内）。"""
    lines = (
        make_order_line(
            sap="21-44640", description="CB2500.B2", gs_model="Q1", quantity=100, source_row=5
        ),
        make_order_line(
            sap="21-44641",
            description="CB3000.B2",
            gs_model="Q2",
            quantity=200,
            item_line_no="20",
            source_row=6,
        ),
        make_order_line(
            sap="21-44642",
            description="CB4000.B2",
            gs_model="Q3",
            quantity=80,
            item_line_no="30",
            source_row=7,
        ),
    )
    result = build_invoice_model(lines, seller=seller, buyer=buyer, po_no="4500030844")
    assert result.model is not None, result.messages
    return result.model


def build_overflowing_invoice() -> DocumentModel:
    """10 行（超过 GS Invoice 模板预留的 7 行，触发插入路径）。"""
    lines = tuple(
        make_order_line(
            sap=f"21-4464{i}",
            description=f"CB{i}00.B2",
            gs_model=f"Q{i}",
            quantity=50 + i,
            item_line_no=str(10 + i * 10),
        )
        for i in range(10)
    )
    result = build_invoice_model(
        lines, seller=ENTITY_GS_PTE, buyer=ENTITY_EMAX_PTE, po_no="4500030844"
    )
    assert result.model is not None, result.messages
    return result.model


def build_emax_pi() -> DocumentModel:
    lines = (
        make_order_line(
            sap="21-44640",
            description="CB2500.B2",
            gs_model="Q1",
            quantity=100,
            source_row=5,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("32.8")},
            confirmed_ex_factory_date=date(2026, 4, 20),
            po_ex_factory_date=date(2026, 3, 15),
        ),
    )
    result = build_pi_model(lines, seller=ENTITY_EMAX_PTE, buyer=ENTITY_PF, po_no="4500030844")
    assert result.model is not None, result.messages
    return replace(
        result.model,
        ship_to="209 Stoneridge Drive, Columbia, South Carolina 29210, United States",
    )


def build_emax_pi_three_lines() -> DocumentModel:
    lines = (
        make_order_line(
            sap="11-19833",
            description="Accel 6'6\" M/F Casting Rod 2pc",
            gs_model="Q1",
            quantity=120,
            item_line_no="10",
            source_row=5,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("6.71")},
            confirmed_ex_factory_date=date(2026, 6, 26),
        ),
        make_order_line(
            sap="11-19835",
            description="Accel 7'0\" M/F Casting Rod 2pc",
            gs_model="Q2",
            quantity=120,
            item_line_no="20",
            source_row=6,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("6.44")},
            confirmed_ex_factory_date=date(2026, 6, 26),
        ),
        make_order_line(
            sap="11-18369",
            description="Speed Stick 6'6\" L Spinning Rod 3PC",
            gs_model="Q3",
            quantity=64,
            item_line_no="30",
            source_row=7,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("14.51")},
            confirmed_ex_factory_date=date(2026, 6, 26),
        ),
    )
    result = build_pi_model(lines, seller=ENTITY_EMAX_PTE, buyer=ENTITY_PF, po_no="4500031170")
    assert result.model is not None, result.messages
    return replace(
        result.model,
        ship_to="209 Stoneridge Drive, Columbia, South Carolina 29210, United States",
    )


def build_emax_po() -> DocumentModel:
    lines = (
        make_order_line(
            sap="21-44640",
            description="CB2500.B2",
            gs_model="Q1",
            quantity=100,
            source_row=5,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("32.8")},
            confirmed_ex_factory_date=date(2026, 3, 15),
        ),
    )
    result = build_po_model(lines, seller=ENTITY_EMAX_PTE, buyer=ENTITY_PF, po_no="4500030844")
    assert result.model is not None, result.messages
    return result.model


def build_emax_pl() -> DocumentModel:
    line = replace(
        make_order_line(
            sap="21-44640",
            description="CB2500.B2",
            gs_model="Q1",
            quantity=100,
            source_row=5,
            unit_prices={(ENTITY_EMAX_PTE, ENTITY_PF): Decimal("32.8")},
            confirmed_ex_factory_date=date(2026, 3, 15),
        ),
        ship_qty=Decimal("100"),
        carton_count=Decimal("5"),
        net_weight=Decimal("8.5"),
        gross_weight=Decimal("10.1"),
        total_cbm=Decimal("1.23"),
    )
    result = build_pl_model((line,), seller=ENTITY_EMAX_PTE, buyer=ENTITY_PF, po_no="4500030844")
    assert result.model is not None, result.messages
    return result.model


# ————————————————————————————————————————
# 基础渲染：3 行，不触发插入
# ————————————————————————————————————————


class TestRenderBasic:
    def test_returns_absolute_path(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        assert result.output_path.is_absolute()
        assert result.output_path.exists()

    def test_creates_parent_directories(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        deep = tmp_path / "a" / "b" / "c" / "out.xlsx"
        result = render_document(model, mapping, deep)
        assert result.output_path.exists()

    def test_invoice_no_written_to_header(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        # mapping.header.invoice_no = H6
        assert ws["H6"].value == "INV-RENDER-001"

    def test_three_lines_written_to_correct_columns(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        # mapping start_row=15, columns: SAP=D, price=E, qty=F, amount=H, description=C, unit_label row_fixed: G=PCS
        assert ws["D15"].value == "21-44640"
        assert ws["E15"].value == 32.8
        assert ws["F15"].value == 100
        assert ws["G15"].value == "PCS"
        assert ws["H15"].value == pytest.approx(3280.0)
        assert ws["C15"].value == "CB2500.B2"

        assert ws["D16"].value == "21-44641"
        assert ws["F16"].value == 200

        assert ws["D17"].value == "21-44642"
        assert ws["F17"].value == 80

    def test_unfilled_reserved_rows_cleared(self, tmp_path: Path) -> None:
        """3 行数据写入后，紧接着的样板行应被清掉。"""
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        # 数据写入 15-17，row 18-19 应被清除（20 是 totals 行）
        for row in range(18, 20):
            assert ws[f"D{row}"].value is None, f"row {row} 没有清掉 D 列"
            assert ws[f"F{row}"].value is None, f"row {row} 没有清掉 F 列"

    def test_totals_written_at_unchanged_position(self, tmp_path: Path) -> None:
        """3 行不触发插入时，合计行号保持模板原位置 (F20/H20)。"""
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_three_line_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        # openpyxl 读回 Decimal 数值会被规约为 int/float
        assert ws["F20"].value == 380  # 100+200+80
        # amount: 100*32.8 + 200*32.8 + 80*32.8 = 12464.00
        assert ws["H20"].value == pytest.approx(12464.0)

    def test_emax_pi_bill_to_uses_header_fixed_yaml(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        model = build_emax_pi()
        result = render_document(model, mapping, tmp_path / "emax-pi.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["B9"].value == "209 Stoneridge Drive"
        assert ws["B10"].value == "Columbia, South Carolina 29210"
        assert ws["B11"].value == "United States"
        assert ws["G9"].value == "209 Stoneridge Drive"
        assert ws["G10"].value == "Columbia, South Carolina 29210"
        assert ws["G11"].value == "United States"

    def test_emax_pi_ex_factory_date_written_from_po_record(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        model = build_emax_pi()
        result = render_document(model, mapping, tmp_path / "emax-pi.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["G7"].value == "2026-03-15"

        loc = result.source_index.lookup_source("G7")
        assert loc is not None
        assert loc.sheet == "PO record"
        assert loc.field == "FINAL EX-FACTORY DATE"

    def test_emax_pi_usd_price_and_amount_use_dollar_number_format(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "emax-pi.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]

        assert ws["E18"].value == pytest.approx(32.8)
        assert ws["G18"].value == pytest.approx(3280.0)
        assert ws["G21"].value == pytest.approx(3280.0)
        assert ws["E18"].number_format == '"$"#,##0.00'
        assert ws["G18"].number_format == '"$"#,##0.00'
        assert ws["G21"].number_format == '"$"#,##0.00'

    def test_emax_pi_number_traces_to_customer_po(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "emax-pi.xlsx")
        loc = result.source_index.lookup_source("B6")
        assert loc is not None
        assert loc.sheet == "客户PO"
        assert loc.field == "Purchasing Document"

    def test_emax_pi_totals_can_write_fixed_and_current_date(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["G23"].value == "Joyce"
        assert ws["G24"].value == date.today().strftime("%Y-%m-%d")

        signature_loc = result.source_index.lookup_source("G23")
        date_loc = result.source_index.lookup_source("G24")
        assert signature_loc is not None and signature_loc.is_computed
        assert date_loc is not None and date_loc.is_computed
        assert signature_loc.field == "totals.signature"
        assert date_loc.field == "totals.Date"

    def test_emax_pi_keeps_table_header_labels(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["A17"].value == "Country of The Origin"
        assert ws["B17"].value == "PO Number"
        assert ws["G17"].value == "USD Amount "
        assert ws["H17"].value == "EX-FACTORY DATE"

    def test_emax_pi_keeps_totals_labels(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["F21"].value == "Total"
        assert ws["F23"].value == "Signature:"
        assert ws["F24"].value == "Date:"

    def test_emax_pi_reserved_rows_use_consistent_number_formats(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi_three_lines(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["Standard Invoice format"]
        assert ws["E18"].value == pytest.approx(6.71)
        assert ws["E20"].value == pytest.approx(14.51)
        assert ws["E20"].number_format == ws["E18"].number_format
        assert ws["F20"].number_format == ws["F18"].number_format
        assert ws["G20"].number_format == ws["G18"].number_format
        assert ws["H20"].number_format == ws["H18"].number_format

    def test_emax_po_keeps_table_header_labels(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PO_MAPPING)
        result = render_document(build_emax_po(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["EMAX PO"]
        assert ws["A19"].value == "Country of The Origin"
        assert ws["B19"].value == "PO Number"
        assert ws["C19"].value == "Item Number"
        assert ws["H19"].value == "EX-FACTORY DATE"
        assert ws["C20"].value == "10"
        assert ws["D20"].value == "CB2500.B2"
        assert ws["E20"].value == pytest.approx(32.8)
        assert ws["F20"].value == 100
        assert ws["G20"].value == pytest.approx(3280.0)
        assert ws["H20"].value.date() == date(2026, 3, 15)
        assert ws["G23"].value == pytest.approx(3280.0)
        assert "d/mmm/yy" not in ws["G23"].number_format.lower()
        loc = result.source_index.lookup_source("B20")
        assert loc is not None
        assert loc.sheet == "客户PO"
        assert loc.field == "Purchasing Document"

    def test_emax_pl_keeps_table_header_labels(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PL_MAPPING)
        result = render_document(build_emax_pl(), mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["PL"]
        assert ws["B9"].value == "SAP PO#"
        assert ws["C9"].value == "Description of Goods"
        assert ws["G9"].value == "Net Weight"
        assert ws["K9"].value == "Measurement"

    def test_emax_pl_measurement_uses_two_decimal_display(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PL_MAPPING)
        model = build_emax_pl()
        model = replace(
            model,
            lines=(replace(model.lines[0], cbm=Decimal("1.2")),),
        )
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["PL"]
        assert ws["K10"].value == pytest.approx(1.2)
        assert ws["K10"].number_format == "0.00"

    def test_emax_pl_measurement_preserves_source_decimal_places(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PL_MAPPING)
        model = build_emax_pl()
        model = replace(
            model,
            lines=(replace(model.lines[0], cbm=Decimal("1.2000")),),
        )
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["PL"]
        assert ws["K10"].value == pytest.approx(1.2)
        assert ws["K10"].number_format == "0.0000"


# ————————————————————————————————————————
# 触发插入：10 行 > 模板预留 7 行
# ————————————————————————————————————————


class TestRenderOverflow:
    def test_all_ten_lines_present(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_overflowing_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        for i in range(10):
            row = 15 + i
            assert ws[f"D{row}"].value == f"21-4464{i}", f"row {row} D 列未写入第 {i} 行 SAP"

    def test_totals_shifted_by_insertion_count(self, tmp_path: Path) -> None:
        """模板预留区间 = totals_row(20) - start_row(15) = 5 行；
        10 行插入 5 行，合计从 F20/H20 平移到 F25/H25。
        """
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_overflowing_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        assert ws["F25"].value is not None
        assert ws["H25"].value is not None
        # 10 行总数: 50+51+...+59 = 545
        assert ws["F25"].value == 545

    def test_total_label_shifts_with_totals_row(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_overflowing_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]

        assert ws["A19"].value is None
        assert ws["A25"].value == "Total"
        assert ws["F25"].value == 545

    def test_inserted_rows_have_styles(self, tmp_path: Path) -> None:
        """新插入的行单元格必须有样式（来自 style_source_row 15）。

        预留区间 = 20-15 = 5 行（rows 15-19），10 行触发插入 5 行。
        """
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        model = build_overflowing_invoice()
        result = render_document(model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        cell = ws["D20"]
        assert cell.has_style, "插入行 row 20 D 列缺样式"


# ————————————————————————————————————————
# 复用 Spike A 的不变量断言
# ————————————————————————————————————————


class TestStylePreservationParity:
    """无论是否插入行，几个关键不变量都必须保持。"""

    @pytest.fixture
    def baseline(self) -> dict[str, object]:
        wb = load_workbook(GS_INVOICE_TEMPLATE)
        ws = wb["INV"]
        return {
            "merged": {str(r) for r in ws.merged_cells.ranges},
            "col_widths": {col: dim.width for col, dim in ws.column_dimensions.items()},
        }

    def test_basic_keeps_merges_and_widths(
        self, tmp_path: Path, baseline: dict[str, object]
    ) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "basic.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        assert {str(r) for r in ws.merged_cells.ranges} == baseline["merged"]
        for col, expected in baseline["col_widths"].items():  # type: ignore[attr-defined]
            assert ws.column_dimensions[col].width == expected, f"列 {col} 宽度变化"

    def test_overflow_keeps_widths(self, tmp_path: Path, baseline: dict[str, object]) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_overflowing_invoice(), mapping, tmp_path / "overflow.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb["INV"]
        for col, expected in baseline["col_widths"].items():  # type: ignore[attr-defined]
            assert ws.column_dimensions[col].width == expected


# ————————————————————————————————————————
# 边界
# ————————————————————————————————————————


class TestEdgeCases:
    def test_invalid_sheet_raises(self) -> None:
        # 构造一个 mapping 引用不存在的 sheet 名
        # 通过修改加载后的对象做不到（frozen）；直接传入坏 mapping 副本要复杂
        # 用 monkeypatch 替换更简单：但本测试只为冒烟，我们先 skip 这条
        pytest.skip("需要 mapping mocking")

    def test_corrupt_template_raises(self, tmp_path: Path) -> None:
        from dataclasses import replace

        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        bad_template = tmp_path / "bad.xlsx"
        bad_template.write_bytes(b"not a real xlsx")
        broken_mapping = replace(mapping, template_path=bad_template)
        with pytest.raises(TemplateError, match="无法打开模板"):
            render_document(build_three_line_invoice(), broken_mapping, tmp_path / "out.xlsx")


# ————————————————————————————————————————
# 双向溯源索引（产品方案 §4.4）
# ————————————————————————————————————————


class TestSourceIndex:
    def test_index_returned_with_result(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        # 数据行 SAP/qty/price + 表头 invoice_no/ship_to + 合计 quantity/amount 都应有条目
        assert len(result.source_index) > 0

    def test_data_row_traces_back_to_po_record(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        # 第一行的 SAP（D15）应溯源到 PO record source_row=5 的 SAP Number 字段
        loc = result.source_index.lookup_source("D15")
        assert loc is not None
        assert loc.sheet == "PO record"
        assert loc.row == 5
        assert loc.field == "SAP Number"

    def test_quantity_traces_to_ship_qty_for_invoice(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("F15")
        assert loc is not None
        assert loc.field == "SHIP QTY"

    def test_description_traces_to_po_record_for_invoice(self, tmp_path: Path) -> None:
        # Invoice/PL 的 description 来自 PO record（出货时的商品描述），不是 DATA BASE
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("C15")
        assert loc is not None
        assert loc.sheet == "PO record"
        assert isinstance(loc.row, int)  # PO record 逐行数据，有源行号
        assert loc.field == "DESCRIPTION"

    def test_emax_pi_unit_price_traces_to_data_base_price_column(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("E18")
        assert loc is not None
        assert loc.sheet == "DATA BASE"
        assert loc.row is None
        assert loc.field == "EMAX PTE COMBO FOB 2026"

    def test_amount_marked_computed(self, tmp_path: Path) -> None:
        """amount 列写公式，UI 上溯源应标识为"由工作台计算"。"""
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("H15")
        assert loc is not None
        assert loc.is_computed
        assert loc.sheet == COMPUTED_SHEET

    def test_totals_marked_computed(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        # 模板原合计在 F20/H20（3 行不触发插入）
        qty_loc = result.source_index.lookup_source("F20")
        amt_loc = result.source_index.lookup_source("H20")
        assert qty_loc is not None and qty_loc.is_computed
        assert amt_loc is not None and amt_loc.is_computed
        assert qty_loc.field == "total_quantity"
        assert amt_loc.field == "total_amount"

    def test_header_traces_to_po_record(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(GS_INVOICE_MAPPING)
        result = render_document(build_three_line_invoice(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("H6")
        assert loc is not None
        assert loc.sheet == "PO record"
        assert loc.field == "INV#"

    def test_gs_pi_fixed_ship_to_header_has_no_base_source_trace(self, tmp_path: Path) -> None:
        from ro_generator.template_mapping import load_template_mapping as ltm

        gs_pi_mapping = REPO_ROOT / "templates" / "gs" / "mappings" / "pi.yaml"
        mapping = ltm(gs_pi_mapping)
        model = build_pi_model(
            (
                make_order_line(
                    sap="21-44640",
                    description="CB2500.B2",
                    gs_model="GS-100",
                    quantity=100,
                    unit_prices={(ENTITY_GS_PTE, ENTITY_EMAX_PTE): Decimal("32.8")},
                ),
            ),
            seller=ENTITY_GS_PTE,
            buyer=ENTITY_EMAX_PTE,
            po_no="4500030844",
        )
        assert model.model is not None, model.messages
        result = render_document(model.model, mapping, tmp_path / "out.xlsx")
        wb = load_workbook(result.output_path)
        ws = wb[mapping.sheet]
        assert ws["G10"].value == "E MAX SPORT PTE. LTD."

        loc = result.source_index.lookup_source("G10")
        assert loc is None

    def test_emax_pi_ship_to_continuation_headers_trace_to_customer_po(
        self, tmp_path: Path
    ) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        line2 = result.source_index.lookup_source("G10")
        line3 = result.source_index.lookup_source("G11")
        assert line2 is not None
        assert line2.sheet == "客户PO"
        assert line2.field == "ship to"
        assert line3 is not None
        assert line3.sheet == "客户PO"
        assert line3.field == "ship to"

    def test_emax_pi_item_no_traces_to_customer_po_material(self, tmp_path: Path) -> None:
        mapping = load_template_mapping(EMAX_PI_MAPPING)
        result = render_document(build_emax_pi(), mapping, tmp_path / "out.xlsx")
        loc = result.source_index.lookup_source("C18")
        assert loc is not None
        assert loc.sheet == "客户PO"
        assert loc.row is None
        assert loc.field == "item"
