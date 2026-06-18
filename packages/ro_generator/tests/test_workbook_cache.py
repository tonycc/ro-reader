"""WorkbookCacheManager 单元测试。"""

from __future__ import annotations

import threading
import time

from openpyxl import Workbook
from ro_generator.workbook_cache import WorkbookCacheManager

# —————————————————————————————————————
# Fixtures
# —————————————————————————————————————

DATA_BASE_HEADER = [
    "SAP",
    "Material Description",
    "Category",
    "GS MODEL",
    "round value",
    "L",
    "W",
    "H",
]
PO_RECORD_HEADER = [
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "FINALQTY",
    "GS-SK/YM USD FOB",
    "EMAX-GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "SHIP QTY",
    "CTNS",
    "TOTAL CBM",
    "外箱(最终出口装箱率)",
    "N/W",
    "G/W",
    "L",
    "W",
    "H",
]


def _write_sheet(ws, headers, rows, header_row=4, first_data_row=5):
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            if header in row and row[header] is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=row[header])


def make_base_file(tmp_path, *, data_base_rows, po_record_rows, name="base.xlsx"):
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(ws_db, DATA_BASE_HEADER, data_base_rows)
    ws_po = wb.create_sheet("PO record")
    _write_sheet(ws_po, PO_RECORD_HEADER, po_record_rows)
    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(
        ws_cp,
        ["Purchasing Document", "Material", "Order Quantity"],
        [],
        header_row=1,
        first_data_row=2,
    )
    path = tmp_path / name
    wb.save(path)
    return str(path)


COMBO_PRODUCT = {
    "SAP": "21-44640",
    "Material Description": "CB2500.B2",
    "Category": 1,
    "GS MODEL": "Q1",
    "round value": 24,
    "L": 60,
    "W": 40,
    "H": 30,
}


def basic_po_row(**overrides):
    base = {
        "PO NO.": "4500030844",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2",
        "FINALQTY": 100,
        "GS-SK/YM USD FOB": 28.0,
        "EMAX-GS PTE FOB": 32.8,
        "EMAX PTE": 38.0,
        "INV#": "INV-001",
        "SHIP QTY": 100,
        "外箱(最终出口装箱率)": 24,
        "CTNS": 5,
        "TOTAL CBM": 0.36,
        "N/W": 8.5,
        "G/W": 10.1,
        "L": 48,
        "W": 31,
        "H": 35,
    }
    base.update(overrides)
    return base


# —————————————————————————————————————
# Tests
# —————————————————————————————————————


class TestCacheHit:
    def test_repeated_get_returns_same_snapshot(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        s1 = cache.get_snapshot(path)
        s2 = cache.get_snapshot(path)
        assert s1 is s2  # same object — cache hit

    def test_second_get_is_fast(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        cache.get_snapshot(path)  # prime cache
        t0 = time.time()
        cache.get_snapshot(path)
        t1 = time.time()
        # cache hit should be near-instant (sub-1ms)
        assert (t1 - t0) < 0.1


class TestCacheInvalidation:
    def test_invalidate_forces_rebuild(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        s1 = cache.get_snapshot(path)
        cache.invalidate(path)
        s2 = cache.get_snapshot(path)
        assert s1 is not s2

    def test_file_modification_triggers_rebuild(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        s1 = cache.get_snapshot(path)

        # Modify file (must wait for mtime to change)
        time.sleep(0.01)
        make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"FINALQTY": 200})],
            name="base.xlsx",
        )
        s2 = cache.get_snapshot(path)
        assert s1 is not s2

    def test_invalidate_nonexistent_no_error(self, tmp_path):
        cache = WorkbookCacheManager(ttl_seconds=3600)
        cache.invalidate("/nonexistent/path.xlsx")  # should not raise


class TestCacheExpiry:
    def test_clear_expired_removes_old_entries(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        # Use very short TTL
        cache = WorkbookCacheManager(ttl_seconds=0)
        cache.get_snapshot(path)
        time.sleep(0.01)  # ensure TTL exceeded
        removed = cache.clear_expired()
        assert removed >= 1

    def test_clear_expired_does_not_remove_fresh(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        cache.get_snapshot(path)
        removed = cache.clear_expired()
        assert removed == 0  # just accessed, not expired


class TestConcurrency:
    def test_concurrent_same_file_builds_once(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)

        # We can't easily patch the internal build, so we verify by
        # checking that concurrent access returns the same object
        results = []

        def fetch():
            results.append(cache.get_snapshot(path))

        threads = [threading.Thread(target=fetch) for _ in range(5)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        # All should get the same snapshot object
        first = results[0]
        for r in results[1:]:
            assert r is first

    def test_different_files_independent(self, tmp_path):
        path1 = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
            name="base1.xlsx",
        )
        path2 = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"PO NO.": "4500099999"})],
            name="base2.xlsx",
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        s1 = cache.get_snapshot(path1)
        s2 = cache.get_snapshot(path2)
        assert s1 is not s2
        # Invalidate one doesn't affect the other
        cache.invalidate(path1)
        s3 = cache.get_snapshot(path2)
        assert s2 is s3  # still cached


class TestBuildLockCleanup:
    def test_invalidate_cleans_up_build_lock(self, tmp_path):
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        cache = WorkbookCacheManager(ttl_seconds=3600)
        cache.get_snapshot(path)
        assert path in cache._build_locks  # pyright: ignore[reportPrivateUsage]
        cache.invalidate(path)
        assert path not in cache._build_locks  # pyright: ignore[reportPrivateUsage]
