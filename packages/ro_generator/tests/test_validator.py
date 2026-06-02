"""Validator 测试：覆盖结构校验的正反例。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from ro_generator.validator import (
    CODE_HEADER_MISSING,
    CODE_SHEET_MISSING,
    validate_workbook_structure,
)
from ro_generator.workbook_reader import WorkbookReader


def make_workbook(tmp_path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    """构造合成 base，复用 reader 测试中的同样布局：表头第 4 行、数据第 5 行。"""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    path = tmp_path / "base.xlsx"
    wb.save(path)
    return path


def _blank_rows(n: int) -> list[list[Any]]:
    return [[None] for _ in range(n)]


# 完整 fixture：DATA BASE + PO record 都齐备
def standard_base(tmp_path: Path) -> Path:
    return make_workbook(
        tmp_path,
        sheets={
            "DATA BASE": [
                *_blank_rows(3),
                ["SAP", "Material Description", "Category"],
                ["21-44640", "CB2500.B2", 1],
            ],
            "PO record": [
                *_blank_rows(3),
                ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                ["4500030844", "10", "21-44640", 100],
            ],
        },
    )


class TestValidWorkbook:
    def test_no_messages_when_structure_complete(self, tmp_path: Path) -> None:
        path = standard_base(tmp_path)
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)
        assert messages == ()


class TestMissingSheets:
    def test_missing_data_base(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        sheet_msgs = [m for m in messages if m.code == CODE_SHEET_MISSING]
        assert len(sheet_msgs) == 1
        assert sheet_msgs[0].sheet == "DATA BASE"
        assert sheet_msgs[0].kind == "blocking_error"

    def test_missing_po_record(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["SAP", "Material Description", "Category"],
                    ["21-44640", "CB2500.B2", 1],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        sheet_msgs = [m for m in messages if m.code == CODE_SHEET_MISSING]
        assert len(sheet_msgs) == 1
        assert sheet_msgs[0].sheet == "PO record"

    def test_both_sheets_missing(self, tmp_path: Path) -> None:
        # 完全无关 sheet
        path = make_workbook(tmp_path, sheets={"unrelated": [[None]]})
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        sheet_msgs = [m for m in messages if m.code == CODE_SHEET_MISSING]
        assert {m.sheet for m in sheet_msgs} == {"DATA BASE", "PO record"}

    def test_missing_sheet_does_not_trigger_header_check(self, tmp_path: Path) -> None:
        """sheet 缺失时不应再为该 sheet 报表头缺失，避免噪声。"""
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["SAP", "Material Description", "Category"],
                    ["21-44640", "CB2500.B2", 1],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        # 只期待 1 条 SHEET_MISSING（PO record），不应有针对 PO record 的 HEADER_MISSING
        po_record_header_msgs = [
            m for m in messages if m.code == CODE_HEADER_MISSING and m.sheet == "PO record"
        ]
        assert po_record_header_msgs == []


class TestMissingHeaders:
    def test_missing_required_header_in_data_base(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["SAP", "Material Description"],  # 缺 Category
                    ["21-44640", "CB2500.B2"],
                ],
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        header_msgs = [m for m in messages if m.code == CODE_HEADER_MISSING]
        assert len(header_msgs) == 1
        msg = header_msgs[0]
        assert msg.sheet == "DATA BASE"
        assert msg.field == "Category"
        assert msg.kind == "blocking_error"

    def test_multiple_missing_headers_all_reported(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["SAP"],  # 缺 Material Description, Category
                    ["21-44640"],
                ],
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        header_msgs = [m for m in messages if m.code == CODE_HEADER_MISSING]
        missing_fields = {m.field for m in header_msgs}
        assert "Material Description" in missing_fields
        assert "Category" in missing_fields

    def test_completely_empty_sheet_reports_all_headers_missing(self, tmp_path: Path) -> None:
        """sheet 存在但行数不到 HEADER_ROW，应对每个必需表头报缺失。"""
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [],  # 完全空
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)
        # DATA BASE 的全部必需表头都该报
        data_base_msgs = [
            m for m in messages if m.code == CODE_HEADER_MISSING and m.sheet == "DATA BASE"
        ]
        assert {m.field for m in data_base_msgs} == {
            "SAP",
            "Material Description",
            "Category",
        }

    def test_normalized_header_match(self, tmp_path: Path) -> None:
        """带换行/多余空格的表头应能被识别为有效。"""
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["  SAP  ", "Material\nDescription", "Category"],
                    ["21-44640", "CB2500.B2", 1],
                ],
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)
        assert messages == ()


class TestMessageQuality:
    """阻断消息的元数据要够丰富，UI / CLI 才能定位问题。"""

    def test_header_message_includes_sheet_and_field(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [
                    *_blank_rows(3),
                    ["SAP", "Material Description"],
                    ["21-44640", "CB2500.B2"],
                ],
                "PO record": [
                    *_blank_rows(3),
                    ["PO NO.", "ITEM LINE#", "SAP Number", "FINALQTY"],
                    ["4500030844", "10", "21-44640", 100],
                ],
            },
        )
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)

        header_msgs = [m for m in messages if m.code == CODE_HEADER_MISSING]
        assert header_msgs
        msg = header_msgs[0]
        assert msg.sheet == "DATA BASE"
        assert msg.field == "Category"
        assert "Category" in msg.message
        assert "DATA BASE" in msg.message

    def test_sheet_message_includes_sheet(self, tmp_path: Path) -> None:
        path = make_workbook(tmp_path, sheets={"unrelated": [[None]]})
        with WorkbookReader(path) as reader:
            messages = validate_workbook_structure(reader)
        sheet_msgs = [m for m in messages if m.code == CODE_SHEET_MISSING]
        assert {m.sheet for m in sheet_msgs} == {"DATA BASE", "PO record"}
        for m in sheet_msgs:
            assert m.kind == "blocking_error"


@pytest.mark.parametrize(
    "_label",
    ["smoke"],
)
def test_returns_tuple(tmp_path: Path, _label: str) -> None:
    """返回类型必须是 tuple，避免下游意外修改消息列表。"""
    path = standard_base(tmp_path)
    with WorkbookReader(path) as reader:
        messages = validate_workbook_structure(reader)
    assert isinstance(messages, tuple)
