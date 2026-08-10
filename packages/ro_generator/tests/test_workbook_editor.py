"""workbook_editor 测试：字段写回。"""

from __future__ import annotations

from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, cast

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from ro_generator.profiles import create_pf_profile
from ro_generator.workbook_editor import edit_workbook_cell


def _make_temp_xlsx(
    headers: dict[int, str], data: dict[int, dict[int, object]] | None = None
) -> str:
    """创建临时 xlsx，表头在第 4 行，返回路径。"""
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "PO record"
    for col, name in headers.items():
        ws.cell(row=4, column=col, value=name)
    if data:
        for row, cols in data.items():
            for col, val in cols.items():
                ws.cell(row=row, column=col, value=cast(Any, val))
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    wb.close()
    return tmp_path


def test_edit_existing_field():
    path = _make_temp_xlsx(
        {1: "PO NO.", 2: "FINALQTY"},
        {5: {1: "4500099999", 2: 100}},
    )
    try:
        result = edit_workbook_cell(path, "PO record", row=5, field="FINALQTY", value=200)
        assert result.ok
        assert "已更新" in result.message
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_unknown_field():
    path = _make_temp_xlsx({1: "PO NO."})
    try:
        result = edit_workbook_cell(path, "PO record", row=5, field="nonexistent", value=1)
        assert not result.ok
        assert "找不到字段" in result.message
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_internal_field_key():
    path = _make_temp_xlsx(
        {1: "PO NO.", 2: "FINALQTY"},
        {5: {1: "4500099999", 2: 100}},
    )
    try:
        result = edit_workbook_cell(path, "PO record", row=5, field="quantity", value=200)
        assert result.ok

        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb["PO record"]
        assert ws.cell(row=5, column=2).value == 200
        wb.close()
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_unknown_sheet_returns_result():
    path = _make_temp_xlsx({1: "PO NO."})
    try:
        result = edit_workbook_cell(path, "UNKNOWN", row=5, field="po_no", value=1)
        assert not result.ok
        assert "不存在" in result.message
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_missing_sheet():
    path = _make_temp_xlsx({1: "PO NO."})
    try:
        result = edit_workbook_cell(path, "DATA BASE", row=5, field="po_no", value=1)
        assert not result.ok
        assert "不存在" in result.message
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_missing_file():
    result = edit_workbook_cell("/nonexistent/file.xlsx", "PO record", row=5, field="x", value=1)
    assert not result.ok
    assert "不存在" in result.message


def test_edit_preserves_other_data():
    path = _make_temp_xlsx(
        {1: "PO NO.", 2: "FINALQTY"},
        {5: {1: "4500099999", 2: 100}},
    )
    try:
        result = edit_workbook_cell(path, "PO record", row=5, field="PO NO.", value="4500088888")
        assert result.ok
        # Verify via openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(path)
        ws = wb["PO record"]
        assert ws.cell(row=5, column=1).value == "4500088888"
        assert ws.cell(row=5, column=2).value == 100
        wb.close()
    finally:
        Path(path).unlink(missing_ok=True)


def test_edit_profile_resolves_logical_sheet_to_physical_sheet(tmp_path: Path) -> None:
    wb = Workbook()
    data_base = wb.active
    assert data_base is not None
    data_base.title = "DATA BASE TEMPLATE"
    data_base.cell(row=2, column=1, value="SAP")
    po_record = wb.create_sheet("PO RECORD 26")
    po_record.cell(row=1, column=1, value="SAP Number")
    po_record.cell(row=2, column=1, value="OLD")
    customer_po = wb.create_sheet("new PO template")
    customer_po.cell(row=1, column=1, value="PO#")
    base_file = tmp_path / "pf-base.xlsx"
    wb.save(base_file)
    wb.close()

    result = edit_workbook_cell(
        str(base_file),
        "PO record",
        row=2,
        field="sap",
        value="NEW",
        profile=create_pf_profile(),
    )

    assert result.ok, result.message
    edited = load_workbook(base_file, data_only=True)
    assert edited["PO RECORD 26"]["A2"].value == "NEW"
    edited.close()
