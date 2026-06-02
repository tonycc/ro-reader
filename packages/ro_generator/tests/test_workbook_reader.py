"""Workbook reader 测试：用合成 xlsx fixture 覆盖各种结构场景。"""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from ro_generator.errors import WorkbookOpenError
from ro_generator.workbook_reader import (
    ROW_NUMBER_KEY,
    SheetData,
    WorkbookReader,
)

# ————————————————————————————————————————
# Fixture helpers
# ————————————————————————————————————————


def make_workbook(
    tmp_path: Path,
    *,
    name: str = "test.xlsx",
    sheets: dict[str, list[list[Any]]] | None = None,
) -> Path:
    """构造一个最小可用的 xlsx fixture。

    `sheets` 的值是行列表，row 1 = sheets[name][0]，让测试代码自然读写。
    """
    wb = Workbook()
    # 删除默认 Sheet
    default = wb.active
    if default is not None:
        wb.remove(default)
    if sheets is None:
        sheets = {"Sheet1": [["a"]]}
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    path = tmp_path / name
    wb.save(path)
    return path


def standard_po_record_fixture(tmp_path: Path) -> Path:
    """生成符合 base 文件布局（表头第 4 行、数据第 5 行）的合成 PO record。"""
    rows: list[list[Any]] = [
        ["", "", "", ""],  # row 1
        ["", "", "", ""],  # row 2
        ["", "", "", ""],  # row 3
        ["PO NO.", "SAP Number", "FINALQTY", "INV#"],  # row 4 = header
        ["4500030844", "21-44640", 100, "INV-001"],  # row 5
        ["4500030844", "21-44641", 200, "INV-001"],  # row 6
    ]
    return make_workbook(tmp_path, sheets={"PO record": rows})


# ————————————————————————————————————————
# Open / context manager
# ————————————————————————————————————————


class TestOpen:
    def test_missing_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(WorkbookOpenError, match="不存在"):
            WorkbookReader(tmp_path / "nope.xlsx")

    def test_directory_not_file_raises(self, tmp_path: Path) -> None:
        with pytest.raises(WorkbookOpenError, match="不是文件"):
            WorkbookReader(tmp_path)

    def test_invalid_xlsx_raises(self, tmp_path: Path) -> None:
        bogus = tmp_path / "bogus.xlsx"
        bogus.write_bytes(b"not really xlsx")
        with pytest.raises(WorkbookOpenError, match="无法打开"):
            WorkbookReader(bogus)

    def test_context_manager(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            assert reader.has_sheet("PO record")
        # 离开 context 后应已关闭，再次访问不抛出异常即可

    def test_str_path_accepted(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(str(path)) as reader:
            assert "PO record" in reader.sheet_names()


# ————————————————————————————————————————
# Sheet 元信息
# ————————————————————————————————————————


class TestSheetMeta:
    def test_sheet_names_returns_all(self, tmp_path: Path) -> None:
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": [["SAP"]],
                "PO record": [["PO NO."]],
                "extra": [["x"]],
            },
        )
        with WorkbookReader(path) as reader:
            assert set(reader.sheet_names()) == {"DATA BASE", "PO record", "extra"}

    def test_has_sheet(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            assert reader.has_sheet("PO record")
            assert not reader.has_sheet("nope")


# ————————————————————————————————————————
# read_sheet 主路径
# ————————————————————————————————————————


class TestReadSheet:
    def test_missing_sheet_raises(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with (
            WorkbookReader(path) as reader,
            pytest.raises(WorkbookOpenError, match="找不到 sheet"),
        ):
            reader.read_sheet("DATA BASE")

    def test_returns_sheet_data(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert isinstance(data, SheetData)
        assert data.sheet_name == "PO record"

    def test_headers_normalized_and_ordered(self, tmp_path: Path) -> None:
        rows: list[list[Any]] = [
            [None, None, None, None],
            [None, None, None, None],
            [None, None, None, None],
            ["PO NO.", "  SAP \nNumber  ", "FINALQTY", "INV#"],
            ["4500030844", "21-44640", 100, "INV-001"],
        ]
        path = make_workbook(tmp_path, sheets={"PO record": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        # 顺序保留 + 表头被规范化
        assert data.headers == ("PO NO.", "SAP Number", "FINALQTY", "INV#")

    def test_header_columns_use_1_based_indices(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert data.header_columns == {
            "PO NO.": 1,
            "SAP Number": 2,
            "FINALQTY": 3,
            "INV#": 4,
        }

    def test_rows_keyed_by_header(self, tmp_path: Path) -> None:
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert len(data.rows) == 2
        first = data.rows[0]
        assert first["PO NO."] == "4500030844"
        assert first["SAP Number"] == "21-44640"
        assert first["FINALQTY"] == 100
        assert first["INV#"] == "INV-001"

    def test_rows_carry_row_number(self, tmp_path: Path) -> None:
        """__row_number__ 是 1-based openpyxl 行号。

        合成数据：表头在 row 4，数据从 row 5 开始。
        """
        path = standard_po_record_fixture(tmp_path)
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert data.rows[0][ROW_NUMBER_KEY] == 5
        assert data.rows[1][ROW_NUMBER_KEY] == 6


# ————————————————————————————————————————
# 空白行 / 稀疏数据
# ————————————————————————————————————————


class TestBlankAndSparseRows:
    def test_blank_rows_skipped(self, tmp_path: Path) -> None:
        rows: list[list[Any]] = [
            *_blank_rows(3),
            ["PO NO.", "SAP Number", "FINALQTY", "INV#"],  # row 4
            ["4500030844", "21-44640", 100, "INV-001"],  # row 5
            [None, None, None, None],  # row 6 完全空白，跳过
            ["", "", "", ""],  # row 7 全空字符串，跳过
            ["4500030844", "21-44642", 80, "INV-001"],  # row 8
        ]
        path = make_workbook(tmp_path, sheets={"PO record": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert len(data.rows) == 2
        assert data.rows[0][ROW_NUMBER_KEY] == 5
        assert data.rows[1][ROW_NUMBER_KEY] == 8  # 行号反映原始位置

    def test_partially_filled_row_kept(self, tmp_path: Path) -> None:
        """只要有任意一列非空，就保留该行（让 validator 报缺字段）。"""
        rows: list[list[Any]] = [
            *_blank_rows(3),
            ["PO NO.", "SAP Number", "FINALQTY", "INV#"],
            ["4500030844", None, None, None],  # 只有 PO 列有值
        ]
        path = make_workbook(tmp_path, sheets={"PO record": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert len(data.rows) == 1
        assert data.rows[0]["PO NO."] == "4500030844"
        assert data.rows[0]["SAP Number"] is None
        assert data.rows[0]["FINALQTY"] is None

    def test_whitespace_only_string_treated_as_value(self, tmp_path: Path) -> None:
        """`" "` 不算空白：用户可能故意留占位，让 validator 在数值列上报错。"""
        rows: list[list[Any]] = [
            *_blank_rows(3),
            ["PO NO.", "SAP Number", "FINALQTY", "INV#"],
            [" ", None, None, None],
        ]
        path = make_workbook(tmp_path, sheets={"PO record": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert len(data.rows) == 1


# ————————————————————————————————————————
# 表头边界
# ————————————————————————————————————————


class TestHeaderEdgeCases:
    def test_empty_header_cells_skipped(self, tmp_path: Path) -> None:
        """表头行中间有空列时，那一列不参与索引（也不会出现在 row dict 中）。"""
        rows: list[list[Any]] = [
            *_blank_rows(3),
            ["PO NO.", None, "FINALQTY"],  # 中间空一列
            ["4500030844", "应被跳过", 100],
        ]
        path = make_workbook(tmp_path, sheets={"PO record": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("PO record")
        assert data.headers == ("PO NO.", "FINALQTY")
        # 第二列的数据没有进入 row dict（因为没有可用的 header key）
        row = data.rows[0]
        assert "应被跳过" not in row.values()
        assert row["PO NO."] == "4500030844"
        assert row["FINALQTY"] == 100

    def test_duplicate_headers_first_wins(self, tmp_path: Path) -> None:
        rows: list[list[Any]] = [
            *_blank_rows(3),
            ["SAP", "OTHER", "SAP"],  # SAP 重复
            ["A", "x", "B"],
        ]
        path = make_workbook(tmp_path, sheets={"Sheet": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("Sheet")
        # 重复表头只索引到首次出现的列
        assert data.header_columns == {"SAP": 1, "OTHER": 2}
        assert data.rows[0]["SAP"] == "A"  # 取列 1，不是列 3

    def test_header_row_below_data_returns_empty(self, tmp_path: Path) -> None:
        """sheet 行数不足 header_row 时，headers 为空，rows 为空。

        validator 据此报"必需表头缺失"。
        """
        rows: list[list[Any]] = [["x"], ["y"]]  # 仅 2 行
        path = make_workbook(tmp_path, sheets={"Sheet": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("Sheet")
        assert data.headers == ()
        assert data.header_columns == {}
        assert data.rows == ()

    def test_custom_header_row(self, tmp_path: Path) -> None:
        """允许传 header_row 参数，便于将来其他 sheet 布局接入。"""
        rows: list[list[Any]] = [
            ["PO NO.", "SAP Number"],  # row 1 = header
            ["4500030844", "21-44640"],  # row 2 = data
        ]
        path = make_workbook(tmp_path, sheets={"Sheet": rows})
        with WorkbookReader(path) as reader:
            data = reader.read_sheet("Sheet", header_row=1, first_data_row=2)
        assert data.headers == ("PO NO.", "SAP Number")
        assert len(data.rows) == 1
        assert data.rows[0]["PO NO."] == "4500030844"


# ————————————————————————————————————————
# 多 sheet 共存
# ————————————————————————————————————————


class TestMultipleSheets:
    def test_read_each_sheet_independently(self, tmp_path: Path) -> None:
        data_base_rows: list[list[Any]] = [
            *_blank_rows(3),
            ["SAP", "Material Description", "Category"],
            ["21-44640", "CB2500.B2", 1],
            ["21-44641", "CB3000.B2", 1],
        ]
        po_record_rows: list[list[Any]] = [
            *_blank_rows(3),
            ["PO NO.", "SAP Number", "FINALQTY"],
            ["4500030844", "21-44640", 100],
        ]
        path = make_workbook(
            tmp_path,
            sheets={
                "DATA BASE": data_base_rows,
                "PO record": po_record_rows,
            },
        )
        with WorkbookReader(path) as reader:
            db = reader.read_sheet("DATA BASE")
            po = reader.read_sheet("PO record")
        assert db.headers == ("SAP", "Material Description", "Category")
        assert len(db.rows) == 2
        assert po.headers == ("PO NO.", "SAP Number", "FINALQTY")
        assert len(po.rows) == 1


# ————————————————————————————————————————
# helpers
# ————————————————————————————————————————


def _blank_rows(n: int) -> Iterable[list[Any]]:
    for _ in range(n):
        yield [None]
