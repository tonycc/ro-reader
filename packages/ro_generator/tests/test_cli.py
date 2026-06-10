"""CLI 测试：参数解析、JSON 模式、退出码、错误路径。

覆盖：
- 退出码 0/1/2/3
- --json 模式 stdout 是合法 JSON
- 非 JSON 模式人类可读输出
- --input request.json 和命令行参数合并
- 参数错误捕获
"""

from __future__ import annotations

import io
import json
from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook
from ro_generator.cli import (
    EXIT_ERROR,
    EXIT_NEEDS_INPUT,
    EXIT_SUCCESS,
    EXIT_USAGE,
    main,
)

# ————————————————————————————————————————
# 复用 generator 测试的 fixture builder
# ————————————————————————————————————————

DATA_BASE_HEADER = [
    "SAP",
    "Material Description",
    "Category",
    "GS MODEL",
    "round value",
    "L",
    "W",
    "H",
]

PO_RECORD_HEADER = [
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "FINALQTY",
    "GS-SK/YM USD FOB",
    "EMAX-GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "SHIP QTY",
    "CTNS",
    "TOTAL CBM",
    "外箱(最终出口装箱率)",
    "N/W",
    "G/W",
]

CUSTOMER_PO_HEADER = ["Purchasing Document", "Material", "Order Quantity"]


def _write_sheet(ws: Any, headers: list[str], rows: list[dict[str, Any]],
                 header_row: int = 4, first_data_row: int = 5) -> None:
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            if value is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=value)


def _default_customer_po_rows(po_record_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in po_record_rows:
        po_no = row.get("PO NO.")
        material = row.get("SAP Number")
        if po_no is None or material is None:
            continue
        rows.append({
            "Purchasing Document": po_no,
            "Material": material,
            "Order Quantity": row.get("FINALQTY", 100),
        })
    return rows


def make_base_file(tmp_path: Path) -> Path:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(
        ws_db,
        DATA_BASE_HEADER,
        [
            {
                "SAP": "21-44640",
                "Material Description": "CB2500.B2",
                "Category": 1,
                "GS MODEL": "Q1",
                "round value": 24,
                "L": 60,
                "W": 40,
                "H": 30,
            }
        ],
    )
    ws_po = wb.create_sheet("PO record")
    _write_sheet(
        ws_po,
        PO_RECORD_HEADER,
        [
            {
                "PO NO.": "4500030844",
                "ITEM LINE#": "10",
                "SAP Number": "21-44640",
                "DESCRIPTION": "CB2500.B2",
                "FINALQTY": 100,
                "GS-SK/YM USD FOB": Decimal("28.0"),
                "EMAX-GS PTE FOB": Decimal("32.8"),
                "EMAX PTE": Decimal("38.0"),
                "INV#": "INV-001",
                "SHIP QTY": 100,
                "外箱(最终出口装箱率)": 24,
                "CTNS": 5,
                "TOTAL CBM": Decimal("0.36"),
                "N/W": Decimal("8.5"),
                "G/W": Decimal("10.1"),
            }
        ],
    )
    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(
        ws_cp,
        CUSTOMER_PO_HEADER,
        _default_customer_po_rows(
            [
                {
                    "PO NO.": "4500030844",
                    "SAP Number": "21-44640",
                    "FINALQTY": 100,
                }
            ]
        ),
        header_row=1,
        first_data_row=2,
    )
    path = tmp_path / "base.xlsx"
    wb.save(path)
    return path


def _run(argv: list[str]) -> tuple[int, str, str]:
    out = io.StringIO()
    err = io.StringIO()
    code = main(argv, stdout=out, stderr=err)
    return code, out.getvalue(), err.getvalue()


# ————————————————————————————————————————
# 成功路径
# ————————————————————————————————————————


class TestSuccess:
    def test_cli_args_success(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        code, stdout, _stderr = _run(
            [
                "--base",
                str(base),
                "--po",
                "4500030844",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--invoice-no",
                "INV-001",
                "--output-dir",
                str(out_dir),
            ]
        )
        assert code == EXIT_SUCCESS
        assert "装配成功" in stdout

    def test_json_mode_outputs_valid_json(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        code, stdout, _ = _run(
            [
                "--base",
                str(base),
                "--po",
                "4500030844",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--invoice-no",
                "INV-001",
                "--output-dir",
                str(out_dir),
                "--json",
            ]
        )
        assert code == EXIT_SUCCESS
        payload = json.loads(stdout)
        assert payload["status"] == "success"
        assert payload["output_file"]
        assert isinstance(payload["files"], list)
        assert isinstance(payload["source_index"], list)
        assert len(payload["source_index"]) > 0
        # 每个条目格式正确
        first = payload["source_index"][0]
        assert "doc_cell" in first and "source" in first
        assert {"sheet", "row", "field", "is_computed"} <= set(first["source"])

    def test_json_stdout_only_contains_json(self, tmp_path: Path) -> None:
        """JSON 模式下 stdout 必须只有 JSON，没有其他文本。"""
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        code, stdout, _ = _run(
            [
                "--base",
                str(base),
                "--po",
                "4500030844",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--invoice-no",
                "INV-001",
                "--output-dir",
                str(out_dir),
                "--json",
            ]
        )
        assert code == EXIT_SUCCESS
        # 整个 stdout 应该正好是一个 JSON 对象
        stripped = stdout.strip()
        assert stripped.startswith("{")
        assert stripped.endswith("}")
        json.loads(stripped)


# ————————————————————————————————————————
# --input request.json
# ————————————————————————————————————————


class TestInputFile:
    def test_request_json_drives_full_request(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        request_json = tmp_path / "request.json"
        request_json.write_text(
            json.dumps(
                {
                    "base_file": str(base),
                    "po_no": "4500030844",
                    "documents": ["INVOICE"],
                    "seller": "GS PTE",
                    "buyer": "EMAX PTE",
                    "invoice_no": "INV-001",
                    "output_dir": str(out_dir),
                }
            ),
            encoding="utf-8",
        )
        code, stdout, _ = _run(["--input", str(request_json), "--json"])
        assert code == EXIT_SUCCESS
        payload = json.loads(stdout)
        assert payload["status"] == "success"

    def test_cli_overrides_input_json(self, tmp_path: Path) -> None:
        """命令行参数应覆盖 JSON 中的同名字段。"""
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        request_json = tmp_path / "request.json"
        request_json.write_text(
            json.dumps(
                {
                    "base_file": str(base),
                    "po_no": "WILL_BE_OVERRIDDEN",
                    "documents": ["INVOICE"],
                    "seller": "GS PTE",
                    "buyer": "EMAX PTE",
                    "invoice_no": "INV-001",
                    "output_dir": str(out_dir),
                }
            ),
            encoding="utf-8",
        )
        code, _stdout, _stderr = _run(
            [
                "--input",
                str(request_json),
                "--po",
                "4500030844",  # 命令行覆盖
                "--json",
            ]
        )
        assert code == EXIT_SUCCESS

    def test_missing_input_file(self, tmp_path: Path) -> None:
        code, _stdout, stderr = _run(["--input", str(tmp_path / "nope.json")])
        assert code == EXIT_USAGE
        assert "不存在" in stderr

    def test_invalid_json_in_input(self, tmp_path: Path) -> None:
        f = tmp_path / "request.json"
        f.write_text("{not valid json", encoding="utf-8")
        code, _stdout, stderr = _run(["--input", str(f)])
        assert code == EXIT_USAGE
        assert "JSON" in stderr


# ————————————————————————————————————————
# 参数错误（退出码 2）
# ————————————————————————————————————————


class TestUsageErrors:
    def test_missing_base(self) -> None:
        code, _, stderr = _run(["--po", "4500030844", "--docs", "invoice"])
        assert code == EXIT_USAGE
        assert "base" in stderr

    def test_missing_po(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        code, _, stderr = _run(["--base", str(base), "--docs", "invoice"])
        assert code == EXIT_USAGE
        assert "po_no" in stderr or "po" in stderr.lower()

    def test_missing_docs(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        code, _, _stderr = _run(["--base", str(base), "--po", "4500030844"])
        assert code == EXIT_USAGE

    def test_invalid_doc_type(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        code, _, stderr = _run(["--base", str(base), "--po", "4500030844", "--docs", "RECEIPT"])
        assert code == EXIT_USAGE
        assert "RECEIPT" in stderr or "documents" in stderr

    def test_argparse_unknown_flag(self) -> None:
        code, _, _ = _run(["--unknown-flag", "x"])
        assert code == EXIT_USAGE


# ————————————————————————————————————————
# 阻断错误（退出码 1）
# ————————————————————————————————————————


class TestErrorExitCode:
    def test_unknown_po_returns_error(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        code, _stdout, stderr = _run(
            [
                "--base",
                str(base),
                "--po",
                "9999",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--invoice-no",
                "INV-001",
                "--output-dir",
                str(out_dir),
            ]
        )
        assert code == EXIT_ERROR
        assert "PO_NOT_FOUND" in stderr

    def test_error_in_json_mode(self, tmp_path: Path) -> None:
        base = make_base_file(tmp_path)
        out_dir = tmp_path / "out"
        code, stdout, _ = _run(
            [
                "--base",
                str(base),
                "--po",
                "9999",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--invoice-no",
                "INV-001",
                "--output-dir",
                str(out_dir),
                "--json",
            ]
        )
        assert code == EXIT_ERROR
        payload = json.loads(stdout)
        assert payload["status"] == "error"
        assert payload["errors"]
        assert any(e["code"] == "PO_NOT_FOUND" for e in payload["errors"])


# ————————————————————————————————————————
# needs_input（退出码 3）
# ————————————————————————————————————————


class TestNeedsInputExitCode:
    def test_multi_month_auto_selects_first(self, tmp_path: Path) -> None:
        """多月→自动选第一月，退出码 0（success）。"""
        wb = Workbook()
        default = wb.active
        if default is not None:
            wb.remove(default)
        ws_db = wb.create_sheet("DATA BASE")
        _write_sheet(
            ws_db,
            DATA_BASE_HEADER,
            [
                {
                    "SAP": "21-44640",
                    "Material Description": "CB2500.B2",
                    "Category": 1,
                    "GS MODEL": "Q1",
                    "round value": 24,
                }
            ],
        )
        ws_po = wb.create_sheet("PO record")
        _write_sheet(
            ws_po,
            PO_RECORD_HEADER,
            [
                {
                    "PO NO.": "4500030844",
                    "ITEM LINE#": "10",
                    "SAP Number": "21-44640",
                    "FINALQTY": 300,
                    "GS-SK/YM USD FOB": Decimal("28.0"),
                    "EMAX-GS PTE FOB": Decimal("32.8"),
                    "EMAX PTE": Decimal("38.0"),
                    "INV#": "INV-001",
                    "SHIP QTY": 300,
                    "CTNS": 12,
                    "TOTAL CBM": Decimal("0.86"),
                    "外箱(最终出口装箱率)": 24,
                }
            ],
        )
        ws_cp = wb.create_sheet("客户PO")
        _write_sheet(
            ws_cp,
            CUSTOMER_PO_HEADER,
            [
                {
                    "Purchasing Document": "4500030844",
                    "Material": "21-44640",
                    "Order Quantity": 300,
                }
            ],
            header_row=1,
            first_data_row=2,
        )
        base = tmp_path / "base.xlsx"
        wb.save(base)

        code, stdout, _ = _run(
            [
                "--base",
                str(base),
                "--po",
                "4500030844",
                "--docs",
                "invoice",
                "--seller",
                "GS PTE",
                "--buyer",
                "EMAX PTE",
                "--output-dir",
                str(tmp_path / "out"),
                "--json",
            ]
        )
        assert code == EXIT_SUCCESS
        payload = json.loads(stdout)
        assert payload["status"] == "success"


# ————————————————————————————————————————
# 退出码常量稳定性（不允许重排）
# ————————————————————————————————————————


@pytest.mark.parametrize(
    ("constant", "expected"),
    [
        (EXIT_SUCCESS, 0),
        (EXIT_ERROR, 1),
        (EXIT_USAGE, 2),
        (EXIT_NEEDS_INPUT, 3),
    ],
)
def test_exit_code_values_stable(constant: int, expected: int) -> None:
    assert constant == expected
