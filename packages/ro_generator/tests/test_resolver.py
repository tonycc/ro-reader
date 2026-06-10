"""Resolver 测试：覆盖 SAP join、价格选择、月度切片、公式回退、错误传播。"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from ro_generator.models import OrderLine
from ro_generator.resolver import (
    CODE_FORMULA_FALLBACK,
    CODE_NO_PRICES,
    CODE_PO_NOT_FOUND,
    CODE_QTY_INVALID,
    CODE_QTY_MISSING,
    CODE_SAP_MISSING,
    CODE_SAP_NOT_IN_DATA_BASE,
    PO_PRICE_COLUMNS,
    ResolveResult,
    resolve_po_lines,
)
from ro_generator.schema import (
    ENTITY_EMAX_PTE,
    ENTITY_GS_PTE,
    ENTITY_PF,
    ENTITY_YM,
)
from ro_generator.workbook_reader import WorkbookReader

# ————————————————————————————————————————
# Fixture builders
# ————————————————————————————————————————

# DATA BASE 默认表头：覆盖必需字段及物流字段
DATA_BASE_HEADER: list[Any] = [
    "SAP",
    "Material Description",
    "Category",
    "MOQ",
    "FOB LT",
    "GS MODEL",
    "GS-SK/YM COMBO FOB 2026",
    "EMAX-GS PTE COMBO FOB 2026",
    "EMAX PTE COMBO FOB 2026",
    "品牌",
    "RFID",
    "包装",
    "inner case value",
    "round value",
    "N/W",
    "G/W",
    "L",
    "W",
    "H",
    "CBM",
    "主件编号",
]

# PO record 默认表头：覆盖所需的所有列
PO_RECORD_HEADER: list[Any] = [
    "SHIP TO",
    "PO NO.",
    "ITEM LINE#",
    "SAP Number",
    "DESCRIPTION",
    "BRAND",
    "FINALQTY",
    "GS-SK/YM USD FOB",
    "EMAX-GS PTE FOB",
    "EMAX PTE",
    "INV#",
    "SHIP QTY",
    "CTNS",
    "N/W",
    "G/W",
    "TOTAL CBM",
    "外箱(最终出口装箱率)",
]

CUSTOMER_PO_HEADER: list[Any] = ["Purchasing Document", "Material", "ship to", "Order Quantity"]


def make_base_file(
    tmp_path: Path,
    *,
    data_base_rows: list[dict[str, Any]] | None = None,
    po_record_rows: list[dict[str, Any]] | None = None,
    customer_po_rows: list[dict[str, Any]] | None = None,
) -> Path:
    """构造合成 base 文件。每行用 dict 给定，未填字段补 None。"""
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    ws_db = wb.create_sheet("DATA BASE")
    _write_sheet(ws_db, DATA_BASE_HEADER, data_base_rows or [])

    ws_po = wb.create_sheet("PO record")
    _write_sheet(ws_po, PO_RECORD_HEADER, po_record_rows or [])

    ws_cp = wb.create_sheet("客户PO")
    _write_sheet(
        ws_cp,
        CUSTOMER_PO_HEADER,
        customer_po_rows if customer_po_rows is not None else _default_customer_po_rows(po_record_rows or []),
        header_row=1,
        first_data_row=2,
    )

    path = tmp_path / "base.xlsx"
    wb.save(path)
    return path


def _write_sheet(ws: Any, headers: list[Any], rows: list[dict[str, Any]],
                 header_row: int = 4, first_data_row: int = 5) -> None:
    """写入 sheet 表头和数据。"""
    for c_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c_idx, value=header)
    for r_offset, row in enumerate(rows):
        for c_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            if value is not None:
                ws.cell(row=first_data_row + r_offset, column=c_idx, value=value)


def _default_customer_po_rows(po_record_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in po_record_rows:
        po_no = row.get("PO NO.")
        material = row.get("SAP Number")
        if po_no is None or material is None:
            continue
        rows.append({
            "Purchasing Document": po_no,
            "Material": material,
            "ship to": "Customer PO Ship To",
            "Order Quantity": row.get("FINALQTY", 100),
        })
    return rows


# 一个常见 DATA BASE 产品定义：combo 类，包装齐全
COMBO_PRODUCT = {
    "SAP": "21-44640",
    "Material Description": "CB2500.B2",
    "Category": 1,
    "GS MODEL": "Q1",
    "GS-SK/YM COMBO FOB 2026": Decimal("28.0"),
    "EMAX-GS PTE COMBO FOB 2026": Decimal("32.8"),
    "EMAX PTE COMBO FOB 2026": Decimal("38.0"),
    "品牌": "Quantum",
    "round value": 24,
    "N/W": Decimal("12.5"),
    "G/W": Decimal("13.8"),
    "L": Decimal("60"),
    "W": Decimal("40"),
    "H": Decimal("30"),
}


def basic_po_row(**overrides: Any) -> dict[str, Any]:
    """生成一条具备所有必需字段的 PO record 行。"""
    base: dict[str, Any] = {
        "PO NO.": "4500030844",
        "ITEM LINE#": "10",
        "SAP Number": "21-44640",
        "DESCRIPTION": "CB2500.B2",
        "BRAND": "Quantum",
        "FINALQTY": 100,
        "GS-SK/YM USD FOB": Decimal("28.0"),
        "EMAX-GS PTE FOB": Decimal("32.8"),
        "EMAX PTE": Decimal("38.0"),
        "INV#": "INV-001",
        "SHIP QTY": 100,
        "外箱(最终出口装箱率)": 24,
    }
    base.update(overrides)
    return base


# ————————————————————————————————————————
# PO 整体：找不到、找到
# ————————————————————————————————————————


class TestPoLookup:
    def test_unknown_po_returns_blocking(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "9999999999")
        assert result.lines == ()
        assert len(result.messages) == 1
        msg = result.messages[0]
        assert msg.code == CODE_PO_NOT_FOUND
        assert msg.kind == "blocking_error"

    def test_known_po_resolves_one_line(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert isinstance(result, ResolveResult)
        assert len(result.lines) == 1
        line = result.lines[0]
        assert isinstance(line, OrderLine)
        assert line.po_no == "4500030844"
        assert line.sap == "21-44640"
        assert line.quantity == Decimal("100")
        assert line.category == 1
        assert line.invoice_no == "INV-001"
        assert line.invoice_no == "INV-001"

    def test_multiple_po_rows_resolved_in_order(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                COMBO_PRODUCT,
                {"SAP": "21-44641", "Material Description": "CB3000.B2", "Category": 1},
            ],
            po_record_rows=[
                basic_po_row(SAP_NUMBER_OVERRIDE="不被使用"),  # 顺序占位
                basic_po_row(**{"SAP Number": "21-44641", "FINALQTY": 200}),
            ],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert len(result.lines) == 2
        assert result.lines[0].sap == "21-44640"
        assert result.lines[1].sap == "21-44641"
        assert result.lines[1].quantity == Decimal("200")

    def test_ship_to_reads_from_customer_po_material_match(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SHIP TO": "PO Record Ship To"})],
            customer_po_rows=[{
                "Purchasing Document": "4500030844",
                "Material": "21-44640",
                "ship to": "Customer PO Ship To",
                "Order Quantity": 100,
            }],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines[0].ship_to == "Customer PO Ship To"

    def test_item_line_no_reads_from_customer_po_material(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"ITEM LINE#": "10"})],
            customer_po_rows=[{
                "Purchasing Document": "4500030844",
                "Material": "CP-MATERIAL-001",
                "ship to": "Customer PO Ship To",
                "Order Quantity": 100,
            }],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines[0].item_line_no == "CP-MATERIAL-001"

    def test_description_reads_from_data_base_material_description(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"DESCRIPTION": "PO Record Description"})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines[0].description == "CB2500.B2"

    def test_quantity_reads_from_customer_po_order_quantity(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=100)],
            customer_po_rows=[{
                "Purchasing Document": "4500030844",
                "Material": "21-44640",
                "ship to": "Customer PO Ship To",
                "Order Quantity": 240,
            }],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines[0].quantity == Decimal("240")

    def test_prices_read_from_data_base_price_matrix(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[{
                **COMBO_PRODUCT,
                "EMAX-GS PTE COMBO FOB 2026": Decimal("66.6"),
                "EMAX PTE COMBO FOB 2026": Decimal("88.8"),
            }],
            po_record_rows=[basic_po_row(**{
                "EMAX-GS PTE FOB": Decimal("32.8"),
                "EMAX PTE": Decimal("38.0"),
            })],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.prices[(ENTITY_GS_PTE, ENTITY_EMAX_PTE)] == Decimal("66.6")
        assert line.prices[(ENTITY_EMAX_PTE, ENTITY_PF)] == Decimal("88.8")

    def test_ship_to_falls_back_to_same_po_first_non_empty_customer_po_value(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SHIP TO": "PO Record Ship To"})],
            customer_po_rows=[
                {
                    "Purchasing Document": "4500030844",
                    "Material": "OTHER-MATERIAL",
                    "ship to": None,
                    "Order Quantity": None,
                },
                {
                    "Purchasing Document": "4500030844",
                    "Material": "DIFFERENT-MATERIAL",
                    "ship to": "Fallback Customer Ship To",
                    "Order Quantity": 100,
                },
            ],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines[0].ship_to == "Fallback Customer Ship To"


# ————————————————————————————————————————
# SAP join 的反例
# ————————————————————————————————————————


class TestSapJoin:
    def test_missing_sap_blocks_row(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SAP Number": None})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines == ()
        codes = [m.code for m in result.messages]
        assert CODE_SAP_MISSING in codes

    def test_sap_not_in_data_base_blocks_row(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"SAP Number": "99-99999"})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert result.lines == ()
        codes = [m.code for m in result.messages]
        assert CODE_SAP_NOT_IN_DATA_BASE in codes

    def test_data_base_row_without_sap_skipped(self, tmp_path: Path) -> None:
        """DATA BASE 中没有 SAP 的行应被忽略，不影响别的 SAP 解析。"""
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                {"SAP": None, "Category": 1},  # 应被跳过
                COMBO_PRODUCT,
            ],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert len(result.lines) == 1

    def test_data_base_row_without_category_skipped(self, tmp_path: Path) -> None:
        """category 缺失的产品视为不可用——下游若命中会拿到不完整数据。"""
        path = make_base_file(
            tmp_path,
            data_base_rows=[{"SAP": "21-44640", "Material Description": "x", "Category": None}],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        # SAP 在 PO record 中存在，但 DATA BASE 该行因 category 缺失被跳过 → 视为不存在
        codes = [m.code for m in result.messages]
        assert CODE_SAP_NOT_IN_DATA_BASE in codes


class TestQty:
    def test_missing_qty_blocks_row(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=100)],
            customer_po_rows=[{
                "Purchasing Document": "4500030844",
                "Material": "21-44640",
                "ship to": "Customer PO Ship To",
                "Order Quantity": None,
            }],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        codes = [m.code for m in result.messages]
        assert CODE_QTY_MISSING in codes
        assert result.lines == ()

    def test_invalid_qty_blocks_row(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=100)],
            customer_po_rows=[{
                "Purchasing Document": "4500030844",
                "Material": "21-44640",
                "ship to": "Customer PO Ship To",
                "Order Quantity": "abc",
            }],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        codes = [m.code for m in result.messages]
        assert CODE_QTY_INVALID in codes


# ————————————————————————————————————————
# 价格按链段读取
# ————————————————————————————————————————


class TestPrices:
    def test_all_three_segments_present(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.prices[(ENTITY_YM, ENTITY_GS_PTE)] == Decimal("28.0")
        assert line.prices[(ENTITY_GS_PTE, ENTITY_EMAX_PTE)] == Decimal("32.8")
        assert line.prices[(ENTITY_EMAX_PTE, ENTITY_PF)] == Decimal("38.0")

    def test_subtotals_match_quantity_times_price(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=100)],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.subtotals[(ENTITY_YM, ENTITY_GS_PTE)] == Decimal("2800.00")
        assert line.subtotals[(ENTITY_GS_PTE, ENTITY_EMAX_PTE)] == Decimal("3280.00")
        assert line.subtotals[(ENTITY_EMAX_PTE, ENTITY_PF)] == Decimal("3800.00")

    def test_only_one_segment_priced(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[{
                **COMBO_PRODUCT,
                "GS-SK/YM COMBO FOB 2026": None,
                "EMAX-GS PTE COMBO FOB 2026": Decimal("32.8"),
                "EMAX PTE COMBO FOB 2026": None,
            }],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert (ENTITY_GS_PTE, ENTITY_EMAX_PTE) in line.prices
        assert (ENTITY_YM, ENTITY_GS_PTE) not in line.prices
        assert (ENTITY_EMAX_PTE, ENTITY_PF) not in line.prices

    def test_no_prices_warns_but_still_resolves(self, tmp_path: Path) -> None:
        """缺价格→high warning，仍创建 OrderLine（prices={}）。"""
        path = make_base_file(
            tmp_path,
            data_base_rows=[{
                **COMBO_PRODUCT,
                "GS-SK/YM COMBO FOB 2026": None,
                "EMAX-GS PTE COMBO FOB 2026": None,
                "EMAX PTE COMBO FOB 2026": None,
            }],
            po_record_rows=[basic_po_row()],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        codes = [m.code for m in result.messages]
        assert CODE_NO_PRICES in codes
        assert all(m.kind == "warning" for m in result.messages)
        assert len(result.lines) == 1

    def test_price_columns_constant_covers_all_legal_segments(self) -> None:
        """常量表必须覆盖所有合法链段，否则会有段被静默跳过。"""
        from ro_generator.schema import LEGAL_CHAIN_SEGMENTS

        for seg in LEGAL_CHAIN_SEGMENTS:
            assert seg in PO_PRICE_COLUMNS


# ————————————————————————————————————————
# 月度出货
# ————————————————————————————————————————


class TestShipQty:
    def test_ship_qty_read(self, tmp_path: Path) -> None:
        row = basic_po_row(**{"SHIP QTY": 200})
        path = make_base_file(tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[row])
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.ship_qty == Decimal("200")

    def test_ship_qty_none_when_not_provided(self, tmp_path: Path) -> None:
        row = basic_po_row(**{"SHIP QTY": None})
        path = make_base_file(tmp_path, data_base_rows=[COMBO_PRODUCT], po_record_rows=[row])
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.ship_qty is None


# ————————————————————————————————————————
# 公式回退
# ————————————————————————————————————————


class TestFormulaFallback:
    def test_ctns_present_uses_value(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=240, CTNS=10)],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.carton_count == Decimal("10")
        # CTNS 不应触发 fallback warning（TOTAL CBM 仍可能 fallback，因为本测试未提供）
        ctns_fallback_msgs = [
            m for m in result.messages if m.code == CODE_FORMULA_FALLBACK and m.field == "CTNS"
        ]
        assert ctns_fallback_msgs == []

    def test_ctns_missing_falls_back_to_quantity_div_carton(self, tmp_path: Path) -> None:
        # 240 / 24 = 10
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=240, CTNS=None, **{"外箱": 24})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.carton_count == Decimal("10.0000")
        fallback_msgs = [m for m in result.messages if m.code == CODE_FORMULA_FALLBACK]
        # 至少有 CTNS 的回退 warning
        assert any(m.field == "CTNS" for m in fallback_msgs)
        for m in fallback_msgs:
            assert m.kind == "warning"
            assert m.severity == "high"

    def test_total_cbm_falls_back_to_dimensions(self, tmp_path: Path) -> None:
        # 60 * 40 * 30 / 1_000_000 = 0.072  ; 0.072 * 10 = 0.72
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(FINALQTY=240, CTNS=10, **{"TOTAL CBM": None, "外箱": 24})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.total_cbm == Decimal("0.7200")
        fallback_msgs = [
            m for m in result.messages if m.code == CODE_FORMULA_FALLBACK and m.field == "TOTAL CBM"
        ]
        assert len(fallback_msgs) == 1

    def test_zero_net_weight_and_gross_weight_are_preserved(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"N/W": Decimal("0"), "G/W": Decimal("0")})],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.net_weight == Decimal("0")
        assert line.gross_weight == Decimal("0")

    def test_total_cbm_preserves_source_decimal_places(self, tmp_path: Path) -> None:
        path = make_base_file(
            tmp_path,
            data_base_rows=[COMBO_PRODUCT],
            po_record_rows=[basic_po_row(**{"TOTAL CBM": 1.2})],
        )
        wb = load_workbook(path)
        ws = wb["PO record"]
        total_cbm_col = PO_RECORD_HEADER.index("TOTAL CBM") + 1
        ws.cell(row=5, column=total_cbm_col).number_format = "0.0000"
        wb.save(path)
        wb.close()

        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        line = result.lines[0]
        assert line.total_cbm == Decimal("1.2000")


# ————————————————————————————————————————
# 错误传播
# ————————————————————————————————————————


class TestPartialFailures:
    def test_one_good_one_bad_line(self, tmp_path: Path) -> None:
        """一行成功 + 一行失败：成功的进 lines，失败的进 messages。"""
        path = make_base_file(
            tmp_path,
            data_base_rows=[
                COMBO_PRODUCT,
                {"SAP": "21-44641", "Material Description": "CB3000.B2", "Category": 1},
            ],
            po_record_rows=[
                basic_po_row(),
                basic_po_row(**{"SAP Number": "99-99999"}),  # 不存在的 SAP
            ],
        )
        with WorkbookReader(path) as reader:
            result = resolve_po_lines(reader, "4500030844")
        assert len(result.lines) == 1
        codes = [m.code for m in result.messages]
        assert CODE_SAP_NOT_IN_DATA_BASE in codes


@pytest.mark.parametrize("_label", ["smoke"])
def test_resolve_result_is_immutable(tmp_path: Path, _label: str) -> None:
    path = make_base_file(
        tmp_path,
        data_base_rows=[COMBO_PRODUCT],
        po_record_rows=[basic_po_row()],
    )
    with WorkbookReader(path) as reader:
        result = resolve_po_lines(reader, "4500030844")
    assert isinstance(result.lines, tuple)
    assert isinstance(result.messages, tuple)
