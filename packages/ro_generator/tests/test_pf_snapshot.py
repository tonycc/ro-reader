"""PF Profile 的客户 PO 先行、数字月份表头和提醒集成测试。"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook, load_workbook
from ro_generator.document_model import CODE_LINE_NOT_PRICED
from ro_generator.generator import (
    export_invoice_group_from_snapshot,
    generate,
    preview_from_snapshot,
)
from ro_generator.invoice_inspection import inspect_invoice_group_from_snapshot
from ro_generator.models import DocumentRequest
from ro_generator.order_constraints import CODE_FULL_CARTON_NOT_MET, CODE_MOQ_NOT_MET
from ro_generator.profiles import GenerationContext, create_pf_profile
from ro_generator.resolver import resolve_po_rows
from ro_generator.workbook_snapshot import build_workbook_snapshot


def _make_pf_base(tmp_path: Path, *, category: str = "Single Rod") -> Path:
    workbook = Workbook()
    data_base = workbook.active
    assert data_base is not None
    data_base.title = "DATA BASE TEMPLATE"
    data_base.append([])
    data_base.append(
        [
            "SAP",
            "Material Description",
            "Category",
            "round value",
            "MOQ",
            "GS-SK/YM COMBO FOB 20260612-NEW PO",
            "EMAX-GS COMBO FOB 20260612-NEW PO 留下3%",
            "PF-EMAX COMBO DDP 2026 EFFECTIVE AS OF JUN/12/26-NEW PO",
        ]
    )
    data_base.append(["10001", "PF test item", category, 24, 100, 10, 11, 12])

    po_record = workbook.create_sheet("PO RECORD 26")
    po_record.append(
        ["PO NO.", "ITEM LINE#", "SAP Number", "INV#", 2601, 2602, "NEW DATE EX -FACTORY DATE"]
    )

    customer_po = workbook.create_sheet("new PO template")
    customer_po.append(
        [
            "PO Creation Date",
            "PO#",
            "PO-Item",
            "Material",
            "Material Description",
            "PO requested ex-fty date",
            "Order Quantity",
        ]
    )
    customer_po.append(["2026-08-01", "4500000001", 10, "10001", "PF test item", "2026-09-01", 90])

    path = tmp_path / "pf-base.xlsx"
    workbook.save(path)
    return path


def test_pf_snapshot_includes_customer_po_before_po_record_and_surfaces_constraints(
    tmp_path: Path,
) -> None:
    base_file = _make_pf_base(tmp_path)
    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)

    snapshot = build_workbook_snapshot(str(base_file), context=context)

    assert snapshot.profile_id == "pf"
    assert "2601" in snapshot.headers_po_record
    assert snapshot.po_summary[0].po_no == "4500000001"
    assert snapshot.po_summary[0].status == "ready"
    assert snapshot.po_summary[0].exportable_documents_by_seller["GS PTE"] == ("PI", "PO")

    result = resolve_po_rows(
        snapshot.po_rows_for_po("4500000001"),
        snapshot.product_index,
        po_no="4500000001",
        customer_po_rows=snapshot.customer_po_rows_for_po("4500000001"),
        profile=profile,
    )
    warning_codes = [message.code for message in result.messages if message.kind == "warning"]
    assert warning_codes == [CODE_MOQ_NOT_MET, CODE_FULL_CARTON_NOT_MET]


def test_pf_invoice_and_pl_from_separate_templates_export_as_zip(tmp_path: Path) -> None:
    base_file = _make_pf_base(tmp_path)
    workbook = load_workbook(base_file)
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "G26010101", 90, None])
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    summary = snapshot.invoice_summary[0]

    assert summary.seller_invoice_numbers == {
        "GS PTE": "G26010101",
        "EMAX PTE": "G26010101",
    }

    result = export_invoice_group_from_snapshot(
        snapshot,
        summary.invoice_group_key,
        seller="GS PTE",
        documents=("INVOICE", "PL"),
        output_dir=str(tmp_path / "output"),
        context=context,
    )

    assert result.status == "success", result.errors
    assert len(result.files) == 2
    assert result.output_file is not None
    assert result.output_file.endswith(".zip")
    with ZipFile(result.output_file) as archive:
        assert sorted(archive.namelist()) == sorted(result.files)


def test_pf_gs_pl_writes_sequential_carton_numbers(tmp_path: Path) -> None:
    base_file = _make_pf_base(tmp_path)
    workbook = load_workbook(base_file)
    data_base = workbook["DATA BASE TEMPLATE"]
    data_base["D3"] = 30
    data_base["I2"] = "N/W"
    data_base["J2"] = "G/W"
    data_base["K2"] = "L"
    data_base["L2"] = "W"
    data_base["M2"] = "H"
    data_base["I3"] = 1.1
    data_base["J3"] = 2.0
    data_base["K3"] = 119
    data_base["L3"] = 11.5
    data_base["M3"] = 17
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "G26010101", 90, None])
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PL",),
        seller="GS PTE",
        invoice_no="G26010101",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    assert preview_result.preview.lines[0]["carton_from"] == 1
    assert preview_result.preview.lines[0]["carton_to"] == 3

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["PL"]
    assert sheet["A21"].value == 1
    assert sheet["B21"].value == 3
    exported.close()


def test_pf_invoice_inspection_uses_context_for_monthly_ship_quantity(tmp_path: Path) -> None:
    base_file = _make_pf_base(tmp_path)
    workbook = load_workbook(base_file)
    workbook["PO RECORD 26"].append(["4500000001", 10, "10001", "G26010101", 90, None])
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)

    result = inspect_invoice_group_from_snapshot(
        snapshot,
        snapshot.invoice_summary[0].invoice_group_key,
        context=context,
    )

    assert [(row.invoice_no, row.ship_qty) for row in result.rows] == [("G26010101", Decimal("90"))]


def test_pf_gs_invoice_uses_invoice_source_rules_and_combo_breakdown(
    tmp_path: Path,
) -> None:
    base_file = _make_pf_base(tmp_path, category="Combo")
    workbook = load_workbook(base_file)
    data_base = workbook["DATA BASE TEMPLATE"]
    data_base["I2"] = "EMAX-GS ROD FOB 20260612-NEW PO 留下3%"
    data_base["J2"] = "EMAX-GS REEL FOB 20260612-NEW PO 留下3%"
    data_base["I3"] = 7
    data_base["J3"] = 4
    po_record = workbook["PO RECORD 26"]
    po_record["H1"] = "DESCRIPTION"
    po_record["I1"] = "ACTUAL EX FACTORY"
    po_record["J1"] = "ETD ON BOARD"
    po_record.append(
        [
            "4500000001",
            20,
            "10001",
            "G26010101",
            90,
            None,
            "2026-09-15",
            "PO Record Description",
            "2026-09-15",
            "2026-09-20",
        ]
    )
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("INVOICE",),
        seller="GS PTE",
        invoice_no="G26010101",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.resolved_values["etd_baseline"] == "2026-09-20"
    assert preview.resolved_values["number_of_cartons"] == ""
    assert preview.lines[0]["po_no"] == "4500000001"
    assert preview.lines[0]["item_line_no"] == "20"
    assert preview.lines[0]["sap"] == "10001"
    assert preview.lines[0]["description"] == "PO Record Description"
    assert preview.lines[0]["quantity"] == "90"
    assert preview.cost_breakdown == [
        {
            "po_no": "4500000001",
            "item_line_no": "20",
            "item_number": "10001",
            "description": "PO Record Description - RODS",
            "unit_price": "$7.00",
            "_index": 0,
            "_source_row": 2,
            "component": "RODS",
        },
        {
            "po_no": "4500000001",
            "item_line_no": "20",
            "item_number": "10001",
            "description": "PO Record Description - REELS",
            "unit_price": "$4.00",
            "_index": 1,
            "_source_row": 2,
            "component": "REELS",
        },
    ]
    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert source_by_field["invoice_no"]["field"] == "INV#"
    assert "按 INV# 筛选" in str(source_by_field["invoice_no"]["rule"])
    assert source_by_field["etd_baseline"]["field"] == "ETD ON BOARD"
    assert source_by_field["ex_factory_date"]["field"] == "ACTUAL EX FACTORY"
    assert "按 INV# 筛选" in str(source_by_field["ex_factory_date"]["rule"])
    assert source_by_field["line[0].po_no"]["field"] == "PO NO."
    assert source_by_field["line[0].item_line_no"]["field"] == "ITEM LINE#"
    assert source_by_field["line[0].quantity"]["field"] == "2601"

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["Sheet1"]
    assert str(sheet["G6"].value) == "2026-09-20"
    assert sheet["B20"].value == "4500000001"
    assert sheet["C20"].value == "20"
    assert sheet["D20"].value == "10001"
    assert sheet["E20"].value == "PO Record Description"
    assert sheet["H21"].value == 990
    assert sheet["H22"].value == 990
    assert sheet["G24"].value == "TINA"
    assert sheet["G25"].value
    assert sheet["B31"].value == "4500000001"
    assert sheet["E31"].value.endswith("- RODS")
    assert sheet["F31"].value == 7
    assert sheet["E32"].value.endswith("- REELS")
    assert sheet["F32"].value == 4
    for addr in ("E20", "A29", "B30", "B31", "E31", "E32"):
        font = sheet[addr].font
        assert font.name == "Arial", addr
        assert font.size == 9, addr
    exported.close()


def test_pf_gs_invoice_single_rod_uses_combo_fob_column(tmp_path: Path) -> None:
    """Single Rod 的 GS PTE 卖价在 COMBO FOB 列；ROD 列是 Combo 组件价，不能当主单价。"""

    base_file = _make_pf_base(tmp_path, category="Single Rod")
    workbook = load_workbook(base_file)
    data_base = workbook["DATA BASE TEMPLATE"]
    data_base["I2"] = "EMAX-GS ROD FOB 20260612-NEW PO 留下3%"
    data_base["I3"] = None
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 20, "10001", "G26010101", 90, None])
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("INVOICE",),
        seller="GS PTE",
        invoice_no="G26010101",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.lines[0]["unit_price"] == "$11.00"
    assert all(message.code != CODE_LINE_NOT_PRICED for message in preview_result.warnings)
    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert source_by_field["line[0].unit_price"]["field"] == (
        "EMAX-GS COMBO FOB 20260612-NEW PO 留下3%"
    )


def test_pf_invoice_combo_breakdown_prefers_po_record_category(tmp_path: Path) -> None:
    """Invoice 截图标注的 CATEGORY 来源是 PO RECORD，而不是只看 DATA BASE。"""

    base_file = _make_pf_base(tmp_path, category="Single Rod")
    workbook = load_workbook(base_file)
    data_base = workbook["DATA BASE TEMPLATE"]
    data_base["I2"] = "EMAX-GS ROD FOB 20260612-NEW PO 留下3%"
    data_base["J2"] = "EMAX-GS REEL FOB 20260612-NEW PO 留下3%"
    data_base["I3"] = 7
    data_base["J3"] = 4
    po_record = workbook["PO RECORD 26"]
    po_record["H1"] = "DESCRIPTION"
    po_record["I1"] = "ACTUAL EX FACTORY"
    po_record["J1"] = "ETD ON BOARD"
    po_record["K1"] = "CATEGORY"
    po_record.append(
        [
            "4500000001",
            20,
            "10001",
            "G26010101",
            90,
            None,
            "2026-09-15",
            "PO Record Description",
            "2026-09-15",
            "2026-09-20",
            "Combo",
        ]
    )
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("INVOICE",),
        seller="GS PTE",
        invoice_no="G26010101",
        output_dir=str(tmp_path / "output"),
    )

    result = preview_from_snapshot(snapshot, request, context=context)
    assert result.status == "success", result.errors
    assert result.preview is not None
    assert [row["component"] for row in result.preview.cost_breakdown] == ["RODS", "REELS"]


@pytest.mark.parametrize("category", ["Single Reel", "Single Rod"])
def test_pf_gs_pi_uses_screenshot_field_rules_for_preview_and_export(
    tmp_path: Path,
    category: str,
) -> None:
    base_file = _make_pf_base(tmp_path, category=category)
    workbook = load_workbook(base_file)
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "", None, None, "2026-09-15"])
    customer_po = workbook["new PO template"]
    customer_po["E2"] = "Customer PO description"
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PI",),
        seller="GS PTE",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.resolved_values["pi_no"] == "4500000001"
    assert preview.resolved_values["document_date"] == "2026-08-01"
    assert preview.resolved_values["ex_factory_date"] == "SEE BELOW"
    assert preview.resolved_values["manufacturer"] == ""
    assert preview.resolved_values["manufacturer_address"] == ""
    assert preview.resolved_values["manufacturer_address_2"] == ""
    assert preview.lines[0]["po_no"] == "4500000001"
    assert preview.lines[0]["item_number"] == "10001"
    assert preview.lines[0]["description"] == "Customer PO description"
    assert str(preview.lines[0]["confirmed_ex_factory_date"]) == "2026-09-15"
    assert preview.totals["signature_date"] == "2026-08-01"

    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert (
        source_by_field["line[0].po_no"]["sheet"],
        source_by_field["line[0].po_no"]["field"],
    ) == (
        "new PO template",
        "PO#",
    )
    assert (
        source_by_field["line[0].item_number"]["sheet"],
        source_by_field["line[0].item_number"]["field"],
    ) == ("new PO template", "Material")
    assert (
        source_by_field["line[0].description"]["sheet"],
        source_by_field["line[0].description"]["field"],
    ) == ("new PO template", "Material Description")
    assert (
        source_by_field["line[0].confirmed_ex_factory_date"]["sheet"],
        source_by_field["line[0].confirmed_ex_factory_date"]["field"],
    ) == ("PO RECORD 26", "NEW DATE EX -FACTORY DATE")
    assert "SK/YM确认的 PI 交期" in str(
        source_by_field["line[0].confirmed_ex_factory_date"]["rule"]
    )
    assert source_by_field["document_date"]["field"] == "PO Creation Date"
    assert source_by_field["totals.signature_date"]["field"] == "PO Creation Date"
    assert source_by_field["manufacturer"]["source_type"] == "manual_input"
    assert source_by_field["manufacturer_address"]["source_type"] == "manual_input"
    assert source_by_field["manufacturer_address_2"]["source_type"] == "manual_input"

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["Sheet1"]
    assert sheet["G7"].value == "2026-08-01"
    assert sheet["G8"].value == "SEE BELOW"
    assert sheet["G15"].value in (None, "")
    assert sheet["G16"].value in (None, "")
    assert sheet["G17"].value in (None, "")
    assert sheet["D20"].value == "10001"
    assert sheet["E20"].value == "Customer PO description"
    assert sheet["I20"].value.strftime("%Y-%m-%d") == "2026-09-15"
    assert sheet["F21"].value == "Sub-Total "
    assert sheet["G22"].value == "TOTAL EXCLUDING EXCISE TAX  (USD)"
    assert sheet["G24"].value == "TINA"
    assert sheet["G25"].value == "2026-08-01"
    exported.close()


def test_pf_emax_pi_uses_screenshot_field_rules_for_preview_and_export(
    tmp_path: Path,
) -> None:
    base_file = _make_pf_base(tmp_path, category="Single Rod")
    workbook = load_workbook(base_file)
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "", None, None, "2026-09-15"])
    customer_po = workbook["new PO template"]
    customer_po["E2"] = "Customer PO description"
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PI",),
        seller="EMAX PTE",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.resolved_values["pi_no"] == "4500000001"
    assert preview.resolved_values["document_date"] == "2026-08-01"
    assert preview.resolved_values["ex_factory_date"] == "SEE BELOW"
    assert preview.lines[0]["po_no"] == "4500000001"
    assert preview.lines[0]["item_line_no"] == "10"
    assert preview.lines[0]["sap"] == "10001"
    assert preview.lines[0]["description"] == "Customer PO description"
    assert preview.lines[0]["unit_price"] == "$12.00"
    assert preview.lines[0]["quantity"] == "90"
    assert str(preview.lines[0]["confirmed_ex_factory_date"]) == "2026-09-15"
    assert preview.totals["signature_date"] == "2026-08-01"

    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert source_by_field["document_date"]["field"] == "PO Creation Date"
    assert source_by_field["document_date"]["source_type"] == "base_field"
    assert source_by_field["line[0].description"]["field"] == "Material Description"
    assert source_by_field["line[0].description"]["sheet"] == "new PO template"
    assert source_by_field["line[0].unit_price"]["field"] == (
        "PF-EMAX COMBO DDP 2026 EFFECTIVE AS OF JUN/12/26-NEW PO"
    )
    assert source_by_field["line[0].confirmed_ex_factory_date"]["field"] == (
        "NEW DATE EX -FACTORY DATE"
    )
    assert source_by_field["totals.signature_date"]["field"] == "PO Creation Date"

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["PF Standard Invoice format"]
    assert sheet["G7"].value == "2026-08-01"
    assert sheet["G8"].value == "SEE BELOW"
    assert sheet["D20"].value == "10001"
    assert sheet["E20"].value == "Customer PO description"
    assert sheet["F20"].value == 12
    assert sheet["G20"].value == 90
    assert sheet["I20"].value.strftime("%Y-%m-%d") == "2026-09-15"
    assert sheet["G29"].value == "2026-08-01"
    exported.close()


def test_pf_emax_po_uses_screenshot_field_rules_for_preview_and_export(
    tmp_path: Path,
) -> None:
    base_file = _make_pf_base(tmp_path, category="Single Rod")
    workbook = load_workbook(base_file)
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "", None, None, "2026-09-15"])
    customer_po = workbook["new PO template"]
    customer_po["E2"] = "Customer PO description"
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PO",),
        seller="EMAX PTE",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.resolved_values["po_no"] == "4500000001"
    assert preview.resolved_values["document_date"] == "2026-08-01"
    assert preview.resolved_values["ex_factory_date"] == "SEE BELOW"
    assert preview.resolved_values["bill_to_line2"] == ""
    assert preview.resolved_values["ship_to_line2"] == ""
    assert preview.resolved_values["manufacturer_address"] == ""
    assert preview.lines[0]["po_no"] == "4500000001"
    assert preview.lines[0]["item_line_no"] == "10"
    assert preview.lines[0]["item_number"] == "10001"
    assert preview.lines[0]["description"] == "Customer PO description"
    assert preview.lines[0]["unit_price"] == "$12.00"
    assert preview.lines[0]["quantity"] == "90"
    assert str(preview.lines[0]["confirmed_ex_factory_date"]) == "2026-09-01"
    assert preview.totals["signature_date"] == "2026-08-01"

    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert source_by_field["po_no"]["field"] == "PO#"
    assert source_by_field["document_date"]["field"] == "PO Creation Date"
    assert source_by_field["ex_factory_date"]["source_type"] == "template_content"
    assert source_by_field["line[0].po_no"]["field"] == "PO#"
    assert source_by_field["line[0].item_line_no"]["field"] == "PO-Item"
    assert source_by_field["line[0].item_number"]["field"] == "Material"
    assert source_by_field["line[0].description"]["field"] == "Material Description"
    assert source_by_field["line[0].description"]["sheet"] == "new PO template"
    assert source_by_field["line[0].unit_price"]["field"] == (
        "PF-EMAX COMBO DDP 2026 EFFECTIVE AS OF JUN/12/26-NEW PO"
    )
    assert source_by_field["line[0].quantity"]["field"] == "Order Quantity"
    assert source_by_field["line[0].confirmed_ex_factory_date"]["field"] == (
        "PO requested ex-fty date"
    )
    assert source_by_field["totals.signature_date"]["field"] == "PO Creation Date"

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["PO"]
    assert sheet["B6"].value == "4500000001"
    assert sheet["B10"].value == "E MAX SPORT PTE. LTD."
    assert sheet["G7"].value == "2026-08-01"
    assert sheet["G8"].value == "SEE BELOW"
    assert sheet["B11"].value in (None, "")
    assert sheet["G11"].value in (None, "")
    assert sheet["G16"].value in (None, "")
    assert sheet["D20"].value == "10001"
    assert sheet["E20"].value == "Customer PO description"
    assert sheet["F20"].value == 12
    assert sheet["G20"].value == 90
    assert sheet["I20"].value.strftime("%Y-%m-%d") == "2026-09-01"
    assert sheet["F21"].value == "Sub-Total "
    assert sheet["G22"].value == "TOTAL EXCLUDING EXCISE TAX  (USD)"
    assert sheet["G24"].value == "Tina"
    assert sheet["G25"].value == "2026-08-01"
    exported.close()


def test_pf_emax_po_preserves_footer_when_detail_rows_are_inserted(tmp_path: Path) -> None:
    base_file = _make_pf_base(tmp_path, category="Single Rod")
    workbook = load_workbook(base_file)
    data_base = workbook["DATA BASE TEMPLATE"]
    data_base.append(["10002", "PF second item", "Single Rod", 24, 100, 10, 11, 12])
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "", None, None, "2026-09-15"])
    po_record.append(["4500000001", 20, "10002", "", None, None, "2026-09-16"])
    customer_po = workbook["new PO template"]
    customer_po.append(
        ["2026-08-01", "4500000001", 20, "10002", "PF second item", "2026-09-02", 90]
    )
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PO",),
        seller="EMAX PTE",
        output_dir=str(tmp_path / "output"),
    )

    result = generate(request, context=context)
    assert result.status == "success", result.errors
    assert result.output_file is not None
    exported = load_workbook(result.output_file, data_only=False)
    sheet = exported["PO"]
    assert sheet["D20"].value == "10001"
    assert sheet["D21"].value == "10002"
    assert sheet["F22"].value == "Sub-Total "
    assert sheet["G23"].value == "TOTAL EXCLUDING EXCISE TAX  (USD)"
    assert sheet["G25"].value == "Tina"
    assert sheet["G26"].value == "2026-08-01"
    exported.close()


@pytest.mark.parametrize(
    ("category", "expected_manufacturer", "expected_address", "expected_address_2"),
    [
        (
            "Single Reel",
            "GUANGDONG GLOBALSINO OUTDOOR SPORTS EQUIPMENT LIMITED",
            "NO.40 BAIJIA ST12 NO. 93 GRAPE BEACH ROAD",
            "DEVELOPMENT ZONE QINGYUAN GUANGDONG CHINA",
        ),
        (
            "Single Rod",
            "WEIHAI E-MAX SPORT APPARATUS CO.LTD",
            "NO. 93 GRAPE BEACH ROAD",
            "SUNJIATUAN TOWN, WEIHAI, SHANGDONG, CHINA",
        ),
    ],
)
def test_pf_gs_po_uses_screenshot_field_rules_for_preview_and_export(
    tmp_path: Path,
    category: str,
    expected_manufacturer: str,
    expected_address: str,
    expected_address_2: str,
) -> None:
    base_file = _make_pf_base(tmp_path, category=category)
    workbook = load_workbook(base_file)
    po_record = workbook["PO RECORD 26"]
    po_record.append(["4500000001", 10, "10001", "", None, None, "2026-09-15"])
    customer_po = workbook["new PO template"]
    customer_po["E2"] = "Customer PO description"
    workbook.save(base_file)
    workbook.close()

    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PO",),
        seller="GS PTE",
        output_dir=str(tmp_path / "output"),
    )

    preview_result = preview_from_snapshot(snapshot, request, context=context)
    assert preview_result.status == "success", preview_result.errors
    assert preview_result.preview is not None
    preview = preview_result.preview
    assert preview.title == "PURCHASE ORDER"
    assert preview.resolved_values["po_no"] == "4500000001"
    assert preview.resolved_values["incoterm"] == "FOB Qingdao"
    assert preview.resolved_values["payment_terms"] == "Net 75 days"
    assert preview.resolved_values["document_date"] == "2026-08-01"
    assert preview.resolved_values["ex_factory_date"] == "SEE BELOW"
    assert preview.resolved_values["manufacturer"] == expected_manufacturer
    assert preview.resolved_values["manufacturer_address"] == expected_address
    assert preview.resolved_values["manufacturer_address_2"] == expected_address_2
    assert preview.lines[0]["po_no"] == "4500000001"
    assert preview.lines[0]["item_number"] == "10001"
    assert preview.lines[0]["description"] == "Customer PO description"
    assert preview.lines[0]["unit_price"] == "$10.00"
    assert preview.lines[0]["quantity"] == "90"
    assert str(preview.lines[0]["confirmed_ex_factory_date"]) == "2026-09-01"
    assert preview.totals["signature_date"] == "2026-08-01"

    source_by_field = {entry["preview_field"]: entry for entry in preview.source_entries}
    assert source_by_field["po_no"]["field"] == "PO#"
    assert source_by_field["document_date"]["field"] == "PO Creation Date"
    assert source_by_field["manufacturer"]["field"] == "Category"
    assert source_by_field["line[0].po_no"]["field"] == "PO#"
    assert source_by_field["line[0].item_line_no"]["field"] == "PO-Item"
    assert source_by_field["line[0].item_number"]["field"] == "Material"
    assert source_by_field["line[0].description"]["field"] == "Material Description"
    assert source_by_field["line[0].quantity"]["field"] == "Order Quantity"
    assert source_by_field["line[0].confirmed_ex_factory_date"]["field"] == (
        "PO requested ex-fty date"
    )
    assert source_by_field["totals.signature_date"]["field"] == "PO Creation Date"

    export_result = generate(request, context=context)
    assert export_result.status == "success", export_result.errors
    assert export_result.output_file is not None
    exported = load_workbook(export_result.output_file, data_only=False)
    sheet = exported["Sheet1"]
    assert sheet["B6"].value == "4500000001"
    assert sheet["B7"].value == "FOB Qingdao"
    assert sheet["G7"].value == "2026-08-01"
    assert sheet["G8"].value == "SEE BELOW"
    assert sheet["G15"].value == expected_manufacturer
    assert sheet["G16"].value == expected_address
    assert sheet["G17"].value == expected_address_2
    assert sheet["D20"].value == "10001"
    assert sheet["E20"].value == "Customer PO description"
    assert sheet["F20"].value == 10
    assert sheet["G20"].value == 90
    assert sheet["I20"].value.strftime("%Y-%m-%d") == "2026-09-01"
    assert sheet["G24"].value == "TINA"
    assert sheet["G25"].value == "2026-08-01"
    exported.close()


@pytest.mark.parametrize(
    ("seller", "category", "expected_company"),
    [
        (
            "SK",
            "Single Reel",
            "GUANGDONG GLOBALSINO OUTDOOR SPORTS EQUIPMENT LIMITED",
        ),
        ("YM", "Single Rod", "WEIHAI E-MAX SPORT APPARATUS CO.,LTD"),
    ],
)
def test_pf_sk_ym_pi_preview_uses_template_document_header(
    tmp_path: Path,
    seller: str,
    category: str,
    expected_company: str,
) -> None:
    base_file = _make_pf_base(tmp_path, category=category)
    profile = create_pf_profile()
    context = GenerationContext(profile=profile, base_file=base_file)
    snapshot = build_workbook_snapshot(str(base_file), context=context)
    request = DocumentRequest(
        base_file=str(base_file),
        po_no="4500000001",
        documents=("PI",),
        seller=seller,
    )

    result = preview_from_snapshot(snapshot, request, context=context)

    assert result.status == "success", result.errors
    assert result.preview is not None
    preview = result.preview
    assert preview.title == "PROFORMA INVOICE"
    assert preview.seller_info[0] == expected_company
    assert preview.header_labels["pi_no"] == "PI Number:"
    assert preview.header_labels["etd_baseline"] == "ETD (Baseline Date for FOB Term):"
    assert preview.layout["top"] == {
        "left": [],
        "center": ["seller_info", "title"],
        "right": [],
    }
    assert preview.resolved_values["bill_to_line2"] == (
        "10 KAKI BUKIT ROAD 2, #01-37, FIRST EAST CENTRE"
    )
    if seller == "YM":
        assert preview.resolved_values["manufacturer_address"] == "NO.25 TONGYI NORTH ROAD,"
        assert preview.resolved_values["manufacturer_address_2"] == (
            "HUANCUI DISTRICT, WEIHAI, SHANDONG, CHINA."
        )
    po_no_entry = next(
        entry for entry in preview.source_entries if entry["preview_field"] == "line[0].po_no"
    )
    assert (po_no_entry["sheet"], po_no_entry["field"]) == ("new PO template", "PO#")
