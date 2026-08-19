"""schema_inspect 结构问题探测的单元测试。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from ro_generator.base_schema import base_schema
from ro_generator.schema_inspect import inspect_schema, sheet_header_candidates
from ro_generator.workbook_reader import WorkbookReader


def _make_workbook(tmp_path: Path, sheets: dict[str, list[list[Any]]]) -> Path:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    path = tmp_path / "base.xlsx"
    wb.save(path)
    return path


def _headers_row(*headers: str) -> list[list[Any]]:
    # 表头在第 4 行，前 3 行留空
    return [[None], [None], [None], list(headers)]


# RO DATA BASE 的全部价格列表头（data_base_price_columns 去重后的值）。
_RO_PRICE_HEADERS = (
    "GS-SK/YM COMBO FOB 2026",
    "GS-SK/YM YM ROD FOB 2026",
    "GS-SK/YM SK REEL FOB 2026",
    "EMAX-GS PTE COMBO FOB 2026",
    "EMAX-GS PTE ROD FOB 2026",
    "EMAX-GS PTE REEL FOB 2026",
    "EMAX PTE COMBO FOB 2026",
    "EMAX PTE ROD FOB 2026",
    "EMAX PTE REEL FOB 2026",
)


def test_sheet_header_candidates_returns_letters(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path,
        {
            "DATA BASE": _headers_row("SAP", "Material Description"),
            "PO record": _headers_row("PO NO."),
            "客户PO": [["Purchasing Document"]],
        },
    )
    schema = base_schema()
    reader = WorkbookReader(str(path), schema=schema)
    headers, letters = sheet_header_candidates(reader, schema, "DATA BASE", "DATA BASE")
    assert "SAP" in headers
    assert letters["SAP"] == "A"


def test_no_issues_when_structure_matches(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path,
        {
            "DATA BASE": _headers_row(
                "SAP", "Material Description", "Category", *_RO_PRICE_HEADERS
            ),
            "PO record": _headers_row("PO NO.", "ITEM LINE#", "SAP Number"),
            "客户PO": [["Purchasing Document", "Material", "Order Quantity"]],
        },
    )
    reader = WorkbookReader(str(path), schema=base_schema())
    inspection = inspect_schema(reader, base_schema())
    assert not inspection.has_issues()
    assert inspection.field_issues == ()
    assert inspection.sheet_issues == ()
    assert inspection.price_issues == ()


def test_missing_header_produces_field_issue_with_candidates(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path,
        {
            # SAP 被改名为 SAP Code
            "DATA BASE": _headers_row(
                "SAP Code", "Material Description", "Category", *_RO_PRICE_HEADERS
            ),
            "PO record": _headers_row("PO NO.", "ITEM LINE#", "SAP Number"),
            "客户PO": [["Purchasing Document", "Material", "Order Quantity"]],
        },
    )
    reader = WorkbookReader(str(path), schema=base_schema())
    inspection = inspect_schema(reader, base_schema())
    assert inspection.has_issues()
    assert len(inspection.field_issues) == 1
    issue = inspection.field_issues[0]
    assert issue.logical_sheet == "DATA BASE"
    assert issue.internal_key == "sap"
    assert issue.expected_header == "SAP"
    assert "SAP Code" in issue.available_headers
    assert issue.column_letters["SAP Code"] == "A"


def test_missing_sheet_produces_sheet_issue(tmp_path: Path) -> None:
    path = _make_workbook(
        tmp_path,
        {
            "PO record": _headers_row("PO NO.", "ITEM LINE#", "SAP Number"),
            "客户PO": [["Purchasing Document", "Material", "Order Quantity"]],
        },
    )
    reader = WorkbookReader(str(path), schema=base_schema())
    inspection = inspect_schema(reader, base_schema())
    assert inspection.has_issues()
    assert any(item.logical_sheet == "DATA BASE" for item in inspection.sheet_issues)
    # DATA BASE 缺失时不应再为它产出逐列问题
    assert all(item.logical_sheet != "DATA BASE" for item in inspection.field_issues)


def test_missing_price_column_produces_price_issue(tmp_path: Path) -> None:
    # EMAX PTE COMBO FOB 2026 被改名为 EMAX PTE COMBO FOB 2027
    renamed = tuple(
        "EMAX PTE COMBO FOB 2027" if h == "EMAX PTE COMBO FOB 2026" else h
        for h in _RO_PRICE_HEADERS
    )
    path = _make_workbook(
        tmp_path,
        {
            "DATA BASE": _headers_row("SAP", "Material Description", "Category", *renamed),
            "PO record": _headers_row("PO NO.", "ITEM LINE#", "SAP Number"),
            "客户PO": [["Purchasing Document", "Material", "Order Quantity"]],
        },
    )
    reader = WorkbookReader(str(path), schema=base_schema())
    inspection = inspect_schema(reader, base_schema())
    assert inspection.has_issues()
    assert inspection.field_issues == ()
    # 该列只被 EMAX PTE/combo 一个价格键引用
    assert len(inspection.price_issues) == 1
    issue = inspection.price_issues[0]
    assert issue.logical_sheet == "DATA BASE"
    assert issue.internal_key == "EMAX PTE/combo"
    assert issue.expected_header == "EMAX PTE COMBO FOB 2026"
    assert "EMAX PTE COMBO FOB 2027" in issue.available_headers
    assert issue.column_letters["EMAX PTE COMBO FOB 2027"] == "J"
